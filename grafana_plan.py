from __future__ import annotations

import json
import re
from collections import defaultdict
from datetime import datetime
from typing import Any, Callable, Dict, Iterable, List, Sequence, Tuple

from openpyxl import Workbook, load_workbook  # type: ignore

import config
from api_clients import GrafanaAPI
from common import autosize_columns, join_sorted, normalize_values
from grafana_audit import _field_kind, _iter_strings
from mapping_plan import is_selected


GRAFANA_PLAN_HEADERS: List[str] = [
    "apply",
    "grafana_org_id",
    "dashboard_uid",
    "dashboard_title",
    "folder_title",
    "dashboard_url",
    "panel_url",
    "panel_id",
    "panel_title",
    "panel_type",
    "location_kind",
    "variable_name",
    "variable_type",
    "datasource_names",
    "field_path",
    "json_path",
    "field_kind",
    "old_group",
    "new_group",
    "replace_count",
    "source_value",
    "planned_value",
    "change_kind",
    "change_mode",
    "manual_required",
    "status",
    "comment",
]

GRAFANA_APPLY_RESULT_HEADERS: List[str] = [
    "grafana_org_id",
    "dashboard_uid",
    "dashboard_title",
    "dashboard_url",
    "panel_url",
    "panel_id",
    "panel_title",
    "panel_type",
    "location_kind",
    "variable_name",
    "field_path",
    "json_path",
    "field_kind",
    "old_group",
    "new_group",
    "before_value",
    "after_value",
    "status",
    "applied",
    "message",
]


def _log(log: Callable[[str], None] | None, message: str) -> None:
    if log is not None:
        log(message)


def load_grafana_org_report(path: str) -> Dict[str, Any]:
    with open(path, "r", encoding="utf-8") as handle:
        return json.load(handle)


def load_impact_plan(path: str) -> Dict[str, Any]:
    with open(path, "r", encoding="utf-8") as handle:
        return json.load(handle)


def save_grafana_plan_json(data: Dict[str, Any], path: str) -> None:
    with open(path, "w", encoding="utf-8") as handle:
        json.dump(data, handle, ensure_ascii=False, indent=2)


def _append_rows(ws, rows: Sequence[Dict[str, Any]], headers: Sequence[str]) -> None:
    ws.append(list(headers))
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    autosize_columns(ws)


def write_grafana_plan_xlsx(data: Dict[str, Any], path: str) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    summary_ws = wb.create_sheet("SUMMARY")
    summary_ws.append(["key", "value"])
    for key, value in data.get("summary", {}).items():
        if isinstance(value, list):
            rendered = ", ".join(str(item) for item in value)
        else:
            rendered = value
        summary_ws.append([key, rendered])
    autosize_columns(summary_ws)

    plan_ws = wb.create_sheet("PLAN")
    _append_rows(plan_ws, data.get("plan_rows") or [], GRAFANA_PLAN_HEADERS)

    missing_ws = wb.create_sheet("REVIEW_ROWS")
    _append_rows(
        missing_ws,
        data.get("review_rows") or data.get("missing_variables") or [],
        [
            "grafana_org_id",
            "dashboard_uid",
            "dashboard_title",
            "dashboard_url",
            "variable_name",
            "variable_type",
            "location_kind",
            "field_kind",
            "json_path",
            "old_group",
            "new_group",
            "status",
            "message",
        ],
    )

    wb.save(path)


def write_grafana_apply_xlsx(data: Dict[str, Any], path: str) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    summary_ws = wb.create_sheet("SUMMARY")
    summary_ws.append(["key", "value"])
    for key, value in data.get("summary", {}).items():
        if isinstance(value, list):
            rendered = ", ".join(str(item) for item in value)
        else:
            rendered = value
        summary_ws.append([key, rendered])
    autosize_columns(summary_ws)

    results_ws = wb.create_sheet("RESULTS")
    _append_rows(results_ws, data.get("results") or [], GRAFANA_APPLY_RESULT_HEADERS)

    dashboards_ws = wb.create_sheet("DASHBOARDS")
    _append_rows(
        dashboards_ws,
        data.get("dashboards") or [],
        [
            "grafana_org_id",
            "dashboard_uid",
            "dashboard_title",
            "dashboard_url",
            "changes_requested",
            "changes_applied",
            "status",
            "message",
        ],
    )

    wb.save(path)


def load_grafana_plan_rows(path: str) -> List[Dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=False)
    ws = wb["PLAN"] if "PLAN" in wb.sheetnames else wb[wb.sheetnames[0]]
    raw_rows = list(ws.iter_rows(values_only=True))
    if not raw_rows:
        return []

    headers = [str(item or "").strip() for item in raw_rows[0]]
    rows: List[Dict[str, Any]] = []
    for raw in raw_rows[1:]:
        if not raw:
            continue
        row = {headers[index]: raw[index] for index in range(min(len(headers), len(raw))) if headers[index]}
        if any(str(value or "").strip() for value in row.values()):
            rows.append(row)
    return rows


def get_selected_grafana_changes(rows: Sequence[Dict[str, Any]]) -> List[Dict[str, str]]:
    selected: List[Dict[str, str]] = []
    for row in rows:
        if not is_selected(row.get("apply")):
            continue
        grafana_org_id = str(row.get("grafana_org_id") or "").strip()
        dashboard_uid = str(row.get("dashboard_uid") or "").strip()
        location_kind = str(row.get("location_kind") or "").strip()
        variable_name = str(row.get("variable_name") or "").strip()
        field_path = str(row.get("field_path") or "").strip()
        old_group = str(row.get("old_group") or "").strip()
        new_group = str(row.get("new_group") or "").strip()
        if not all([grafana_org_id, dashboard_uid, location_kind, field_path, old_group, new_group]):
            raise RuntimeError("Selected Grafana plan row must contain org_id, dashboard_uid, location_kind, field_path, old_group, new_group.")
        selected.append(
            {
                "grafana_org_id": grafana_org_id,
                "dashboard_uid": dashboard_uid,
                "dashboard_title": str(row.get("dashboard_title") or "").strip(),
                "dashboard_url": str(row.get("dashboard_url") or "").strip(),
                "panel_url": str(row.get("panel_url") or "").strip(),
                "panel_id": str(row.get("panel_id") or "").strip(),
                "panel_title": str(row.get("panel_title") or "").strip(),
                "panel_type": str(row.get("panel_type") or "").strip(),
                "location_kind": location_kind,
                "variable_name": variable_name,
                "field_path": field_path,
                "json_path": str(row.get("json_path") or "").strip(),
                "field_kind": str(row.get("field_kind") or "").strip(),
                "old_group": old_group,
                "new_group": new_group,
                "source_value": str(row.get("source_value") or ""),
                "planned_value": str(row.get("planned_value") or ""),
                "change_mode": str(row.get("change_mode") or "").strip(),
            }
        )
    if not selected:
        raise RuntimeError("No rows with apply=yes found in Grafana plan.")
    return selected


def _find_variable(dashboard: Dict[str, Any], variable_name: str) -> Tuple[int, Dict[str, Any]] | Tuple[None, None]:
    templating = (dashboard.get("templating") or {}).get("list") or []
    for index, variable in enumerate(templating):
        if str(variable.get("name") or "").strip() == variable_name:
            return index, variable
    return None, None


def _shorten(text: str, limit: int = 500) -> str:
    value = str(text or "")
    if len(value) <= limit:
        return value
    return value[: limit - 3] + "..."


def _is_manual_change_mode(field_kind: str) -> bool:
    return str(field_kind or "").strip() in {"regex", "query", "definition", "expression", "sql"}


def _resolve_change_mode(row: Dict[str, Any]) -> str:
    explicit = str(row.get("change_mode") or "").strip()
    if explicit:
        return explicit
    if _is_manual_change_mode(str(row.get("field_kind") or "")):
        return "manual_regex"
    return "exact"


_PATH_TOKEN_RX = re.compile(r"([^.\[]+)|\[(\d+)\]")


def _parse_field_path(path: str) -> List[Any]:
    tokens: List[Any] = []
    for part in str(path or "").split("."):
        if not part:
            continue
        for match in _PATH_TOKEN_RX.finditer(part):
            key, index = match.groups()
            if key is not None:
                tokens.append(key)
            elif index is not None:
                tokens.append(int(index))
    return tokens


def _get_path_value(node: Any, path: str) -> Any:
    current = node
    for token in _parse_field_path(path):
        if isinstance(token, int):
            current = current[token]
        else:
            current = current[token]
    return current


def _set_path_value(node: Any, path: str, value: Any) -> None:
    tokens = _parse_field_path(path)
    if not tokens:
        raise RuntimeError(f"Invalid field path: {path}")
    current = node
    for token in tokens[:-1]:
        if isinstance(token, int):
            current = current[token]
        else:
            current = current[token]
    tail = tokens[-1]
    if isinstance(tail, int):
        current[tail] = value
    else:
        current[tail] = value


def _group_mappings_by_old(selected_mappings: Sequence[Dict[str, str]]) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    seen: set[tuple[str, str]] = set()
    for row in selected_mappings:
        old_group = str(row.get("old_group") or "").strip()
        new_group = str(row.get("new_group") or "").strip()
        if not old_group or not new_group:
            continue
        key = (old_group, new_group)
        if key in seen:
            continue
        seen.add(key)
        rows.append(
            {
                "old_group": old_group,
                "new_group": new_group,
            }
        )
    rows.sort(key=lambda item: (-len(item["old_group"]), item["old_group"].lower()))
    return rows


def _allowed_mapping_pairs(selected_mappings: Sequence[Dict[str, str]]) -> set[tuple[str, str]]:
    out: set[tuple[str, str]] = set()
    for row in selected_mappings:
        old_group = str(row.get("old_group") or "").strip()
        new_group = str(row.get("new_group") or "").strip()
        if old_group and new_group:
            out.add((old_group, new_group))
    return out


def _collect_variable_targets(org_audit_report: Dict[str, Any]) -> List[Dict[str, str]]:
    targets: Dict[Tuple[str, str, str], Dict[str, str]] = {}
    for row in org_audit_report.get("variables") or []:
        variable_name = str(row.get("variable_name") or "").strip()
        variable_type = str(row.get("variable_type") or "").strip()
        if not variable_name or variable_type == "datasource":
            continue
        key = (
            str(row.get("grafana_org_id") or "").strip(),
            str(row.get("dashboard_uid") or "").strip(),
            variable_name,
        )
        if key in targets:
            continue
        targets[key] = {
            "grafana_org_id": key[0],
            "dashboard_uid": key[1],
            "dashboard_title": str(row.get("dashboard_title") or "").strip(),
            "folder_title": str(row.get("folder_title") or "").strip(),
            "dashboard_url": str(row.get("dashboard_url") or "").strip(),
            "variable_name": variable_name,
            "variable_type": variable_type,
            "datasource_names": str(row.get("datasource_names") or "").strip(),
        }
    return list(targets.values())


def _root_relative_field_path(json_path: str) -> str:
    text = str(json_path or "").strip()
    if text.startswith("dashboard."):
        return text[len("dashboard.") :]
    if text == "dashboard":
        return ""
    return text


def _is_supported_plan_field(location_kind: str, json_path: str, field_kind: str) -> bool:
    lower_path = str(json_path or "").lower()
    lower_kind = str(field_kind or "").strip().lower()
    if any(marker in lower_path for marker in (".host.filter", ".hosts.filter", ".host_filter", ".hosts_filter")):
        return False
    if location_kind == "variable":
        if lower_kind in {"query", "definition", "regex"}:
            return True
        return any(marker in lower_path for marker in (".current.", ".options["))
    if location_kind == "panel":
        return ".targets[" in lower_path and ".group.filter" in lower_path
    return False


def build_grafana_plan(
    conn: config.GrafanaConnection,
    org_audit_report: Dict[str, Any],
    selected_mappings: Sequence[Dict[str, str]],
    log: Callable[[str], None] | None = None,
) -> Dict[str, Any]:
    mapping_rows = _group_mappings_by_old(selected_mappings)
    if not mapping_rows:
        raise RuntimeError("Selected mappings are empty. Prepare mapping_plan.xlsx first.")

    variable_targets = _collect_variable_targets(org_audit_report)
    api_cache: Dict[int, GrafanaAPI] = {}
    dashboard_cache: Dict[Tuple[int, str], Dict[str, Any]] = {}
    plan_rows: List[Dict[str, Any]] = []
    review_rows: List[Dict[str, Any]] = []
    seen_rows: set[tuple[str, ...]] = set()

    _log(log, f"grafana-plan: start variable_targets={len(variable_targets)} selected_mappings={len(mapping_rows)}")

    for target_index, target in enumerate(variable_targets, start=1):
        org_id = int(target["grafana_org_id"] or 0)
        dashboard_uid = target["dashboard_uid"]
        variable_name = target["variable_name"]
        if target_index == 1 or target_index % 25 == 0:
            _log(log, f"grafana-plan: progress target={target_index}/{len(variable_targets)} org_id={org_id} uid={dashboard_uid} variable={variable_name}")

        api = api_cache.get(org_id)
        if api is None:
            api = GrafanaAPI(conn.base_url, conn.username, conn.password, org_id=org_id, timeout_sec=int(config.HTTP_TIMEOUT_SEC))
            api_cache[org_id] = api

        cache_key = (org_id, dashboard_uid)
        dashboard_payload = dashboard_cache.get(cache_key)
        if dashboard_payload is None:
            dashboard_payload = api.get_dashboard_by_uid(dashboard_uid)
            dashboard_cache[cache_key] = dashboard_payload
        dashboard = dashboard_payload.get("dashboard") or dashboard_payload

        variable_index, variable = _find_variable(dashboard, variable_name)
        if variable is None or variable_index is None:
            review_rows.append(
                {
                    "grafana_org_id": str(org_id),
                    "dashboard_uid": dashboard_uid,
                    "dashboard_title": target["dashboard_title"],
                    "dashboard_url": target["dashboard_url"],
                    "variable_name": variable_name,
                    "variable_type": target["variable_type"],
                    "status": "variable_not_found",
                    "message": "Variable missing in current dashboard JSON.",
                }
            )
            continue

        base_path = f"dashboard.templating.list[{variable_index}]"
        string_rows = _iter_strings(variable, path=base_path)
        field_hits = 0
        for json_path, text in string_rows:
            source_value = str(text or "")
            if not source_value:
                continue
            relative_path = json_path[len(base_path) + 1 :] if json_path.startswith(base_path + ".") else json_path[len(base_path) :]
            relative_path = relative_path.lstrip(".")
            if not relative_path:
                continue
            if not _is_supported_variable_field(relative_path):
                continue
            field_kind = _field_kind(json_path)
            matched_rows = [row for row in mapping_rows if row["old_group"] in source_value]
            if not matched_rows:
                continue
            field_hits += 1
            change_mode = "manual_regex" if _is_manual_change_mode(field_kind) else "exact"
            manual_required = change_mode == "manual_regex"
            for mapping in matched_rows:
                old_group = mapping["old_group"]
                new_group = mapping["new_group"]
                replace_count = source_value.count(old_group)
                planned_value = source_value.replace(old_group, new_group)
                row = {
                    "apply": "",
                    "grafana_org_id": str(org_id),
                    "dashboard_uid": dashboard_uid,
                    "dashboard_title": target["dashboard_title"],
                    "folder_title": target["folder_title"],
                    "dashboard_url": target["dashboard_url"],
                    "variable_name": variable_name,
                    "variable_type": target["variable_type"],
                    "datasource_names": target["datasource_names"],
                    "field_path": relative_path,
                    "field_kind": field_kind,
                    "old_group": old_group,
                    "new_group": new_group,
                    "replace_count": replace_count,
                    "source_value": source_value,
                    "planned_value": planned_value,
                    "change_kind": "substring_replace",
                    "change_mode": change_mode,
                    "manual_required": "yes" if manual_required else "",
                    "status": "review_regex" if manual_required else "candidate",
                    "comment": "",
                }
                row_key = (
                    row["grafana_org_id"],
                    row["dashboard_uid"],
                    row["variable_name"],
                    row["field_path"],
                    row["old_group"],
                    row["new_group"],
                )
                if row_key in seen_rows:
                    continue
                seen_rows.add(row_key)
                plan_rows.append(row)
        if field_hits:
            _log(log, f"grafana-plan: hit org_id={org_id} uid={dashboard_uid} variable={variable_name} fields={field_hits}")

    org_ids = normalize_values(org_audit_report.get("summary", {}).get("grafana_org_ids") or [])
    summary = {
        "grafana_org_ids": org_ids,
        "variable_targets": len(variable_targets),
        "selected_mappings": len(mapping_rows),
        "plan_rows": len(plan_rows),
        "manual_rows": sum(1 for row in plan_rows if str(row.get("manual_required") or "").strip()),
        "review_rows": len(review_rows),
    }
    _log(log, f"grafana-plan: summary={summary}")
    return {
        "summary": summary,
        "plan_rows": plan_rows,
        "review_rows": review_rows,
        "missing_variables": review_rows,
    }


def build_grafana_plan_from_impact(
    impact_plan: Dict[str, Any],
    log: Callable[[str], None] | None = None,
) -> Dict[str, Any]:
    plan_rows: List[Dict[str, Any]] = []
    review_rows: List[Dict[str, Any]] = []
    seen_rows: set[tuple[str, ...]] = set()
    grafana_rows = list(impact_plan.get("grafana_changes") or []) + list(impact_plan.get("grafana_manual_review") or [])

    _log(log, f"grafana-plan: start impact_rows={len(grafana_rows)}")

    for row in grafana_rows:
        org_id = str(row.get("grafana_org_id") or "").strip()
        dashboard_uid = str(row.get("dashboard_uid") or "").strip()
        dashboard_title = str(row.get("dashboard_title") or "").strip()
        dashboard_url = str(row.get("dashboard_url") or "").strip()
        panel_url = str(row.get("panel_url") or "").strip()
        panel_id = str(row.get("panel_id") or "").strip()
        panel_title = str(row.get("panel_title") or "").strip()
        panel_type = str(row.get("panel_type") or "").strip()
        variable_name = str(row.get("variable_name") or "").strip()
        variable_type = str(row.get("variable_type") or "").strip()
        location_kind = str(row.get("location_kind") or "").strip()
        field_kind = str(row.get("field_kind") or "").strip()
        json_path = str(row.get("json_path") or "").strip()
        field_path = _root_relative_field_path(json_path)
        old_group = str(row.get("old_group") or "").strip()
        new_group = str(row.get("new_group") or "").strip()
        source_value = str(row.get("source_text") or "")
        change_kind = str(row.get("change_kind") or "").strip()

        base_review = {
            "grafana_org_id": org_id,
            "dashboard_uid": dashboard_uid,
            "dashboard_title": dashboard_title,
            "dashboard_url": dashboard_url,
            "panel_url": panel_url,
            "panel_id": panel_id,
            "panel_title": panel_title,
            "panel_type": panel_type,
            "location_kind": location_kind,
            "variable_name": variable_name,
            "variable_type": variable_type,
            "field_kind": field_kind,
            "json_path": json_path,
            "old_group": old_group,
            "new_group": new_group,
        }

        if location_kind not in {"variable", "panel", "dashboard"}:
            review_rows.append(
                {
                    **base_review,
                    "status": "unsupported_location",
                    "message": "Unsupported Grafana location_kind for automatic apply.",
                }
            )
            continue
        if location_kind == "dashboard":
            review_rows.append(
                {
                    **base_review,
                    "status": "unsupported_location",
                    "message": "Dashboard-level matches stay in review; automatic apply is limited to variables and panel group.filter.",
                }
            )
            continue
        if not field_path or not _is_supported_plan_field(location_kind, json_path, field_kind):
            review_rows.append(
                {
                    **base_review,
                    "status": "unsupported_field",
                    "message": f"Unsupported Grafana field path for automatic apply: {field_path or json_path}",
                }
            )
            continue
        if not source_value:
            review_rows.append(
                {
                    **base_review,
                    "status": "empty_source_value",
                    "message": "Impact plan row does not contain source_text.",
                }
            )
            continue

        if change_kind == "replace_exact_string":
            if not old_group or not new_group or old_group not in source_value:
                review_rows.append(
                    {
                        **base_review,
                        "status": "invalid_exact_change",
                        "message": "Exact Grafana change is missing old/new group or old group is absent in source_text.",
                    }
                )
                continue
            change_mode = "manual_regex" if _is_manual_change_mode(field_kind) else "exact"
            manual_required = "yes" if change_mode == "manual_regex" else str(row.get("manual_required") or "")
            planned_value = source_value.replace(old_group, new_group)
            status = "review_regex" if manual_required else "candidate"
        elif change_kind == "review_pattern":
            if not old_group or not new_group:
                review_rows.append(
                    {
                        **base_review,
                        "status": "unresolved_pattern",
                        "message": "Pattern row is not mapped uniquely in impact plan.",
                    }
                )
                continue
            change_mode = "manual_regex"
            manual_required = "yes"
            planned_value = source_value.replace(old_group, new_group)
            status = "review_pattern"
        else:
            review_rows.append(
                {
                    **base_review,
                    "status": "unsupported_change_kind",
                    "message": f"Unsupported Grafana change_kind: {change_kind}",
                }
            )
            continue

        row_key = (org_id, dashboard_uid, location_kind, field_path, old_group, new_group)
        if row_key in seen_rows:
            continue
        seen_rows.add(row_key)
        plan_rows.append(
            {
                "apply": "",
                "grafana_org_id": org_id,
                "dashboard_uid": dashboard_uid,
                "dashboard_title": dashboard_title,
                "folder_title": str(row.get("folder_title") or ""),
                "dashboard_url": dashboard_url,
                "panel_url": panel_url,
                "panel_id": panel_id,
                "panel_title": panel_title,
                "panel_type": panel_type,
                "location_kind": location_kind,
                "variable_name": variable_name,
                "variable_type": variable_type,
                "datasource_names": str(row.get("datasource_names") or ""),
                "field_path": field_path,
                "json_path": json_path,
                "field_kind": field_kind,
                "old_group": old_group,
                "new_group": new_group,
                "replace_count": source_value.count(old_group) if old_group else 0,
                "source_value": source_value,
                "planned_value": planned_value,
                "change_kind": change_kind,
                "change_mode": change_mode,
                "manual_required": manual_required,
                "status": status,
                "comment": "",
            }
        )

    summary = {
        "scope_as": (impact_plan.get("summary") or {}).get("scope_as") or [],
        "scope_env": str((impact_plan.get("summary") or {}).get("scope_env") or "").strip(),
        "scope_gas": (impact_plan.get("summary") or {}).get("scope_gas") or [],
        "grafana_input_rows": len(grafana_rows),
        "plan_rows": len(plan_rows),
        "manual_rows": sum(1 for row in plan_rows if str(row.get("manual_required") or "").strip()),
        "review_rows": len(review_rows),
    }
    _log(log, f"grafana-plan: summary={summary}")
    return {
        "summary": summary,
        "plan_rows": plan_rows,
        "review_rows": review_rows,
        "missing_variables": review_rows,
    }


def apply_grafana_plan(
    conn: config.GrafanaConnection,
    selected_rows: Sequence[Dict[str, str]],
    selected_mappings: Sequence[Dict[str, str]],
    dry_run: bool = True,
    log: Callable[[str], None] | None = None,
) -> Dict[str, Any]:
    allowed_pairs = _allowed_mapping_pairs(selected_mappings)
    grouped: Dict[Tuple[int, str], List[Dict[str, str]]] = defaultdict(list)
    for row in selected_rows:
        grouped[(int(row["grafana_org_id"]), row["dashboard_uid"])].append(row)

    results: List[Dict[str, Any]] = []
    dashboards: List[Dict[str, Any]] = []
    api_cache: Dict[int, GrafanaAPI] = {}
    _log(log, f"grafana-apply: start dashboards={len(grouped)} mode={'dry-run' if dry_run else 'apply'} rows={len(selected_rows)}")

    for (org_id, dashboard_uid), rows in sorted(grouped.items(), key=lambda item: (item[0][0], item[0][1])):
        api = api_cache.get(org_id)
        if api is None:
            api = GrafanaAPI(conn.base_url, conn.username, conn.password, org_id=org_id, timeout_sec=int(config.HTTP_TIMEOUT_SEC))
            api_cache[org_id] = api

        payload = api.get_dashboard_by_uid(dashboard_uid)
        dashboard = payload.get("dashboard") or payload
        meta = payload.get("meta") or {}
        dashboard_title = str(dashboard.get("title") or rows[0].get("dashboard_title") or "")
        dashboard_url = str(meta.get("url") or rows[0].get("dashboard_url") or "")
        if dashboard_url and not dashboard_url.startswith("http://") and not dashboard_url.startswith("https://"):
            dashboard_url = f"{conn.base_url.rstrip('/')}{dashboard_url}"

        changed = 0
        requested = 0
        field_groups: Dict[Tuple[str, str], List[Dict[str, str]]] = defaultdict(list)
        for row in rows:
            requested += 1
            field_groups[(row["location_kind"], row["field_path"])].append(row)

        for (location_kind, field_path), field_rows in sorted(field_groups.items(), key=lambda item: (item[0][0].lower(), item[0][1].lower())):
            field_kind = str(field_rows[0].get("field_kind") or "")
            variable_name = str(field_rows[0].get("variable_name") or "")

            def append_result(row: Dict[str, str], status: str, before_value: str, after_value: str, message: str, applied: str = "") -> None:
                results.append(
                    {
                        "grafana_org_id": str(org_id),
                        "dashboard_uid": dashboard_uid,
                        "dashboard_title": dashboard_title,
                        "dashboard_url": dashboard_url,
                        "panel_url": str(row.get("panel_url") or ""),
                        "panel_id": str(row.get("panel_id") or ""),
                        "panel_title": str(row.get("panel_title") or ""),
                        "panel_type": str(row.get("panel_type") or ""),
                        "location_kind": str(row.get("location_kind") or ""),
                        "variable_name": str(row.get("variable_name") or ""),
                        "field_path": field_path,
                        "json_path": str(row.get("json_path") or ""),
                        "field_kind": field_kind,
                        "old_group": row["old_group"],
                        "new_group": row["new_group"],
                        "before_value": _shorten(before_value),
                        "after_value": _shorten(after_value),
                        "status": status,
                        "applied": applied,
                        "message": message,
                    }
                )

            try:
                current_value = _get_path_value(dashboard, field_path)
            except Exception as exc:
                for row in field_rows:
                    append_result(row, "field_not_found", "", "", str(exc))
                continue

            before_value = str(current_value or "")
            source_values = {str(row.get("source_value") or "") for row in field_rows}
            if len(source_values) != 1:
                for row in field_rows:
                    append_result(row, "plan_source_conflict", before_value, before_value, "Selected rows for the same field have different source_value.")
                continue

            source_value = next(iter(source_values))
            if before_value != source_value:
                for row in field_rows:
                    append_result(row, "source_mismatch", before_value, before_value, "Current field value differs from source_value saved in Grafana plan.")
                continue

            change_modes = {_resolve_change_mode(row) for row in field_rows}
            if len(change_modes) != 1:
                for row in field_rows:
                    append_result(row, "mixed_change_modes", before_value, before_value, "Selected rows for the same field contain different change_mode values.")
                continue

            change_mode = next(iter(change_modes))
            invalid_pair = next(
                (
                    row
                    for row in field_rows
                    if (str(row.get("old_group") or ""), str(row.get("new_group") or "")) not in allowed_pairs
                ),
                None,
            )
            if invalid_pair is not None:
                for row in field_rows:
                    append_result(row, "mapping_mismatch", before_value, before_value, "Grafana plan row is not aligned with selected impact mappings.")
                continue

            after_value = before_value
            if change_mode == "manual_regex":
                if len(field_rows) != 1:
                    for row in field_rows:
                        append_result(row, "multiple_manual_rows", before_value, before_value, "manual_regex requires exactly one selected row per field.")
                    continue
                manual_row = field_rows[0]
                after_value = str(manual_row.get("planned_value") or "")
                if not after_value:
                    append_result(manual_row, "empty_planned_value", before_value, "", "manual_regex row must contain planned_value.")
                    continue
            else:
                exact_rows = sorted(field_rows, key=lambda row: (-len(str(row.get("old_group") or "")), str(row.get("old_group") or "").lower()))
                for row in exact_rows:
                    expected_planned = source_value.replace(row["old_group"], row["new_group"])
                    if str(row.get("planned_value") or "") != expected_planned:
                        append_result(row, "exact_row_edited", before_value, before_value, "Exact row planned_value must stay derived from source_value and selected mappings.")
                        break
                else:
                    for row in exact_rows:
                        after_value = after_value.replace(row["old_group"], row["new_group"])
                    _set_path_value(dashboard, field_path, after_value)
                    changed += 1
                    for row in field_rows:
                        append_result(row, "dry_run_changed" if dry_run else "changed", before_value, after_value, "", "no" if dry_run else "yes")
                    continue
                continue

            _set_path_value(dashboard, field_path, after_value)
            changed += 1
            for row in field_rows:
                append_result(row, "dry_run_changed" if dry_run else "changed", before_value, after_value, "", "no" if dry_run else "yes")

        dashboard_status = "unchanged"
        dashboard_message = ""
        if changed and not dry_run:
            folder_id = int(meta.get("folderId") or 0)
            message = f"Host-group variable migration ({datetime.now().isoformat(timespec='seconds')})"
            api.update_dashboard(dashboard, folder_id, message)
            dashboard_status = "updated"
            dashboard_message = message
        elif changed:
            dashboard_status = "dry_run"
        dashboards.append(
            {
                "grafana_org_id": str(org_id),
                "dashboard_uid": dashboard_uid,
                "dashboard_title": dashboard_title,
                "dashboard_url": dashboard_url,
                "changes_requested": requested,
                "changes_applied": changed,
                "status": dashboard_status,
                "message": dashboard_message,
            }
        )
        _log(log, f"grafana-apply: dashboard org_id={org_id} uid={dashboard_uid} requested={requested} changed={changed} mode={'dry-run' if dry_run else 'apply'}")

    summary = {
        "mode": "dry-run" if dry_run else "apply",
        "selected_rows": len(selected_rows),
        "dashboards": len(dashboards),
        "changed_rows": sum(1 for row in results if row.get("status") in {"dry_run_changed", "changed"}),
        "updated_dashboards": sum(1 for row in dashboards if row.get("status") in {"dry_run", "updated"}),
        "errors": sum(1 for row in results if row.get("status") not in {"dry_run_changed", "changed"}),
    }
    _log(log, f"grafana-apply: summary={summary}")
    return {
        "summary": summary,
        "dashboards": dashboards,
        "results": results,
    }
