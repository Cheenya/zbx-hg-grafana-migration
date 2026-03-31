from __future__ import annotations

import json
from datetime import datetime
from typing import Any, Dict, Mapping, Sequence

from openpyxl import Workbook  # type: ignore

from common import autosize_columns


HOST_HEADERS = [
    "hostid",
    "host",
    "name",
    "status",
    "status_label",
    "ORG",
    "AS",
    "GAS",
    "GUEST_NAME",
    "OS_FAMILY",
    "ENV_RAW",
    "ENV_SCOPE",
    "has_old_groups",
    "has_standard_groups",
    "missing_any_new_group",
    "old_groups",
    "standard_groups",
    "env_groups",
    "as_groups",
    "gas_groups",
    "os_groups",
    "other_groups",
]

MISMATCH_OLDORG_HEADERS = [
    "hostid",
    "host",
    "name",
    "status",
    "status_label",
    "host_domain_org",
    "old_group_orgs",
    "old_groups",
    "details",
]

MISMATCH_PROXY_HEADERS = [
    "hostid",
    "host",
    "name",
    "status",
    "status_label",
    "host_domain_org",
    "proxyid",
    "proxy_name",
    "proxy_address",
    "proxy_domain_org",
    "details",
]

MISMATCH_LEGACY_ENV_HEADERS = [
    "hostid",
    "host",
    "name",
    "status",
    "status_label",
    "tag_env",
    "old_groups",
    "mismatches",
]

UNKNOWN_HOST_HEADERS = [
    "hostid",
    "host",
    "name",
    "status",
    "status_label",
    "ORG",
    "AS",
    "GAS",
    "GUEST_NAME",
    "OS_FAMILY",
    "ENV_RAW",
    "ENV_SCOPE",
    "groups",
    "unknown_reasons",
]

SKIPPED_HOST_HEADERS = [
    "hostid",
    "host",
    "name",
    "status",
    "status_label",
    "ORG",
    "AS",
    "GAS",
    "GUEST_NAME",
    "OS_FAMILY",
    "ENV_RAW",
    "ENV_SCOPE",
    "skip_reason",
]

GRAFANA_SUMMARY_HEADERS = [
    "AS",
    "grafana_org_id",
    "dashboard_uid",
    "dashboard_title",
    "folder_title",
    "dashboard_url",
    "hits_total",
    "exact_old",
    "pattern_old",
    "variable_hits",
    "panel_hits",
    "dashboard_hits",
    "panels",
    "variables",
]

GRAFANA_DETAIL_HEADERS = [
    "AS",
    "grafana_org_id",
    "dashboard_uid",
    "dashboard_title",
    "folder_title",
    "dashboard_url",
    "panel_url",
    "panel_id",
    "panel_title",
    "panel_type",
    "variable_name",
    "variable_type",
    "location_kind",
    "field_kind",
    "reference_kind",
    "match_type",
    "matched_string",
    "pattern_key",
    "source_text",
    "json_path",
    "count",
]

GRAFANA_ORG_SUMMARY_HEADERS = [
    "grafana_org_id",
    "zabbix_datasources",
    "zabbix_datasource_names",
    "dashboards_total",
    "dashboards_with_zabbix",
    "variables_with_zabbix",
    "panels_with_zabbix",
    "detail_rows",
]

GRAFANA_ORG_DATASOURCE_HEADERS = [
    "grafana_org_id",
    "datasource_id",
    "datasource_uid",
    "datasource_name",
    "datasource_type",
    "access",
    "url",
    "is_default",
    "read_only",
]

GRAFANA_ORG_DASHBOARD_HEADERS = [
    "grafana_org_id",
    "dashboard_uid",
    "dashboard_title",
    "folder_title",
    "dashboard_url",
    "dashboard_datasources",
    "dashboard_datasource_paths",
    "zabbix_variable_count",
    "zabbix_panel_count",
    "detail_rows",
]

GRAFANA_ORG_VARIABLE_HEADERS = [
    "grafana_org_id",
    "dashboard_uid",
    "dashboard_title",
    "folder_title",
    "dashboard_url",
    "variable_name",
    "variable_type",
    "datasource_names",
    "datasource_paths",
    "query",
    "regex",
    "definition",
    "refresh",
    "hide",
    "detail_rows",
]

GRAFANA_ORG_PANEL_HEADERS = [
    "grafana_org_id",
    "dashboard_uid",
    "dashboard_title",
    "folder_title",
    "dashboard_url",
    "panel_url",
    "panel_id",
    "panel_title",
    "panel_type",
    "datasource_names",
    "datasource_paths",
    "targets_total",
    "targets_zabbix",
    "detail_rows",
]

GRAFANA_ORG_DETAIL_HEADERS = [
    "grafana_org_id",
    "dashboard_uid",
    "dashboard_title",
    "folder_title",
    "dashboard_url",
    "panel_url",
    "panel_id",
    "panel_title",
    "panel_type",
    "variable_name",
    "variable_type",
    "location_kind",
    "field_kind",
    "reference_kind",
    "hint_kinds",
    "datasource_names",
    "datasource_paths",
    "source_text",
    "json_path",
]


def _append_rows(ws, rows: Sequence[Dict[str, Any]], headers: Sequence[str]) -> None:
    ws.append(list(headers))
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    autosize_columns(ws)


def _append_titled_table(ws, title: str, rows: Sequence[Dict[str, Any]], headers: Sequence[str]) -> None:
    ws.append([title])
    ws.append(list(headers))
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    ws.append([])


def _apply_hyperlinks(ws, rows: Sequence[Dict[str, Any]], headers: Sequence[str], link_map: Mapping[str, str]) -> None:
    index_by_header = {header: position + 1 for position, header in enumerate(headers)}
    for row_index, row in enumerate(rows, start=2):
        for header, url_field in link_map.items():
            column_index = index_by_header.get(header)
            if not column_index:
                continue
            url = str(row.get(url_field) or "").strip()
            if not url:
                continue
            cell = ws.cell(row=row_index, column=column_index)
            cell.hyperlink = url
            cell.style = "Hyperlink"


def save_inventory_json(report: Dict[str, Any], path: str) -> None:
    payload = {
        "meta": {
            "created_at": datetime.now().isoformat(timespec="seconds"),
            "version": "2.0",
        },
        "summary": report["summary"],
        "inventory": report["inventory"],
        "unknown_hosts": report["unknown_hosts"],
        "hosts": report["hosts"],
        "hosts_replace": report["hosts_replace"],
        "hosts_no_any_new": report["hosts_no_any_new"],
        "host_enrichment": report["host_enrichment"],
        "hosts_need_enrichment": report["hosts_need_enrichment"],
        "hosts_clean": report["hosts_clean"],
        "hosts_skipped_env": report["hosts_skipped_env"],
        "hosts_skipped_gas": report["hosts_skipped_gas"],
        "mismatch_host_oldorg": report["mismatch_host_oldorg"],
        "mismatch_host_proxyorg": report["mismatch_host_proxyorg"],
        "mismatch_legacy_env": report["mismatch_legacy_env"],
        "host_expected_groups": report["host_expected_groups"],
        "env_summary": report["env_summary"],
        "gas_summary": report["gas_summary"],
        "guest_name_summary": report["guest_name_summary"],
        "groups_old": report["groups_old"],
        "groups_new": report["groups_new"],
        "expected_groups": report["expected_groups"],
        "mapping_plan": report["mapping_plan"],
        "zabbix_mapping_preview": report["zabbix_mapping_preview"],
        "actions": report["actions"],
        "usergroups": report["usergroups"],
        "maintenances": report["maintenances"],
        "grafana": report["grafana"],
        "grafana_summary": report["grafana_summary"],
    }
    with open(path, "w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2)


def write_workbook(report: Dict[str, Any], out_path: str) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    summary_ws = wb.create_sheet("SUMMARY")
    summary_ws.append(["key", "value"])
    for key, value in report["summary"].items():
        if isinstance(value, list):
            rendered = ", ".join(str(item) for item in value)
        else:
            rendered = value
        summary_ws.append([key, rendered])
    autosize_columns(summary_ws)

    env_ws = wb.create_sheet("ENV_SUMMARY")
    _append_rows(
        env_ws,
        report["env_summary"],
        ["AS", "ENV_RAW", "ENV_SCOPE", "hosts_count", "enabled_hosts", "disabled_hosts", "legacy_hosts", "sample_hosts"],
    )

    gas_ws = wb.create_sheet("GAS_SUMMARY")
    _append_rows(
        gas_ws,
        report["gas_summary"],
        ["AS", "GAS", "hosts_count", "enabled_hosts", "disabled_hosts", "legacy_hosts", "sample_hosts"],
    )

    guest_ws = wb.create_sheet("GUEST_NAME_SUMMARY")
    _append_rows(
        guest_ws,
        report["guest_name_summary"],
        ["AS", "GUEST_NAME", "hosts_count", "enabled_hosts", "disabled_hosts", "legacy_hosts", "sample_hosts"],
    )

    hosts_ws = wb.create_sheet("HOSTS")
    _append_rows(hosts_ws, report["hosts"], HOST_HEADERS)

    replace_ws = wb.create_sheet("HOSTS_OLD_SCOPE")
    _append_rows(replace_ws, report["hosts_replace"], HOST_HEADERS)

    replace_missing_ws = wb.create_sheet("HOSTS_NO_ANY_NEW")
    _append_rows(replace_missing_ws, report["hosts_no_any_new"], HOST_HEADERS)

    enrichment_ws = wb.create_sheet("HOST_ENRICHMENT")
    _append_rows(
        enrichment_ws,
        report["host_enrichment"],
        [
            "hostid",
            "host",
            "name",
            "status",
            "status_label",
            "ORG",
            "AS",
            "ENV_RAW",
            "ENV_SCOPE",
            "GAS",
            "GUEST_NAME",
            "OS_FAMILY",
            "old_groups",
            "standard_groups",
            "expected_env_groups",
            "expected_as_groups",
            "expected_gas_groups",
            "expected_os_groups",
            "catalog_existing_groups",
            "catalog_missing_groups",
            "host_present_expected_groups",
            "host_missing_expected_groups",
            "suggested_pairs",
            "suggested_new_groups",
            "host_action",
            "unresolved_reasons",
            "manual_required",
        ],
    )

    need_enrichment_ws = wb.create_sheet("HOSTS_NEED_ENRICH")
    _append_rows(
        need_enrichment_ws,
        report["hosts_need_enrichment"],
        [
            "hostid",
            "host",
            "name",
            "status",
            "status_label",
            "ORG",
            "AS",
            "ENV_RAW",
            "ENV_SCOPE",
            "GAS",
            "GUEST_NAME",
            "OS_FAMILY",
            "old_groups",
            "standard_groups",
            "expected_env_groups",
            "expected_as_groups",
            "expected_gas_groups",
            "expected_os_groups",
            "catalog_existing_groups",
            "catalog_missing_groups",
            "host_present_expected_groups",
            "host_missing_expected_groups",
            "suggested_pairs",
            "suggested_new_groups",
            "host_action",
            "unresolved_reasons",
            "manual_required",
        ],
    )

    clean_ws = wb.create_sheet("HOSTS_CLEAN")
    _append_rows(clean_ws, report["hosts_clean"], HOST_HEADERS)

    unknown_ws = wb.create_sheet("UNKNOWN_HOSTS")
    _append_rows(unknown_ws, report["unknown_hosts"], UNKNOWN_HOST_HEADERS)

    skipped_ws = wb.create_sheet("HOSTS_SKIPPED_ENV")
    _append_rows(skipped_ws, report["hosts_skipped_env"], SKIPPED_HOST_HEADERS)

    skipped_gas_ws = wb.create_sheet("HOSTS_SKIPPED_GAS")
    _append_rows(skipped_gas_ws, report["hosts_skipped_gas"], SKIPPED_HOST_HEADERS)

    mismatch_ws = wb.create_sheet("MISMATCHES")
    _append_titled_table(
        mismatch_ws,
        "HOST DOMAIN ORG != OLD GROUP ORG",
        report["mismatch_host_oldorg"],
        MISMATCH_OLDORG_HEADERS,
    )
    _append_titled_table(
        mismatch_ws,
        "HOST DOMAIN ORG != PROXY ORG",
        report["mismatch_host_proxyorg"],
        MISMATCH_PROXY_HEADERS,
    )
    _append_titled_table(
        mismatch_ws,
        "OLD GROUP ENV != TAG ENV",
        report["mismatch_legacy_env"],
        MISMATCH_LEGACY_ENV_HEADERS,
    )
    autosize_columns(mismatch_ws)

    old_ws = wb.create_sheet("GROUPS_OLD")
    _append_rows(
        old_ws,
        report["groups_old"],
        ["group_name", "groupid", "legacy_env_tokens", "org_values", "as_values", "env_raw_values", "env_scope_values", "hosts_count", "sample_hosts"],
    )

    new_ws = wb.create_sheet("GROUPS_NEW")
    _append_rows(
        new_ws,
        report["groups_new"],
        [
            "group_name",
            "groupid",
            "group_kind",
            "org",
            "as_values",
            "env_raw_values",
            "env_scope_values",
            "gas_values",
            "os_families",
            "hosts_count",
            "sample_hosts",
        ],
    )

    expected_ws = wb.create_sheet("EXPECTED_GROUPS")
    _append_rows(
        expected_ws,
        report["expected_groups"],
        [
            "group_name",
            "groupid",
            "group_kind",
            "org",
            "exists_in_zabbix",
            "source_hosts_count",
            "host_present_count",
            "host_missing_count",
            "as_values",
            "env_raw_values",
            "env_scope_values",
            "gas_values",
            "os_families",
            "sample_hosts",
        ],
    )

    expected_host_ws = wb.create_sheet("HOST_EXPECTED")
    _append_rows(
        expected_host_ws,
        report["host_expected_groups"],
        [
            "hostid",
            "host",
            "name",
            "status",
            "status_label",
            "ORG",
            "AS",
            "GAS",
            "GUEST_NAME",
            "OS_FAMILY",
            "ENV_RAW",
            "ENV_SCOPE",
            "group_name",
            "groupid",
            "group_kind",
            "exists_in_zabbix",
            "on_host",
        ],
    )

    mapping_ws = wb.create_sheet("MAPPING_PLAN")
    _append_rows(
        mapping_ws,
        report["mapping_plan"],
        [
            "selected",
            "AS",
            "ORG",
            "old_group",
            "old_groupid",
            "old_group_kind",
            "legacy_env_token",
            "new_group",
            "new_groupid",
            "target_kind",
            "target_exists",
            "candidate_rank",
            "candidate_count",
            "old_hosts_count",
            "target_scope_hosts",
            "new_hosts_count",
            "host_action",
            "hosts_need_add_new",
            "hosts_already_have_new",
            "old_orgs",
            "old_envs",
            "old_env_scopes",
            "target_env_raw",
            "auto_reason",
            "manual_required",
            "status",
            "comment",
        ],
    )

    preview_ws = wb.create_sheet("ZBX_MAP_PREVIEW")
    _append_rows(
        preview_ws,
        report["zabbix_mapping_preview"],
        [
            "object_type",
            "object_id",
            "object_name",
            "where_found",
            "field_path",
            "old_group",
            "old_groupid",
            "candidate_new_group",
            "candidate_new_groupid",
            "candidate_rank",
            "candidate_count",
            "target_kind",
            "target_exists",
            "mapping_status",
            "object_has_candidate_new",
            "include_reason",
            "manual_required",
            "host_action",
            "hosts_need_add_new",
            "hosts_already_have_new",
        ],
    )

    actions_ws = wb.create_sheet("ACTIONS")
    _append_rows(
        actions_ws,
        report["actions"],
        [
            "actionid",
            "name",
            "status",
            "where_found",
            "matched_groupids",
            "matched_group_names",
            "candidate_new_groupids_present",
            "candidate_new_group_names_present",
            "include_reason",
            "recipient_usergroups",
            "recipient_users",
            "recipients_media",
        ],
    )

    usergroups_ws = wb.create_sheet("USERGROUPS")
    _append_rows(
        usergroups_ws,
        report["usergroups"],
        [
            "usrgrpid",
            "name",
            "rights_on_old_groups",
            "rights_on_new_groups",
            "candidate_new_groups_already_present",
            "matching_tag_filters",
            "include_reason",
            "users",
            "users_media",
            "is_action_recipient",
        ],
    )

    maintenances_ws = wb.create_sheet("MAINTENANCES")
    _append_rows(
        maintenances_ws,
        report["maintenances"],
        [
            "maintenanceid",
            "name",
            "matched_groupids",
            "matched_group_names",
            "candidate_new_groupids_present",
            "candidate_new_group_names_present",
            "include_reason",
            "active_since",
            "active_till",
        ],
    )

    grafana_summary_ws = wb.create_sheet("GRAFANA_SUMMARY")
    _append_rows(grafana_summary_ws, report["grafana_summary"], GRAFANA_SUMMARY_HEADERS)
    _apply_hyperlinks(grafana_summary_ws, report["grafana_summary"], GRAFANA_SUMMARY_HEADERS, {"dashboard_url": "dashboard_url"})

    inventory_ws = wb.create_sheet("INVENTORY")
    inventory_ws.append(["section", "value"])
    for key, value in report["inventory"].items():
        inventory_ws.append([key, json.dumps(value, ensure_ascii=False)])
    autosize_columns(inventory_ws)

    wb.save(out_path)


def write_grafana_workbook(report: Dict[str, Any], out_path: str) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    summary_rows = [
        {"key": "scope_as", "value": ", ".join(str(item) for item in report["summary"].get("scope_as") or [])},
        {"key": "scope_env", "value": str(report["summary"].get("scope_env") or "")},
        {"key": "grafana_dashboards", "value": report["summary"].get("grafana_dashboards", len(report["grafana_summary"]))},
        {"key": "grafana_rows", "value": report["summary"].get("grafana_rows", len(report["grafana"]))},
        {"key": "grafana_error", "value": report["summary"].get("grafana_error", "")},
    ]
    summary_ws = wb.create_sheet("SUMMARY")
    _append_rows(summary_ws, summary_rows, ["key", "value"])

    dashboards_ws = wb.create_sheet("DASHBOARDS")
    _append_rows(dashboards_ws, report["grafana_summary"], GRAFANA_SUMMARY_HEADERS)
    _apply_hyperlinks(dashboards_ws, report["grafana_summary"], GRAFANA_SUMMARY_HEADERS, {"dashboard_url": "dashboard_url"})

    details_ws = wb.create_sheet("DETAILS")
    _append_rows(details_ws, report["grafana"], GRAFANA_DETAIL_HEADERS)
    _apply_hyperlinks(
        details_ws,
        report["grafana"],
        GRAFANA_DETAIL_HEADERS,
        {
            "dashboard_url": "dashboard_url",
            "panel_url": "panel_url",
        },
    )

    wb.save(out_path)


def save_grafana_org_json(report: Dict[str, Any], path: str) -> None:
    payload = {
        "meta": {
            "created_at": datetime.now().isoformat(timespec="seconds"),
            "version": "2.0",
        },
        "summary": report["summary"],
        "org_summary": report["org_summary"],
        "datasources": report["datasources"],
        "dashboards": report["dashboards"],
        "variables": report["variables"],
        "panels": report["panels"],
        "details": report["details"],
    }
    with open(path, "w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2)


def write_grafana_org_workbook(report: Dict[str, Any], out_path: str) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    summary_ws = wb.create_sheet("SUMMARY")
    summary_ws.append(["key", "value"])
    for key, value in report["summary"].items():
        if isinstance(value, list):
            rendered = ", ".join(str(item) for item in value)
        else:
            rendered = value
        summary_ws.append([key, rendered])
    autosize_columns(summary_ws)

    org_ws = wb.create_sheet("ORGS")
    _append_rows(org_ws, report["org_summary"], GRAFANA_ORG_SUMMARY_HEADERS)

    datasources_ws = wb.create_sheet("DATASOURCES")
    _append_rows(datasources_ws, report["datasources"], GRAFANA_ORG_DATASOURCE_HEADERS)

    dashboards_ws = wb.create_sheet("DASHBOARDS")
    _append_rows(dashboards_ws, report["dashboards"], GRAFANA_ORG_DASHBOARD_HEADERS)
    _apply_hyperlinks(dashboards_ws, report["dashboards"], GRAFANA_ORG_DASHBOARD_HEADERS, {"dashboard_url": "dashboard_url"})

    variables_ws = wb.create_sheet("VARIABLES")
    _append_rows(variables_ws, report["variables"], GRAFANA_ORG_VARIABLE_HEADERS)
    _apply_hyperlinks(variables_ws, report["variables"], GRAFANA_ORG_VARIABLE_HEADERS, {"dashboard_url": "dashboard_url"})

    panels_ws = wb.create_sheet("PANELS")
    _append_rows(panels_ws, report["panels"], GRAFANA_ORG_PANEL_HEADERS)
    _apply_hyperlinks(
        panels_ws,
        report["panels"],
        GRAFANA_ORG_PANEL_HEADERS,
        {
            "dashboard_url": "dashboard_url",
            "panel_url": "panel_url",
        },
    )

    details_ws = wb.create_sheet("DETAILS")
    _append_rows(details_ws, report["details"], GRAFANA_ORG_DETAIL_HEADERS)
    _apply_hyperlinks(
        details_ws,
        report["details"],
        GRAFANA_ORG_DETAIL_HEADERS,
        {
            "dashboard_url": "dashboard_url",
            "panel_url": "panel_url",
        },
    )

    wb.save(out_path)
