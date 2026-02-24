#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import json
from dataclasses import dataclass
from typing import Any, Dict, List, Optional

import urllib3  # type: ignore
import requests  # type: ignore
from openpyxl import Workbook  # type: ignore
from openpyxl.styles import Font  # type: ignore
from openpyxl.utils import get_column_letter  # type: ignore

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# --- Вводишь здесь напрямую ---
ZBX_URL = "https://zabbix-api.example.ru/api_jsonrpc.php"
ZBX_USER = "login"
ZBX_PASSWORD = "password"

# --- Настройки тегов (как у вас в Zabbix) ---
TAG_AS = "AS"
TAG_ASN = "ASN"
TAG_ENV = "ENV"

# --- Куда сохраняем Excel ---
OUT_XLSX = "zbx_env_values_dump.xlsx"


class ZabbixAPI:
    def __init__(self, api_url: str, timeout: int = 120) -> None:
        self.api_url = api_url
        self.timeout = timeout
        self.auth: Optional[str] = None
        self._id = 1

    def call(self, method: str, params: Dict[str, Any]) -> Any:
        payload: Dict[str, Any] = {
            "jsonrpc": "2.0",
            "method": method,
            "params": params,
            "id": self._id,
        }
        self._id += 1
        if self.auth is not None:
            payload["auth"] = self.auth

        r = requests.post(
            self.api_url,
            json=payload,
            timeout=self.timeout,
            verify=False,  # как просили
        )
        r.raise_for_status()
        data = r.json()
        if "error" in data:
            raise RuntimeError(f"Zabbix API error ({method}): {data['error']}")
        return data["result"]

    def login(self, username: str, password: str) -> None:
        self.auth = self.call("user.login", {"username": username, "password": password})


def get_tag_value(tags: List[Dict[str, Any]], tag_name: str) -> Optional[str]:
    for t in tags or []:
        if t.get("tag") == tag_name:
            v = t.get("value")
            if v is None:
                return None
            v = str(v).strip()
            return v if v else None
    return None


def autosize_columns(ws, min_width: int = 10, max_width: int = 80) -> None:
    for col in ws.columns:
        mx = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is None:
                continue
            mx = max(mx, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max_width, max(min_width, mx + 2))


def fetch_hosts(api: ZabbixAPI) -> List[Dict[str, Any]]:
    # Здесь можно включить только monitored_hosts при необходимости
    return api.call(
        "host.get",
        {
            "output": ["hostid", "host", "name"],
            "selectTags": ["tag", "value"],
            "selectGroups": ["groupid", "name"],
        },
    )


@dataclass
class EnvInfo:
    env: str
    hosts_count: int = 0
    example_host: str = ""
    example_name: str = ""
    example_as: str = ""
    example_asn: str = ""
    example_groups: str = ""
    example_tags_json: str = ""


def main() -> int:
    api = ZabbixAPI(ZBX_URL)
    api.login(ZBX_USER, ZBX_PASSWORD)

    print("Fetching hosts...")
    hosts = fetch_hosts(api)
    print(f"Hosts fetched: {len(hosts)}")

    env_map: Dict[str, EnvInfo] = {}
    no_env_rows: List[Dict[str, Any]] = []

    for h in hosts:
        tags = h.get("tags") or []
        groups = h.get("groups") or []

        env = get_tag_value(tags, TAG_ENV)
        as_val = get_tag_value(tags, TAG_AS) or ""
        asn_val = get_tag_value(tags, TAG_ASN) or ""

        host = str(h.get("host") or "")
        name = str(h.get("name") or "")

        groups_str = ", ".join(sorted({str(g.get("name")) for g in groups if g.get("name")}))

        tags_dict = {str(t.get("tag")): t.get("value") for t in tags if t.get("tag")}
        tags_json = json.dumps(tags_dict, ensure_ascii=False)

        if not env:
            no_env_rows.append(
                {
                    "hostid": h.get("hostid"),
                    "host": host,
                    "name": name,
                    "AS": as_val,
                    "ASN": asn_val,
                    "groups": groups_str,
                    "tags_json": tags_json,
                }
            )
            continue

        if env not in env_map:
            env_map[env] = EnvInfo(
                env=env,
                hosts_count=0,
                example_host=host,
                example_name=name,
                example_as=as_val,
                example_asn=asn_val,
                example_groups=groups_str,
                example_tags_json=tags_json,
            )

        env_map[env].hosts_count += 1

    # --- Excel ---
    wb = Workbook()
    ws = wb.active
    ws.title = "ENV_VALUES"

    ws.append(
        [
            "ENV",
            "hosts_count",
            "example_host",
            "example_name",
            "example_AS",
            "example_ASN",
            "example_groups",
            "example_tags_json",
        ]
    )
    for c in range(1, ws.max_column + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)

    for env in sorted(env_map.keys(), key=lambda x: x.lower()):
        e = env_map[env]
        ws.append(
            [
                e.env,
                e.hosts_count,
                e.example_host,
                e.example_name,
                e.example_as,
                e.example_asn,
                e.example_groups,
                e.example_tags_json,
            ]
        )

    autosize_columns(ws)

    ws2 = wb.create_sheet("NO_ENV_TAG")
    ws2.append(["hostid", "host", "name", "AS", "ASN", "groups", "tags_json"])
    for c in range(1, ws2.max_column + 1):
        ws2.cell(row=1, column=c).font = Font(bold=True)

    for r in no_env_rows:
        ws2.append([r["hostid"], r["host"], r["name"], r["AS"], r["ASN"], r["groups"], r["tags_json"]])

    autosize_columns(ws2)

    wb.save(OUT_XLSX)
    print(f"Saved: {OUT_XLSX}")
    print(f"ENV unique values: {len(env_map)}")
    print(f"Hosts without ENV tag: {len(no_env_rows)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())