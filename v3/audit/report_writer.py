from __future__ import annotations

import json
from datetime import datetime
from typing import Any, Dict, Iterable, List, Mapping, Sequence

from openpyxl import Workbook  # type: ignore
from openpyxl.styles import Font, PatternFill  # type: ignore

from core.common import autosize_columns, join_sorted, parse_standard_group
from planning.mapping_plan import MAPPING_PLAN_HEADERS


MISMATCH_OLDORG_HEADERS = [
    "hostid",
    "host",
    "name",
    "status",
    "status_label",
    "host_domain_org",
    "old_group_orgs",
    "old_groups",
    "details",
]

MISMATCH_PROXY_HEADERS = [
    "hostid",
    "host",
    "name",
    "status",
    "status_label",
    "host_domain_org",
    "proxyid",
    "proxy_name",
    "proxy_address",
    "proxy_domain_org",
    "details",
]

MISMATCH_LEGACY_ENV_HEADERS = [
    "hostid",
    "host",
    "name",
    "status",
    "status_label",
    "tag_env",
    "old_groups",
    "mismatches",
]

GRAFANA_SUMMARY_HEADERS = [
    "AS",
    "grafana_org_id",
    "dashboard_uid",
    "dashboard_title",
    "folder_title",
    "dashboard_url",
    "hits_total",
    "exact_old",
    "pattern_old",
    "variable_hits",
    "panel_hits",
    "dashboard_hits",
    "panels",
    "variables",
]

GRAFANA_DETAIL_HEADERS = [
    "AS",
    "grafana_org_id",
    "dashboard_uid",
    "dashboard_title",
    "folder_title",
    "dashboard_url",
    "panel_url",
    "panel_id",
    "panel_title",
    "panel_type",
    "variable_name",
    "variable_type",
    "location_kind",
    "field_kind",
    "reference_kind",
    "match_type",
    "matched_string",
    "pattern_key",
    "source_text",
    "json_path",
    "count",
    "suggested_old_group",
    "suggested_new_group",
    "suggested_value",
    "suggestion_status",
    "manual_required",
]

GRAFANA_SUGGESTION_HEADERS = [
    "AS",
    "grafana_org_id",
    "dashboard_uid",
    "dashboard_title",
    "folder_title",
    "dashboard_url",
    "panel_url",
    "panel_id",
    "panel_title",
    "panel_type",
    "variable_name",
    "variable_type",
    "location_kind",
    "field_kind",
    "reference_kind",
    "match_type",
    "matched_string",
    "pattern_key",
    "source_text",
    "json_path",
    "suggested_old_group",
    "suggested_new_group",
    "suggested_value",
    "suggestion_status",
    "manual_required",
    "implementation_status",
    "implementation_reason",
]

GRAFANA_ORG_SUMMARY_HEADERS = [
    "grafana_org_id",
    "zabbix_datasources",
    "zabbix_datasource_names",
    "dashboards_total",
    "dashboards_with_zabbix",
    "variables_with_zabbix",
    "panels_with_zabbix",
    "detail_rows",
]

GRAFANA_ORG_DATASOURCE_HEADERS = [
    "grafana_org_id",
    "datasource_id",
    "datasource_uid",
    "datasource_name",
    "datasource_type",
    "access",
    "url",
    "is_default",
    "read_only",
]

GRAFANA_ORG_DASHBOARD_HEADERS = [
    "grafana_org_id",
    "dashboard_uid",
    "dashboard_title",
    "folder_title",
    "dashboard_url",
    "dashboard_datasources",
    "dashboard_datasource_paths",
    "zabbix_variable_count",
    "zabbix_panel_count",
    "detail_rows",
]

GRAFANA_ORG_VARIABLE_HEADERS = [
    "grafana_org_id",
    "dashboard_uid",
    "dashboard_title",
    "folder_title",
    "dashboard_url",
    "variable_name",
    "variable_type",
    "datasource_names",
    "datasource_paths",
    "query",
    "regex",
    "definition",
    "refresh",
    "hide",
    "detail_rows",
]

GRAFANA_ORG_PANEL_HEADERS = [
    "grafana_org_id",
    "dashboard_uid",
    "dashboard_title",
    "folder_title",
    "dashboard_url",
    "panel_url",
    "panel_id",
    "panel_title",
    "panel_type",
    "datasource_names",
    "datasource_paths",
    "targets_total",
    "targets_zabbix",
    "detail_rows",
]

GRAFANA_ORG_DETAIL_HEADERS = [
    "grafana_org_id",
    "dashboard_uid",
    "dashboard_title",
    "folder_title",
    "dashboard_url",
    "panel_url",
    "panel_id",
    "panel_title",
    "panel_type",
    "variable_name",
    "variable_type",
    "location_kind",
    "field_kind",
    "reference_kind",
    "hint_kinds",
    "datasource_names",
    "datasource_paths",
    "source_text",
    "json_path",
]

GRAFANA_ORG_SUGGESTION_HEADERS = [
    "grafana_org_id",
    "dashboard_uid",
    "dashboard_title",
    "folder_title",
    "dashboard_url",
    "panel_url",
    "panel_id",
    "panel_title",
    "panel_type",
    "variable_name",
    "variable_type",
    "location_kind",
    "field_kind",
    "reference_kind",
    "json_path",
    "old_group",
    "new_group",
    "source_text",
    "planned_value",
    "suggestion_status",
    "manual_required",
]

GRAFANA_ORG_ERROR_HEADERS = [
    "grafana_org_id",
    "dashboard_uid",
    "dashboard_title",
    "folder_title",
    "dashboard_url",
    "status",
    "message",
]

EXPECTED_GROUP_HEADERS = [
    "group_name",
    "groupid",
    "group_kind",
    "org",
    "exists_in_zabbix",
    "source_hosts_count",
    "host_present_count",
    "host_missing_count",
    "as_values",
    "env_raw_values",
    "env_scope_values",
    "gas_values",
    "os_families",
    "sample_hosts",
]

HOST_EXPECTED_HEADERS = [
    "hostid",
    "host",
    "name",
    "status",
    "status_label",
    "ORG",
    "AS",
    "GAS",
    "GUEST_NAME",
    "OS_FAMILY",
    "ENV_RAW",
    "ENV_SCOPE",
    "group_name",
    "groupid",
    "group_kind",
    "exists_in_zabbix",
    "on_host",
]

HOST_VIEW_HEADERS = [
    "hostid",
    "host",
    "name",
    "status_label",
    "ORG",
    "AS",
    "GAS",
    "ENV_RAW",
    "ENV_SCOPE",
    "OS_FAMILY",
    "current_groups",
    "add_env_groups",
    "add_as_groups",
    "add_gas_groups",
    "add_os_groups",
    "missing_in_zabbix",
    "manual_required",
    "unresolved_reasons",
]

EXCLUDED_HOST_HEADERS = [
    "exclude_kind",
    "hostid",
    "host",
    "name",
    "status_label",
    "ORG",
    "AS",
    "GAS",
    "ENV_RAW",
    "ENV_SCOPE",
    "OS_FAMILY",
    "current_groups",
    "reason",
]

DISCOVERY_HOST_HEADERS = [
    "hostid",
    "host",
    "name",
    "status_label",
    "ORG",
    "AS",
    "GAS",
    "ENV_RAW",
    "ENV_SCOPE",
    "OS_FAMILY",
    "proxyid",
    "proxy_name",
    "current_groups",
    "flags",
    "discovery_parent_hostid",
    "discovery_parent_itemid",
    "discovery_status",
    "discovery_disable_source",
    "discovery_ts_disable",
    "discovery_ts_delete",
    "discovery_reason",
    "scope_skip_reason",
]

OBJECT_PREVIEW_HEADERS = [
    "object_type",
    "object_id",
    "object_name",
    "where_found",
    "field_paths",
    "reference_status",
    "permission_status",
    "current_groups",
    "add_env_groups",
    "add_as_groups",
    "add_gas_groups",
    "add_os_groups",
    "missing_in_zabbix",
    "include_reason",
    "manual_required",
]

ACTION_VIEW_HEADERS = [
    "actionid",
    "name",
    "status",
    "where_found",
    "reference_status",
    "current_groups",
    "add_env_groups",
    "add_as_groups",
    "add_gas_groups",
    "add_os_groups",
    "missing_in_zabbix",
    "include_reason",
    "recipient_usergroups",
    "recipient_users",
    "recipients_media",
    "manual_required",
]

USERGROUP_VIEW_HEADERS = [
    "usrgrpid",
    "name",
    "current_groups",
    "permission_status",
    "add_env_groups",
    "add_as_groups",
    "add_gas_groups",
    "add_os_groups",
    "missing_in_zabbix",
    "matching_tag_filters",
    "include_reason",
    "users",
    "users_media",
    "is_action_recipient",
    "manual_required",
]

MAINTENANCE_VIEW_HEADERS = [
    "maintenanceid",
    "name",
    "reference_status",
    "current_groups",
    "add_env_groups",
    "add_as_groups",
    "add_gas_groups",
    "add_os_groups",
    "missing_in_zabbix",
    "include_reason",
    "active_since",
    "active_till",
    "manual_required",
]

SUMMARY_TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
SUMMARY_HEADER_FILL = PatternFill("solid", fgColor="D9E2F3")
DEFAULT_HEADER_FILL = PatternFill("solid", fgColor="D9EAD3")
CURRENT_FILL = PatternFill("solid", fgColor="EDEDED")
INFO_FILL = PatternFill("solid", fgColor="DDEBF7")
LINK_FILL = PatternFill("solid", fgColor="EAF2F8")
ENV_FILL = PatternFill("solid", fgColor="FFF2CC")
AS_FILL = PatternFill("solid", fgColor="D9E2F3")
GAS_FILL = PatternFill("solid", fgColor="E2F0D9")
OS_FILL = PatternFill("solid", fgColor="FCE4D6")
ALERT_FILL = PatternFill("solid", fgColor="F4CCCC")
HEADER_FONT = Font(bold=True, color="000000")
TITLE_FONT = Font(bold=True, color="FFFFFF")


def _append_rows(ws, rows: Sequence[Dict[str, Any]], headers: Sequence[str]) -> None:
    ws.append(list(headers))
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    autosize_columns(ws)


def _append_titled_table(ws, title: str, rows: Sequence[Dict[str, Any]], headers: Sequence[str]) -> None:
    ws.append([title])
    ws.append(list(headers))
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    ws.append([])


def _apply_hyperlinks(ws, rows: Sequence[Dict[str, Any]], headers: Sequence[str], link_map: Mapping[str, str]) -> None:
    index_by_header = {header: position + 1 for position, header in enumerate(headers)}
    for row_index, row in enumerate(rows, start=2):
        for header, url_field in link_map.items():
            column_index = index_by_header.get(header)
            if not column_index:
                continue
            url = str(row.get(url_field) or "").strip()
            if not url:
                continue
            cell = ws.cell(row=row_index, column=column_index)
            cell.hyperlink = url
            cell.style = "Hyperlink"


def _style_header_row(ws, row_index: int, headers: Sequence[str], fill: PatternFill = DEFAULT_HEADER_FILL) -> None:
    for column_index, _ in enumerate(headers, start=1):
        cell = ws.cell(row=row_index, column=column_index)
        cell.fill = fill
        cell.font = HEADER_FONT


def _style_title_row(ws, row_index: int) -> None:
    cell = ws.cell(row=row_index, column=1)
    cell.fill = SUMMARY_TITLE_FILL
    cell.font = TITLE_FONT


def _apply_kind_column_fills(ws, headers: Sequence[str]) -> None:
    header_index = {header: index + 1 for index, header in enumerate(headers)}
    fill_by_header = {
        "current_groups": CURRENT_FILL,
        "groups": CURRENT_FILL,
        "old_groups": CURRENT_FILL,
        "standard_groups": CURRENT_FILL,
        "matched_group_names": CURRENT_FILL,
        "rights_on_old_groups": CURRENT_FILL,
        "rights_on_new_groups": CURRENT_FILL,
        "source_text": CURRENT_FILL,
        "planned_value": CURRENT_FILL,
        "suggested_old_group": CURRENT_FILL,
        "add_env_groups": ENV_FILL,
        "expected_env_groups": ENV_FILL,
        "target_env_raw": ENV_FILL,
        "ENV_RAW": ENV_FILL,
        "ENV_SCOPE": ENV_FILL,
        "add_as_groups": AS_FILL,
        "expected_as_groups": AS_FILL,
        "AS": AS_FILL,
        "add_gas_groups": GAS_FILL,
        "expected_gas_groups": GAS_FILL,
        "GAS": GAS_FILL,
        "add_os_groups": OS_FILL,
        "expected_os_groups": OS_FILL,
        "OS_FAMILY": OS_FILL,
        "missing_in_zabbix": ALERT_FILL,
        "reference_status": ALERT_FILL,
        "manual_required": ALERT_FILL,
        "reason": ALERT_FILL,
        "unresolved_reasons": ALERT_FILL,
        "suggestion_status": ALERT_FILL,
        "implementation_status": INFO_FILL,
        "implementation_reason": INFO_FILL,
        "discovery_reason": ALERT_FILL,
        "scope_skip_reason": ALERT_FILL,
        "suggested_new_group": AS_FILL,
        "suggested_value": AS_FILL,
        "include_reason": INFO_FILL,
        "permission_status": INFO_FILL,
        "matching_tag_filters": INFO_FILL,
        "field_paths": INFO_FILL,
        "where_found": INFO_FILL,
        "reference_kind": INFO_FILL,
        "field_kind": INFO_FILL,
        "json_path": INFO_FILL,
        "flags": INFO_FILL,
        "proxy_name": INFO_FILL,
        "proxyid": INFO_FILL,
        "discovery_parent_hostid": INFO_FILL,
        "discovery_parent_itemid": INFO_FILL,
        "discovery_status": INFO_FILL,
        "discovery_disable_source": INFO_FILL,
        "discovery_ts_disable": INFO_FILL,
        "discovery_ts_delete": INFO_FILL,
        "dashboard_url": LINK_FILL,
        "panel_url": LINK_FILL,
    }
    for header, fill in fill_by_header.items():
        column_index = header_index.get(header)
        if not column_index:
            continue
        for row_index in range(1, ws.max_row + 1):
            ws.cell(row=row_index, column=column_index).fill = fill


def _finalize_table_sheet(ws, headers: Sequence[str]) -> None:
    _style_header_row(ws, 1, headers)
    _apply_kind_column_fills(ws, headers)
    ws.freeze_panes = "A2"
    if headers:
        ws.auto_filter.ref = ws.dimensions
    autosize_columns(ws)


def _finalize_summary_sheet(ws, title_rows: Iterable[int], header_rows: Iterable[tuple[int, Sequence[str]]]) -> None:
    for row_index in title_rows:
        _style_title_row(ws, row_index)
    for row_index, headers in header_rows:
        _style_header_row(ws, row_index, headers, SUMMARY_HEADER_FILL)
    ws.freeze_panes = "A2"
    autosize_columns(ws)


def _split_group_values(value: Any) -> List[str]:
    text = str(value or "").strip()
    if not text:
        return []
    return [item.strip() for item in text.split(",") if item.strip()]


def _split_right_group_names(value: Any) -> List[str]:
    items: List[str] = []
    for chunk in str(value or "").split(";"):
        text = chunk.strip()
        if not text:
            continue
        name, _, _ = text.partition(":")
        name = name.strip()
        if name:
            items.append(name)
    return items


def _merge_group_lists(*values: Any) -> str:
    groups = []
    for value in values:
        groups.extend(_split_group_values(value))
    return join_sorted(groups)


def _kind_bucket(group_name: str) -> str:
    parsed = parse_standard_group(group_name)
    if not parsed:
        return ""
    return str(parsed.get("root_kind") or "").upper()


def _split_groups_by_kind(group_names: Iterable[str]) -> Dict[str, str]:
    buckets: Dict[str, List[str]] = {"ENV": [], "AS": [], "GAS": [], "OS": []}
    for group_name in group_names:
        name = str(group_name or "").strip()
        if not name:
            continue
        bucket = _kind_bucket(name)
        if bucket in buckets:
            buckets[bucket].append(name)
    return {
        "ENV": join_sorted(buckets["ENV"]),
        "AS": join_sorted(buckets["AS"]),
        "GAS": join_sorted(buckets["GAS"]),
        "OS": join_sorted(buckets["OS"]),
    }


def _build_host_view_rows(rows: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for row in rows:
        current_groups = _merge_group_lists(row.get("old_groups"), row.get("standard_groups"), row.get("other_groups"))
        out.append(
            {
                "hostid": row.get("hostid", ""),
                "host": row.get("host", ""),
                "name": row.get("name", ""),
                "status_label": row.get("status_label", ""),
                "ORG": row.get("ORG", ""),
                "AS": row.get("AS", ""),
                "GAS": row.get("GAS", ""),
                "ENV_RAW": row.get("ENV_RAW", ""),
                "ENV_SCOPE": row.get("ENV_SCOPE", ""),
                "OS_FAMILY": row.get("OS_FAMILY", ""),
                "current_groups": current_groups,
            }
        )
    return out


def _build_excluded_host_rows(report: Dict[str, Any]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for row in report["discovery_hosts"]:
        rows.append(
            {
                "exclude_kind": "DISCOVERY",
                "hostid": row.get("hostid", ""),
                "host": row.get("host", ""),
                "name": row.get("name", ""),
                "status_label": row.get("status_label", ""),
                "ORG": row.get("ORG", ""),
                "AS": row.get("AS", ""),
                "GAS": row.get("GAS", ""),
                "ENV_RAW": row.get("ENV_RAW", ""),
                "ENV_SCOPE": row.get("ENV_SCOPE", ""),
                "OS_FAMILY": row.get("OS_FAMILY", ""),
                "current_groups": row.get("current_groups", ""),
                "reason": row.get("discovery_reason", ""),
            }
        )
    for row in report["unknown_hosts"]:
        rows.append(
            {
                "exclude_kind": "UNKNOWN",
                "hostid": row.get("hostid", ""),
                "host": row.get("host", ""),
                "name": row.get("name", ""),
                "status_label": row.get("status_label", ""),
                "ORG": row.get("ORG", ""),
                "AS": row.get("AS", ""),
                "GAS": row.get("GAS", ""),
                "ENV_RAW": row.get("ENV_RAW", ""),
                "ENV_SCOPE": row.get("ENV_SCOPE", ""),
                "OS_FAMILY": row.get("OS_FAMILY", ""),
                "current_groups": row.get("groups", ""),
                "reason": row.get("unknown_reasons", ""),
            }
        )
    for row in report["hosts_skipped_env"]:
        rows.append(
            {
                "exclude_kind": "ENV_SCOPE",
                "hostid": row.get("hostid", ""),
                "host": row.get("host", ""),
                "name": row.get("name", ""),
                "status_label": row.get("status_label", ""),
                "ORG": row.get("ORG", ""),
                "AS": row.get("AS", ""),
                "GAS": row.get("GAS", ""),
                "ENV_RAW": row.get("ENV_RAW", ""),
                "ENV_SCOPE": row.get("ENV_SCOPE", ""),
                "OS_FAMILY": row.get("OS_FAMILY", ""),
                "current_groups": row.get("current_groups", ""),
                "reason": row.get("skip_reason", ""),
            }
        )
    for row in report["hosts_skipped_gas"]:
        rows.append(
            {
                "exclude_kind": "GAS_SCOPE",
                "hostid": row.get("hostid", ""),
                "host": row.get("host", ""),
                "name": row.get("name", ""),
                "status_label": row.get("status_label", ""),
                "ORG": row.get("ORG", ""),
                "AS": row.get("AS", ""),
                "GAS": row.get("GAS", ""),
                "ENV_RAW": row.get("ENV_RAW", ""),
                "ENV_SCOPE": row.get("ENV_SCOPE", ""),
                "OS_FAMILY": row.get("OS_FAMILY", ""),
                "current_groups": row.get("current_groups", ""),
                "reason": row.get("skip_reason", ""),
            }
        )
    return rows


def _build_host_enrichment_view_rows(rows: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for row in rows:
        add_groups = _split_groups_by_kind(_split_group_values(row.get("host_missing_expected_groups")))
        out.append(
            {
                "hostid": row.get("hostid", ""),
                "host": row.get("host", ""),
                "name": row.get("name", ""),
                "status_label": row.get("status_label", ""),
                "ORG": row.get("ORG", ""),
                "AS": row.get("AS", ""),
                "GAS": row.get("GAS", ""),
                "ENV_RAW": row.get("ENV_RAW", ""),
                "ENV_SCOPE": row.get("ENV_SCOPE", ""),
                "OS_FAMILY": row.get("OS_FAMILY", ""),
                "current_groups": _merge_group_lists(row.get("old_groups"), row.get("standard_groups")),
                "add_env_groups": add_groups["ENV"],
                "add_as_groups": add_groups["AS"],
                "add_gas_groups": add_groups["GAS"],
                "add_os_groups": add_groups["OS"],
                "missing_in_zabbix": row.get("catalog_missing_groups", ""),
                "manual_required": row.get("manual_required", ""),
                "unresolved_reasons": row.get("unresolved_reasons", ""),
            }
        )
    return out


def _build_preview_index(rows: Sequence[Dict[str, Any]]) -> Dict[tuple[str, str], Dict[str, Any]]:
    index: Dict[tuple[str, str], Dict[str, Any]] = {}
    for row in rows:
        object_type = str(row.get("object_type") or "")
        object_id = str(row.get("object_id") or "")
        key = (object_type, object_id)
        entry = index.setdefault(
            key,
            {
                "object_type": object_type,
                "object_id": object_id,
                "object_name": str(row.get("object_name") or ""),
                "where_found": set(),
                "field_paths": set(),
                "reference_status": set(),
                "permission_status": set(),
                "current_groups": set(),
                "add_env_groups": set(),
                "add_as_groups": set(),
                "add_gas_groups": set(),
                "add_os_groups": set(),
                "missing_in_zabbix": set(),
                "include_reason": set(),
                "manual_required": False,
            },
        )
        old_group = str(row.get("old_group") or "").strip()
        if old_group:
            entry["current_groups"].add(old_group)
        where_found = str(row.get("where_found") or "").strip()
        if where_found:
            entry["where_found"].add(where_found)
        field_path = str(row.get("field_path") or "").strip()
        if field_path:
            entry["field_paths"].add(field_path)
        reference_status = str(row.get("reference_status") or "").strip()
        if reference_status:
            entry["reference_status"].add(reference_status)
        permission_status = str(row.get("permission_status") or "").strip()
        if permission_status:
            entry["permission_status"].add(permission_status)
        include_reason = str(row.get("include_reason") or "").strip()
        if include_reason:
            entry["include_reason"].add(include_reason)

        candidate_new = str(row.get("candidate_new_group") or "").strip()
        target_exists = str(row.get("target_exists") or "").strip().lower() == "yes"
        already_present = str(row.get("object_has_candidate_new") or "").strip().lower() == "yes"
        if candidate_new and already_present:
            entry["current_groups"].add(candidate_new)
        elif candidate_new and target_exists:
            bucket = str(row.get("target_kind") or "").strip().upper()
            if bucket.startswith("ENV"):
                entry["add_env_groups"].add(candidate_new)
            elif bucket.startswith("AS"):
                entry["add_as_groups"].add(candidate_new)
            elif bucket.startswith("GAS"):
                entry["add_gas_groups"].add(candidate_new)
            elif bucket.startswith("OS"):
                entry["add_os_groups"].add(candidate_new)
        elif candidate_new and not target_exists:
            entry["missing_in_zabbix"].add(candidate_new)

        if str(row.get("manual_required") or "").strip().lower() == "yes":
            entry["manual_required"] = True
    return index


def _preview_entry_to_row(entry: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "object_type": entry["object_type"],
        "object_id": entry["object_id"],
        "object_name": entry["object_name"],
        "where_found": join_sorted(entry["where_found"]),
        "field_paths": join_sorted(entry["field_paths"]),
        "reference_status": join_sorted(entry["reference_status"]),
        "permission_status": join_sorted(entry["permission_status"]),
        "current_groups": join_sorted(entry["current_groups"]),
        "add_env_groups": join_sorted(entry["add_env_groups"]),
        "add_as_groups": join_sorted(entry["add_as_groups"]),
        "add_gas_groups": join_sorted(entry["add_gas_groups"]),
        "add_os_groups": join_sorted(entry["add_os_groups"]),
        "missing_in_zabbix": join_sorted(entry["missing_in_zabbix"]),
        "include_reason": join_sorted(entry["include_reason"]),
        "manual_required": "yes" if entry["manual_required"] else "",
    }


def _build_preview_view_rows(rows: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    index = _build_preview_index(rows)
    return [_preview_entry_to_row(index[key]) for key in sorted(index)]


def _build_action_view_rows(action_rows: Sequence[Dict[str, Any]], preview_rows: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    preview_index = _build_preview_index(preview_rows)
    out: List[Dict[str, Any]] = []
    for row in action_rows:
        entry = preview_index.get(("action", str(row.get("actionid") or "")), {})
        out.append(
            {
                "actionid": row.get("actionid", ""),
                "name": row.get("name", ""),
                "status": row.get("status", ""),
                "where_found": row.get("where_found", ""),
                "reference_status": row.get("reference_status", join_sorted(entry.get("reference_status", []))),
                "current_groups": _merge_group_lists(row.get("matched_group_names"), row.get("candidate_new_group_names_present")),
                "add_env_groups": join_sorted(entry.get("add_env_groups", [])),
                "add_as_groups": join_sorted(entry.get("add_as_groups", [])),
                "add_gas_groups": join_sorted(entry.get("add_gas_groups", [])),
                "add_os_groups": join_sorted(entry.get("add_os_groups", [])),
                "missing_in_zabbix": join_sorted(entry.get("missing_in_zabbix", [])),
                "include_reason": row.get("include_reason", ""),
                "recipient_usergroups": row.get("recipient_usergroups", ""),
                "recipient_users": row.get("recipient_users", ""),
                "recipients_media": row.get("recipients_media", ""),
                "manual_required": "yes" if entry.get("manual_required") else "",
            }
        )
    return out


def _build_usergroup_view_rows(usergroup_rows: Sequence[Dict[str, Any]], preview_rows: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    preview_index = _build_preview_index(preview_rows)
    out: List[Dict[str, Any]] = []
    for row in usergroup_rows:
        entry = preview_index.get(("usergroup", str(row.get("usrgrpid") or "")), {})
        current_groups = join_sorted(
            _split_right_group_names(row.get("rights_on_old_groups")) + _split_right_group_names(row.get("rights_on_new_groups"))
        )
        out.append(
            {
                "usrgrpid": row.get("usrgrpid", ""),
                "name": row.get("name", ""),
                "current_groups": current_groups,
                "permission_status": row.get("permission_status", join_sorted(entry.get("permission_status", []))),
                "add_env_groups": join_sorted(entry.get("add_env_groups", [])),
                "add_as_groups": join_sorted(entry.get("add_as_groups", [])),
                "add_gas_groups": join_sorted(entry.get("add_gas_groups", [])),
                "add_os_groups": join_sorted(entry.get("add_os_groups", [])),
                "missing_in_zabbix": join_sorted(entry.get("missing_in_zabbix", [])),
                "matching_tag_filters": row.get("matching_tag_filters", ""),
                "include_reason": row.get("include_reason", ""),
                "users": row.get("users", ""),
                "users_media": row.get("users_media", ""),
                "is_action_recipient": row.get("is_action_recipient", ""),
                "manual_required": "yes" if entry.get("manual_required") else "",
            }
        )
    return out


def _build_maintenance_view_rows(maintenance_rows: Sequence[Dict[str, Any]], preview_rows: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    preview_index = _build_preview_index(preview_rows)
    out: List[Dict[str, Any]] = []
    for row in maintenance_rows:
        entry = preview_index.get(("maintenance", str(row.get("maintenanceid") or "")), {})
        out.append(
            {
                "maintenanceid": row.get("maintenanceid", ""),
                "name": row.get("name", ""),
                "reference_status": row.get("reference_status", join_sorted(entry.get("reference_status", []))),
                "current_groups": _merge_group_lists(row.get("matched_group_names"), row.get("candidate_new_group_names_present")),
                "add_env_groups": join_sorted(entry.get("add_env_groups", [])),
                "add_as_groups": join_sorted(entry.get("add_as_groups", [])),
                "add_gas_groups": join_sorted(entry.get("add_gas_groups", [])),
                "add_os_groups": join_sorted(entry.get("add_os_groups", [])),
                "missing_in_zabbix": join_sorted(entry.get("missing_in_zabbix", [])),
                "include_reason": row.get("include_reason", ""),
                "active_since": row.get("active_since", ""),
                "active_till": row.get("active_till", ""),
                "manual_required": "yes" if entry.get("manual_required") else "",
            }
        )
    return out


def save_inventory_json(report: Dict[str, Any], path: str) -> None:
    payload = {
        "meta": {
            "created_at": datetime.now().isoformat(timespec="seconds"),
            "version": "2.0",
        },
        "summary": report["summary"],
        "inventory": report["inventory"],
        "discovery_hosts": report["discovery_hosts"],
        "unknown_hosts": report["unknown_hosts"],
        "hosts": report["hosts"],
        "hosts_replace": report["hosts_replace"],
        "hosts_no_any_new": report["hosts_no_any_new"],
        "host_enrichment": report["host_enrichment"],
        "hosts_need_enrichment": report["hosts_need_enrichment"],
        "hosts_clean": report["hosts_clean"],
        "hosts_skipped_env": report["hosts_skipped_env"],
        "hosts_skipped_gas": report["hosts_skipped_gas"],
        "mismatch_host_oldorg": report["mismatch_host_oldorg"],
        "mismatch_host_proxyorg": report["mismatch_host_proxyorg"],
        "mismatch_legacy_env": report["mismatch_legacy_env"],
        "host_expected_groups": report["host_expected_groups"],
        "env_summary": report["env_summary"],
        "gas_summary": report["gas_summary"],
        "guest_name_summary": report["guest_name_summary"],
        "groups_old": report["groups_old"],
        "groups_new": report["groups_new"],
        "expected_groups": report["expected_groups"],
        "mapping_plan": report["mapping_plan"],
        "zabbix_mapping_preview": report["zabbix_mapping_preview"],
        "actions": report["actions"],
        "usergroups": report["usergroups"],
        "maintenances": report["maintenances"],
        "grafana": report["grafana"],
        "grafana_summary": report["grafana_summary"],
        "grafana_variables": report.get("grafana_variables", []),
        "grafana_panels": report.get("grafana_panels", []),
        "grafana_suggestions": report.get("grafana_suggestions", []),
    }
    with open(path, "w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2)


def write_workbook(report: Dict[str, Any], out_path: str) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    summary_ws = wb.create_sheet("SUMMARY")
    summary_titles: List[int] = []
    summary_headers: List[tuple[int, Sequence[str]]] = []
    summary_ws.append(["key", "value"])
    for key, value in report["summary"].items():
        if isinstance(value, list):
            rendered = ", ".join(str(item) for item in value)
        else:
            rendered = value
        summary_ws.append([key, rendered])
    _style_header_row(summary_ws, 1, ["key", "value"], SUMMARY_HEADER_FILL)
    summary_ws.append([])
    title_row = summary_ws.max_row + 1
    _append_titled_table(
        summary_ws,
        "ENV SUMMARY",
        report["env_summary"],
        ["AS", "ENV_RAW", "ENV_SCOPE", "hosts_count", "enabled_hosts", "disabled_hosts", "legacy_hosts", "sample_hosts"],
    )
    summary_titles.append(title_row)
    summary_headers.append((title_row + 1, ["AS", "ENV_RAW", "ENV_SCOPE", "hosts_count", "enabled_hosts", "disabled_hosts", "legacy_hosts", "sample_hosts"]))
    title_row = summary_ws.max_row + 1
    _append_titled_table(
        summary_ws,
        "GAS SUMMARY",
        report["gas_summary"],
        ["AS", "GAS", "hosts_count", "enabled_hosts", "disabled_hosts", "legacy_hosts", "sample_hosts"],
    )
    summary_titles.append(title_row)
    summary_headers.append((title_row + 1, ["AS", "GAS", "hosts_count", "enabled_hosts", "disabled_hosts", "legacy_hosts", "sample_hosts"]))
    title_row = summary_ws.max_row + 1
    _append_titled_table(
        summary_ws,
        "GUEST-NAME SUMMARY",
        report["guest_name_summary"],
        ["AS", "GUEST_NAME", "hosts_count", "enabled_hosts", "disabled_hosts", "legacy_hosts", "sample_hosts"],
    )
    summary_titles.append(title_row)
    summary_headers.append((title_row + 1, ["AS", "GUEST_NAME", "hosts_count", "enabled_hosts", "disabled_hosts", "legacy_hosts", "sample_hosts"]))
    _finalize_summary_sheet(summary_ws, summary_titles, summary_headers)

    hosts_ws = wb.create_sheet("HOSTS")
    _append_rows(hosts_ws, _build_host_view_rows(report["hosts"]), HOST_VIEW_HEADERS)
    _finalize_table_sheet(hosts_ws, HOST_VIEW_HEADERS)

    replace_ws = wb.create_sheet("HOSTS_OLD_SCOPE")
    _append_rows(replace_ws, _build_host_view_rows(report["hosts_replace"]), HOST_VIEW_HEADERS)
    _finalize_table_sheet(replace_ws, HOST_VIEW_HEADERS)

    replace_missing_ws = wb.create_sheet("HOSTS_NO_ANY_NEW")
    _append_rows(replace_missing_ws, _build_host_view_rows(report["hosts_no_any_new"]), HOST_VIEW_HEADERS)
    _finalize_table_sheet(replace_missing_ws, HOST_VIEW_HEADERS)

    clean_ws = wb.create_sheet("HOSTS_CLEAN")
    _append_rows(clean_ws, _build_host_view_rows(report["hosts_clean"]), HOST_VIEW_HEADERS)
    _finalize_table_sheet(clean_ws, HOST_VIEW_HEADERS)

    excluded_ws = wb.create_sheet("HOSTS_EXCLUDED")
    _append_rows(excluded_ws, _build_excluded_host_rows(report), EXCLUDED_HOST_HEADERS)
    _finalize_table_sheet(excluded_ws, EXCLUDED_HOST_HEADERS)

    discovery_ws = wb.create_sheet("HOSTS_DISCOVERY")
    _append_rows(discovery_ws, report["discovery_hosts"], DISCOVERY_HOST_HEADERS)
    _finalize_table_sheet(discovery_ws, DISCOVERY_HOST_HEADERS)

    mismatch_ws = wb.create_sheet("MISMATCHES")
    mismatch_titles: List[int] = []
    mismatch_headers: List[tuple[int, Sequence[str]]] = []
    title_row = mismatch_ws.max_row + 1
    _append_titled_table(
        mismatch_ws,
        "HOST DOMAIN ORG != OLD GROUP ORG",
        report["mismatch_host_oldorg"],
        MISMATCH_OLDORG_HEADERS,
    )
    mismatch_titles.append(title_row)
    mismatch_headers.append((title_row + 1, MISMATCH_OLDORG_HEADERS))
    title_row = mismatch_ws.max_row + 1
    _append_titled_table(
        mismatch_ws,
        "HOST DOMAIN ORG != PROXY ORG",
        report["mismatch_host_proxyorg"],
        MISMATCH_PROXY_HEADERS,
    )
    mismatch_titles.append(title_row)
    mismatch_headers.append((title_row + 1, MISMATCH_PROXY_HEADERS))
    title_row = mismatch_ws.max_row + 1
    _append_titled_table(
        mismatch_ws,
        "OLD GROUP ENV != TAG ENV",
        report["mismatch_legacy_env"],
        MISMATCH_LEGACY_ENV_HEADERS,
    )
    mismatch_titles.append(title_row)
    mismatch_headers.append((title_row + 1, MISMATCH_LEGACY_ENV_HEADERS))
    _finalize_summary_sheet(mismatch_ws, mismatch_titles, mismatch_headers)

    old_ws = wb.create_sheet("GROUPS_OLD")
    _append_rows(
        old_ws,
        report["groups_old"],
        ["group_name", "groupid", "legacy_env_tokens", "org_values", "as_values", "env_raw_values", "env_scope_values", "hosts_count", "sample_hosts"],
    )
    _finalize_table_sheet(old_ws, ["group_name", "groupid", "legacy_env_tokens", "org_values", "as_values", "env_raw_values", "env_scope_values", "hosts_count", "sample_hosts"])

    new_ws = wb.create_sheet("GROUPS_NEW")
    groups_new_headers = [
        "group_name",
        "groupid",
        "group_kind",
        "org",
        "as_values",
        "env_raw_values",
        "env_scope_values",
        "gas_values",
        "os_families",
        "hosts_count",
        "sample_hosts",
    ]
    _append_rows(
        new_ws,
        report["groups_new"],
        groups_new_headers,
    )
    _finalize_table_sheet(new_ws, groups_new_headers)

    inventory_ws = wb.create_sheet("INVENTORY")
    inventory_ws.append(["section", "value"])
    for key, value in report["inventory"].items():
        inventory_ws.append([key, json.dumps(value, ensure_ascii=False)])
    _style_header_row(inventory_ws, 1, ["section", "value"], SUMMARY_HEADER_FILL)
    inventory_ws.freeze_panes = "A2"
    autosize_columns(inventory_ws)

    wb.save(out_path)


def write_mapping_workbook(report: Dict[str, Any], out_path: str) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    summary_ws = wb.create_sheet("SUMMARY")
    summary_ws.append(["key", "value"])
    for key, value in report["summary"].items():
        if isinstance(value, list):
            rendered = ", ".join(str(item) for item in value)
        else:
            rendered = value
        summary_ws.append([key, rendered])
    _style_header_row(summary_ws, 1, ["key", "value"], SUMMARY_HEADER_FILL)
    summary_ws.freeze_panes = "A2"
    autosize_columns(summary_ws)

    mapping_ws = wb.create_sheet("MAPPING_PLAN")
    _append_rows(mapping_ws, report["mapping_plan"], MAPPING_PLAN_HEADERS)
    _finalize_table_sheet(mapping_ws, MAPPING_PLAN_HEADERS)

    expected_ws = wb.create_sheet("EXPECTED_GROUPS")
    _append_rows(expected_ws, report["expected_groups"], EXPECTED_GROUP_HEADERS)
    _finalize_table_sheet(expected_ws, EXPECTED_GROUP_HEADERS)

    expected_host_ws = wb.create_sheet("HOST_EXPECTED")
    _append_rows(expected_host_ws, report["host_expected_groups"], HOST_EXPECTED_HEADERS)
    _finalize_table_sheet(expected_host_ws, HOST_EXPECTED_HEADERS)

    enrichment_ws = wb.create_sheet("HOST_ENRICHMENT")
    enrichment_rows = _build_host_enrichment_view_rows(report["host_enrichment"])
    _append_rows(enrichment_ws, enrichment_rows, HOST_VIEW_HEADERS)
    _finalize_table_sheet(enrichment_ws, HOST_VIEW_HEADERS)

    need_enrichment_ws = wb.create_sheet("HOSTS_NEED_ENRICH")
    need_enrichment_rows = _build_host_enrichment_view_rows(report["hosts_need_enrichment"])
    _append_rows(need_enrichment_ws, need_enrichment_rows, HOST_VIEW_HEADERS)
    _finalize_table_sheet(need_enrichment_ws, HOST_VIEW_HEADERS)

    preview_ws = wb.create_sheet("ZBX_MAP_PREVIEW")
    preview_rows = _build_preview_view_rows(report["zabbix_mapping_preview"])
    _append_rows(preview_ws, preview_rows, OBJECT_PREVIEW_HEADERS)
    _finalize_table_sheet(preview_ws, OBJECT_PREVIEW_HEADERS)

    actions_ws = wb.create_sheet("ACTIONS")
    action_rows = _build_action_view_rows(report["actions"], report["zabbix_mapping_preview"])
    _append_rows(actions_ws, action_rows, ACTION_VIEW_HEADERS)
    _finalize_table_sheet(actions_ws, ACTION_VIEW_HEADERS)

    usergroups_ws = wb.create_sheet("USERGROUPS")
    usergroup_rows = _build_usergroup_view_rows(report["usergroups"], report["zabbix_mapping_preview"])
    _append_rows(usergroups_ws, usergroup_rows, USERGROUP_VIEW_HEADERS)
    _finalize_table_sheet(usergroups_ws, USERGROUP_VIEW_HEADERS)

    maintenances_ws = wb.create_sheet("MAINTENANCES")
    maintenance_rows = _build_maintenance_view_rows(report["maintenances"], report["zabbix_mapping_preview"])
    _append_rows(maintenances_ws, maintenance_rows, MAINTENANCE_VIEW_HEADERS)
    _finalize_table_sheet(maintenances_ws, MAINTENANCE_VIEW_HEADERS)

    grafana_summary_ws = wb.create_sheet("GRAFANA_SUMMARY")
    _append_rows(grafana_summary_ws, report["grafana_summary"], GRAFANA_SUMMARY_HEADERS)
    _apply_hyperlinks(grafana_summary_ws, report["grafana_summary"], GRAFANA_SUMMARY_HEADERS, {"dashboard_url": "dashboard_url"})
    _finalize_table_sheet(grafana_summary_ws, GRAFANA_SUMMARY_HEADERS)

    grafana_variables_ws = wb.create_sheet("GRAFANA_VARIABLES")
    _append_rows(grafana_variables_ws, report.get("grafana_variables", []), GRAFANA_DETAIL_HEADERS)
    _apply_hyperlinks(
        grafana_variables_ws,
        report.get("grafana_variables", []),
        GRAFANA_DETAIL_HEADERS,
        {
            "dashboard_url": "dashboard_url",
            "panel_url": "panel_url",
        },
    )
    _finalize_table_sheet(grafana_variables_ws, GRAFANA_DETAIL_HEADERS)

    grafana_panels_ws = wb.create_sheet("GRAFANA_PANELS")
    _append_rows(grafana_panels_ws, report.get("grafana_panels", []), GRAFANA_DETAIL_HEADERS)
    _apply_hyperlinks(
        grafana_panels_ws,
        report.get("grafana_panels", []),
        GRAFANA_DETAIL_HEADERS,
        {
            "dashboard_url": "dashboard_url",
            "panel_url": "panel_url",
        },
    )
    _finalize_table_sheet(grafana_panels_ws, GRAFANA_DETAIL_HEADERS)

    grafana_suggestions_ws = wb.create_sheet("GRAFANA_SUGGESTIONS")
    _append_rows(grafana_suggestions_ws, report.get("grafana_suggestions", []), GRAFANA_SUGGESTION_HEADERS)
    _apply_hyperlinks(
        grafana_suggestions_ws,
        report.get("grafana_suggestions", []),
        GRAFANA_SUGGESTION_HEADERS,
        {
            "dashboard_url": "dashboard_url",
            "panel_url": "panel_url",
        },
    )
    _finalize_table_sheet(grafana_suggestions_ws, GRAFANA_SUGGESTION_HEADERS)

    wb.save(out_path)


def write_grafana_workbook(report: Dict[str, Any], out_path: str) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    summary_rows = [
        {"key": "scope_as", "value": ", ".join(str(item) for item in report["summary"].get("scope_as") or [])},
        {"key": "scope_env", "value": str(report["summary"].get("scope_env") or "")},
        {"key": "scope_gas", "value": ", ".join(str(item) for item in report["summary"].get("scope_gas") or [])},
        {"key": "grafana_dashboards", "value": report["summary"].get("grafana_dashboards", len(report["grafana_summary"]))},
        {"key": "grafana_rows", "value": report["summary"].get("grafana_rows", len(report["grafana"]))},
        {"key": "grafana_variable_rows", "value": report["summary"].get("grafana_variable_rows", len(report.get("grafana_variables", [])))},
        {"key": "grafana_panel_rows", "value": report["summary"].get("grafana_panel_rows", len(report.get("grafana_panels", [])))},
        {"key": "grafana_suggestion_rows", "value": report["summary"].get("grafana_suggestion_rows", len(report.get("grafana_suggestions", [])))},
        {"key": "grafana_error", "value": report["summary"].get("grafana_error", "")},
    ]
    summary_ws = wb.create_sheet("SUMMARY")
    _append_rows(summary_ws, summary_rows, ["key", "value"])
    _finalize_table_sheet(summary_ws, ["key", "value"])

    dashboards_ws = wb.create_sheet("DASHBOARDS")
    _append_rows(dashboards_ws, report["grafana_summary"], GRAFANA_SUMMARY_HEADERS)
    _apply_hyperlinks(dashboards_ws, report["grafana_summary"], GRAFANA_SUMMARY_HEADERS, {"dashboard_url": "dashboard_url"})
    _finalize_table_sheet(dashboards_ws, GRAFANA_SUMMARY_HEADERS)

    details_ws = wb.create_sheet("DETAILS")
    _append_rows(details_ws, report["grafana"], GRAFANA_DETAIL_HEADERS)
    _apply_hyperlinks(
        details_ws,
        report["grafana"],
        GRAFANA_DETAIL_HEADERS,
        {
            "dashboard_url": "dashboard_url",
            "panel_url": "panel_url",
        },
    )
    _finalize_table_sheet(details_ws, GRAFANA_DETAIL_HEADERS)

    variables_ws = wb.create_sheet("VARIABLES")
    _append_rows(variables_ws, report.get("grafana_variables", []), GRAFANA_DETAIL_HEADERS)
    _apply_hyperlinks(
        variables_ws,
        report.get("grafana_variables", []),
        GRAFANA_DETAIL_HEADERS,
        {
            "dashboard_url": "dashboard_url",
            "panel_url": "panel_url",
        },
    )
    _finalize_table_sheet(variables_ws, GRAFANA_DETAIL_HEADERS)

    panels_ws = wb.create_sheet("PANELS")
    _append_rows(panels_ws, report.get("grafana_panels", []), GRAFANA_DETAIL_HEADERS)
    _apply_hyperlinks(
        panels_ws,
        report.get("grafana_panels", []),
        GRAFANA_DETAIL_HEADERS,
        {
            "dashboard_url": "dashboard_url",
            "panel_url": "panel_url",
        },
    )
    _finalize_table_sheet(panels_ws, GRAFANA_DETAIL_HEADERS)

    suggestions_ws = wb.create_sheet("SUGGESTIONS")
    _append_rows(suggestions_ws, report.get("grafana_suggestions", []), GRAFANA_SUGGESTION_HEADERS)
    _apply_hyperlinks(
        suggestions_ws,
        report.get("grafana_suggestions", []),
        GRAFANA_SUGGESTION_HEADERS,
        {
            "dashboard_url": "dashboard_url",
            "panel_url": "panel_url",
        },
    )
    _finalize_table_sheet(suggestions_ws, GRAFANA_SUGGESTION_HEADERS)

    wb.save(out_path)


def save_grafana_org_json(report: Dict[str, Any], path: str) -> None:
    payload = {
        "meta": {
            "created_at": datetime.now().isoformat(timespec="seconds"),
            "version": "2.0",
        },
        "summary": report["summary"],
        "org_summary": report["org_summary"],
        "datasources": report["datasources"],
        "dashboards": report["dashboards"],
        "variables": report["variables"],
        "panels": report["panels"],
        "details": report["details"],
        "suggestions": report.get("suggestions", []),
        "errors": report.get("errors", []),
    }
    with open(path, "w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2)


def write_grafana_org_workbook(report: Dict[str, Any], out_path: str) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    summary_ws = wb.create_sheet("SUMMARY")
    summary_ws.append(["key", "value"])
    for key, value in report["summary"].items():
        if isinstance(value, list):
            rendered = ", ".join(str(item) for item in value)
        else:
            rendered = value
        summary_ws.append([key, rendered])
    _style_header_row(summary_ws, 1, ["key", "value"], SUMMARY_HEADER_FILL)
    summary_ws.freeze_panes = "A2"
    autosize_columns(summary_ws)

    org_ws = wb.create_sheet("ORGS")
    _append_rows(org_ws, report["org_summary"], GRAFANA_ORG_SUMMARY_HEADERS)
    _finalize_table_sheet(org_ws, GRAFANA_ORG_SUMMARY_HEADERS)

    datasources_ws = wb.create_sheet("DATASOURCES")
    _append_rows(datasources_ws, report["datasources"], GRAFANA_ORG_DATASOURCE_HEADERS)
    _finalize_table_sheet(datasources_ws, GRAFANA_ORG_DATASOURCE_HEADERS)

    dashboards_ws = wb.create_sheet("DASHBOARDS")
    _append_rows(dashboards_ws, report["dashboards"], GRAFANA_ORG_DASHBOARD_HEADERS)
    _apply_hyperlinks(dashboards_ws, report["dashboards"], GRAFANA_ORG_DASHBOARD_HEADERS, {"dashboard_url": "dashboard_url"})
    _finalize_table_sheet(dashboards_ws, GRAFANA_ORG_DASHBOARD_HEADERS)

    variables_ws = wb.create_sheet("VARIABLES")
    _append_rows(variables_ws, report["variables"], GRAFANA_ORG_VARIABLE_HEADERS)
    _apply_hyperlinks(variables_ws, report["variables"], GRAFANA_ORG_VARIABLE_HEADERS, {"dashboard_url": "dashboard_url"})
    _finalize_table_sheet(variables_ws, GRAFANA_ORG_VARIABLE_HEADERS)

    panels_ws = wb.create_sheet("PANELS")
    _append_rows(panels_ws, report["panels"], GRAFANA_ORG_PANEL_HEADERS)
    _apply_hyperlinks(
        panels_ws,
        report["panels"],
        GRAFANA_ORG_PANEL_HEADERS,
        {
            "dashboard_url": "dashboard_url",
            "panel_url": "panel_url",
        },
    )
    _finalize_table_sheet(panels_ws, GRAFANA_ORG_PANEL_HEADERS)

    details_ws = wb.create_sheet("DETAILS")
    _append_rows(details_ws, report["details"], GRAFANA_ORG_DETAIL_HEADERS)
    _apply_hyperlinks(
        details_ws,
        report["details"],
        GRAFANA_ORG_DETAIL_HEADERS,
        {
            "dashboard_url": "dashboard_url",
            "panel_url": "panel_url",
        },
    )
    _finalize_table_sheet(details_ws, GRAFANA_ORG_DETAIL_HEADERS)

    suggestions_ws = wb.create_sheet("SUGGESTIONS")
    _append_rows(suggestions_ws, report.get("suggestions", []), GRAFANA_ORG_SUGGESTION_HEADERS)
    _apply_hyperlinks(
        suggestions_ws,
        report.get("suggestions", []),
        GRAFANA_ORG_SUGGESTION_HEADERS,
        {
            "dashboard_url": "dashboard_url",
            "panel_url": "panel_url",
        },
    )
    _finalize_table_sheet(suggestions_ws, GRAFANA_ORG_SUGGESTION_HEADERS)

    errors_ws = wb.create_sheet("ERRORS")
    _append_rows(errors_ws, report.get("errors", []), GRAFANA_ORG_ERROR_HEADERS)
    _apply_hyperlinks(errors_ws, report.get("errors", []), GRAFANA_ORG_ERROR_HEADERS, {"dashboard_url": "dashboard_url"})
    _finalize_table_sheet(errors_ws, GRAFANA_ORG_ERROR_HEADERS)

    wb.save(out_path)
