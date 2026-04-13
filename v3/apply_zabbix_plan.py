#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Применение Zabbix impact plan: host enrichment + безопасные object changes."""

from __future__ import annotations

import argparse
import json
import re
from datetime import datetime
from typing import Any, Dict, List, Sequence, Set, Tuple

from openpyxl import Workbook  # type: ignore

import config
from api_clients import ZabbixAPI
from backup_io import load_backup
from common import (
    autosize_columns,
    build_artifact_path,
    join_sorted,
    normalize_scope_env,
    normalize_values,
    resolve_input_artifact,
)
from impact_plan import fetch_actions, fetch_maintenances, fetch_usergroups


_PATH_TOKEN_RX = re.compile(r"([^\.\[]+)|\[(\d+)\]")
_AUTO_OBJECT_CHANGE_KINDS = {"replace_groupid", "add_group_permission"}


def load_impact_plan(path: str) -> Dict[str, Any]:
    with open(path, "r", encoding="utf-8") as handle:
        return json.load(handle)


def fetch_hosts(api: ZabbixAPI, hostids: Sequence[str]) -> List[Dict[str, Any]]:
    if not hostids:
        return []
    return api.call(
        "host.get",
        {
            "output": ["hostid", "host", "name", "status"],
            "hostids": list(hostids),
            "selectGroups": ["groupid", "name"],
        },
    )


def _parse_path(path: str) -> List[Any]:
    tokens: List[Any] = []
    for part in str(path or "").split("."):
        if not part:
            continue
        for match in _PATH_TOKEN_RX.finditer(part):
            key, index = match.groups()
            if key is not None:
                tokens.append(key)
            elif index is not None:
                tokens.append(int(index))
    return tokens


def _get_path_value(node: Any, path: str) -> Any:
    current = node
    for token in _parse_path(path):
        current = current[token]
    return current


def _set_path_value(node: Any, path: str, value: Any) -> None:
    tokens = _parse_path(path)
    if not tokens:
        raise RuntimeError(f"Invalid path: {path}")
    current = node
    for token in tokens[:-1]:
        current = current[token]
    tail = tokens[-1]
    current[tail] = value



def _top_level_section(path: str) -> str:
    head = str(path or "").split(".", 1)[0]
    return head.split("[", 1)[0]



def _auto_object_rows(impact_plan: Dict[str, Any]) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for row in impact_plan.get("object_mapping_plan") or []:
        change_kind = str(row.get("change_kind") or "").strip()
        manual_required = str(row.get("manual_required") or "").strip().lower() == "yes"
        if manual_required:
            continue
        if change_kind in _AUTO_OBJECT_CHANGE_KINDS:
            out.append(row)
    return out



def _validate_backup_scope(backup_path: str, impact_plan: Dict[str, Any], api: ZabbixAPI) -> None:
    data = load_backup(backup_path)
    summary = impact_plan.get("summary") or {}
    impact_scope_as = normalize_values(summary.get("scope_as") or [])
    impact_scope_env = normalize_scope_env(summary.get("scope_env") or "")
    impact_scope_gas = normalize_values(summary.get("scope_gas") or [])

    if str(data.meta.zabbix_url or "").strip() and str(data.meta.zabbix_url).strip() != str(getattr(api, "api_url", "")).strip():
        raise RuntimeError(
            f"Backup Zabbix URL mismatch: backup={data.meta.zabbix_url} current={getattr(api, 'api_url', '')}"
        )

    backup_scope_as = normalize_values(data.meta.scope_as)
    if sorted(item.lower() for item in backup_scope_as) != sorted(item.lower() for item in impact_scope_as):
        raise RuntimeError(f"Backup scope_as mismatch: backup={backup_scope_as} impact={impact_scope_as}")

    backup_scope_env = normalize_scope_env(data.meta.scope_env)
    if str(backup_scope_env) != str(impact_scope_env):
        raise RuntimeError(f"Backup scope_env mismatch: backup={backup_scope_env} impact={impact_scope_env}")

    backup_scope_gas = normalize_values(data.meta.scope_gas)
    if sorted(item.upper() for item in backup_scope_gas) != sorted(item.upper() for item in impact_scope_gas):
        raise RuntimeError(f"Backup scope_gas mismatch: backup={backup_scope_gas} impact={impact_scope_gas}")

    auto_rows = _auto_object_rows(impact_plan)
    impact_hostids = {
        str(row.get("object_id") or "").strip()
        for row in impact_plan.get("host_enrich_plan") or []
        if str(row.get("object_id") or "").strip()
    }
    impact_actionids = {str(row.get("object_id") or "").strip() for row in auto_rows if str(row.get("object_type") or "") == "action"}
    impact_usergroupids = {str(row.get("object_id") or "").strip() for row in auto_rows if str(row.get("object_type") or "") == "usergroup"}
    impact_maintenanceids = {str(row.get("object_id") or "").strip() for row in auto_rows if str(row.get("object_type") or "") == "maintenance"}

    backup_hostids = {str(item.hostid).strip() for item in data.hosts if str(item.hostid).strip()}
    backup_actionids = {str(item.actionid).strip() for item in data.actions if str(item.actionid).strip()}
    backup_usergroupids = {str(item.usrgrpid).strip() for item in data.usergroups if str(item.usrgrpid).strip()}
    backup_maintenanceids = {str(item.maintenanceid).strip() for item in data.maintenances if str(item.maintenanceid).strip()}

    missing_hostids = sorted(impact_hostids.difference(backup_hostids))
    missing_actionids = sorted(impact_actionids.difference(backup_actionids))
    missing_usergroupids = sorted(impact_usergroupids.difference(backup_usergroupids))
    missing_maintenanceids = sorted(impact_maintenanceids.difference(backup_maintenanceids))

    errors: List[str] = []
    if missing_hostids:
        errors.append(f"hosts: {', '.join(missing_hostids)}")
    if missing_actionids:
        errors.append(f"actions: {', '.join(missing_actionids)}")
    if missing_usergroupids:
        errors.append(f"usergroups: {', '.join(missing_usergroupids)}")
    if missing_maintenanceids:
        errors.append(f"maintenances: {', '.join(missing_maintenanceids)}")
    if errors:
        raise RuntimeError(f"Backup coverage error: {'; '.join(errors)}")



def _group_host_rows(impact_plan: Dict[str, Any]) -> Dict[str, Dict[str, Any]]:
    grouped: Dict[str, Dict[str, Any]] = {}
    for row in impact_plan.get("host_enrich_plan") or []:
        if str(row.get("change_kind") or "") != "add_group":
            continue
        hostid = str(row.get("object_id") or "").strip()
        groupid = str(row.get("new_groupid") or "").strip()
        group_name = str(row.get("new_group") or "").strip()
        if not hostid or not groupid or not group_name:
            continue

        bucket = grouped.setdefault(
            hostid,
            {
                "hostid": hostid,
                "host_name": str(row.get("object_name") or "").strip(),
                "groups": {},
                "details": {},
            },
        )
        bucket["groups"][groupid] = group_name
        detail_key = str(group_name).strip()
        if detail_key:
            bucket["details"][detail_key] = str(row.get("details") or "").strip()
    return grouped



def _assert_method_available(api: ZabbixAPI, method: str, params: Dict[str, Any] | None = None) -> None:
    try:
        api.call(method, params or {})
    except RuntimeError as exc:
        message = str(exc)
        if f'No permissions to call "{method}"' in message:
            raise RuntimeError(f'Zabbix API denies "{method}" for this user/role. Check role API permissions for {method}.') from exc
        if f"Zabbix API error ({method}):" in message:
            return
        raise


def _required_methods_for_impact_plan(impact_plan: Dict[str, Any]) -> Set[str]:
    auto_rows = _auto_object_rows(impact_plan)
    methods_to_check: Set[str] = set()
    if any(str(row.get("object_id") or "").strip() for row in impact_plan.get("host_enrich_plan") or []):
        methods_to_check.add("host.massadd")
    if any(str(row.get("object_type") or "") == "action" for row in auto_rows):
        methods_to_check.add("action.update")
    if any(str(row.get("object_type") or "") == "usergroup" for row in auto_rows):
        methods_to_check.add("usergroup.update")
    if any(str(row.get("object_type") or "") == "maintenance" for row in auto_rows):
        methods_to_check.add("maintenance.update")
    return methods_to_check


def prepare_zabbix_apply(
    api: ZabbixAPI,
    impact_plan: Dict[str, Any],
    backup_path: str = "",
    *,
    dry_run: bool,
    log=print,
) -> None:
    if dry_run:
        return
    for method in sorted(_required_methods_for_impact_plan(impact_plan)):
        log(f"Checking Zabbix API permissions for {method}")
        params = {"hosts": [], "groups": []} if method == "host.massadd" else {}
        _assert_method_available(api, method, params)
    if backup_path:
        log(f"Validating backup against impact plan: {backup_path}")
        _validate_backup_scope(backup_path, impact_plan, api)



def _build_skipped_rows(impact_plan: Dict[str, Any]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for row in impact_plan.get("object_mapping_plan") or []:
        object_type = str(row.get("object_type") or "").strip()
        change_kind = str(row.get("change_kind") or "").strip()
        manual_required = str(row.get("manual_required") or "").strip().lower() == "yes"
        if manual_required:
            reason = "manual_required"
        elif change_kind in _AUTO_OBJECT_CHANGE_KINDS:
            continue
        else:
            reason = f"{object_type}:{change_kind or 'unsupported'}"
        rows.append(
            {
                "object_type": object_type,
                "object_id": str(row.get("object_id") or "").strip(),
                "object_name": str(row.get("object_name") or "").strip(),
                "field_path": str(row.get("field_path") or "").strip(),
                "change_kind": change_kind,
                "old_group": str(row.get("old_group") or "").strip(),
                "new_group": str(row.get("new_group") or "").strip(),
                "reason": reason,
                "details": str(row.get("details") or "").strip(),
            }
        )
    return rows



def _batch_massadd_rows(pending_rows: Sequence[Dict[str, Any]]) -> Dict[Tuple[str, ...], List[Dict[str, Any]]]:
    batches: Dict[Tuple[str, ...], List[Dict[str, Any]]] = {}
    for row in pending_rows:
        key = tuple(sorted(str(item).strip() for item in (row.get("add_groupids") or []) if str(item).strip()))
        if not key:
            continue
        batches.setdefault(key, []).append(row)
    return batches



def _normalize_right(group: Dict[str, Any]) -> Dict[str, str]:
    groupid = str(group.get("groupid") or group.get("id") or group.get("hostgroupid") or "").strip()
    permission = str(group.get("permission") or "")
    if not groupid:
        return {}
    return {"id": groupid, "permission": permission}



def apply_host_enrichment(api: ZabbixAPI, impact_plan: Dict[str, Any], dry_run: bool) -> Dict[str, Any]:
    grouped = _group_host_rows(impact_plan)
    hostids = sorted(grouped)
    live_hosts = fetch_hosts(api, hostids)
    live_by_id = {str(row.get("hostid") or "").strip(): row for row in live_hosts if str(row.get("hostid") or "").strip()}

    results: List[Dict[str, Any]] = []
    pending_rows: List[Dict[str, Any]] = []
    applied_hosts = 0
    changed_hosts = 0
    already_hosts = 0
    missing_hosts = 0
    applied_batches = 0

    for hostid in hostids:
        plan_row = grouped[hostid]
        host = live_by_id.get(hostid)
        if not host:
            missing_hosts += 1
            results.append(
                {
                    "hostid": hostid,
                    "host": "",
                    "name": plan_row.get("host_name") or "",
                    "status": "host_not_found",
                    "mode": "DRY-RUN" if dry_run else "APPLY",
                    "current_groups": "",
                    "planned_groups": join_sorted(plan_row["groups"].values()),
                    "added_groups": "",
                    "details": "host.get did not return this hostid",
                }
            )
            continue

        current_groups = host.get("groups") or []
        current_groupids = {
            str(item.get("groupid") or "").strip()
            for item in current_groups
            if str(item.get("groupid") or "").strip()
        }
        current_group_names = [
            str(item.get("name") or "").strip()
            for item in current_groups
            if str(item.get("name") or "").strip()
        ]
        planned_groupids = sorted(plan_row["groups"])
        add_groupids = [groupid for groupid in planned_groupids if groupid not in current_groupids]
        add_group_names = [plan_row["groups"][groupid] for groupid in add_groupids]

        if not add_groupids:
            already_hosts += 1
            results.append(
                {
                    "hostid": hostid,
                    "host": str(host.get("host") or "").strip(),
                    "name": str(host.get("name") or "").strip(),
                    "status": "already_present",
                    "mode": "DRY-RUN" if dry_run else "APPLY",
                    "current_groups": join_sorted(current_group_names),
                    "planned_groups": join_sorted(plan_row["groups"].values()),
                    "added_groups": "",
                    "details": join_sorted(plan_row["details"].values()),
                }
            )
            continue

        pending_rows.append(
            {
                "hostid": hostid,
                "host": str(host.get("host") or "").strip(),
                "name": str(host.get("name") or "").strip(),
                "status": "dry_run_add" if dry_run else "pending_apply",
                "mode": "DRY-RUN" if dry_run else "APPLY",
                "current_groups": join_sorted(current_group_names),
                "planned_groups": join_sorted(plan_row["groups"].values()),
                "added_groups": join_sorted(add_group_names),
                "details": join_sorted(plan_row["details"].values()),
                "add_groupids": sorted(add_groupids),
            }
        )
        changed_hosts += 1

    if not dry_run and pending_rows:
        for add_groupids, rows in _batch_massadd_rows(pending_rows).items():
            api.call(
                "host.massadd",
                {
                    "hosts": [{"hostid": row["hostid"]} for row in rows],
                    "groups": [{"groupid": groupid} for groupid in add_groupids],
                },
            )
            applied_batches += 1
            applied_hosts += len(rows)
            for row in rows:
                row["status"] = "applied"

    for row in pending_rows:
        row.pop("add_groupids", None)
        results.append(row)

    return {
        "host_results": results,
        "host_summary": {
            "planned_host_rows": len(impact_plan.get("host_enrich_plan") or []),
            "planned_hosts": len(hostids),
            "changed_hosts": changed_hosts,
            "applied_hosts": applied_hosts,
            "applied_batches": applied_batches,
            "already_present_hosts": already_hosts,
            "missing_hosts": missing_hosts,
        },
    }



def apply_object_changes(api: ZabbixAPI, impact_plan: Dict[str, Any], dry_run: bool) -> Dict[str, Any]:
    auto_rows = _auto_object_rows(impact_plan)
    skipped_rows = _build_skipped_rows(impact_plan)
    object_results: List[Dict[str, Any]] = []

    def add_result(row: Dict[str, Any], status: str, message: str = "") -> None:
        object_results.append(
            {
                "object_type": str(row.get("object_type") or "").strip(),
                "object_id": str(row.get("object_id") or "").strip(),
                "object_name": str(row.get("object_name") or "").strip(),
                "field_path": str(row.get("field_path") or "").strip(),
                "change_kind": str(row.get("change_kind") or "").strip(),
                "old_group": str(row.get("old_group") or "").strip(),
                "new_group": str(row.get("new_group") or "").strip(),
                "status": status,
                "mode": "DRY-RUN" if dry_run else "APPLY",
                "message": message,
                "details": str(row.get("details") or "").strip(),
            }
        )

    action_rows_by_id: Dict[str, List[Dict[str, Any]]] = {}
    usergroup_rows_by_id: Dict[str, List[Dict[str, Any]]] = {}
    maintenance_rows_by_id: Dict[str, List[Dict[str, Any]]] = {}
    for row in auto_rows:
        object_type = str(row.get("object_type") or "").strip()
        object_id = str(row.get("object_id") or "").strip()
        if not object_id:
            continue
        if object_type == "action":
            action_rows_by_id.setdefault(object_id, []).append(row)
        elif object_type == "usergroup":
            usergroup_rows_by_id.setdefault(object_id, []).append(row)
        elif object_type == "maintenance":
            maintenance_rows_by_id.setdefault(object_id, []).append(row)
        else:
            skipped_rows.append(
                {
                    "object_type": object_type,
                    "object_id": object_id,
                    "object_name": str(row.get("object_name") or "").strip(),
                    "field_path": str(row.get("field_path") or "").strip(),
                    "change_kind": str(row.get("change_kind") or "").strip(),
                    "old_group": str(row.get("old_group") or "").strip(),
                    "new_group": str(row.get("new_group") or "").strip(),
                    "reason": "unsupported_object_type",
                    "details": str(row.get("details") or "").strip(),
                }
            )

    actions_by_id = {str(row.get("actionid") or "").strip(): row for row in fetch_actions(api, sorted(action_rows_by_id))}
    for object_id, rows in action_rows_by_id.items():
        action = actions_by_id.get(object_id)
        if action is None:
            for row in rows:
                add_result(row, "object_not_found", "action.get did not return this actionid")
            continue
        touched_sections: Set[str] = set()
        changed_rows: List[Dict[str, Any]] = []
        for row in rows:
            field_path = str(row.get("field_path") or "").strip()
            old_groupid = str(row.get("old_groupid") or "").strip()
            new_groupid = str(row.get("new_groupid") or "").strip()
            try:
                before = str(_get_path_value(action, field_path) or "").strip()
            except Exception as exc:
                add_result(row, "field_not_found", str(exc))
                continue
            if before == new_groupid:
                add_result(row, "already_present")
                continue
            if before != old_groupid:
                add_result(row, "source_mismatch", f"current={before} expected_old={old_groupid}")
                continue
            _set_path_value(action, field_path, new_groupid)
            touched_sections.add(_top_level_section(field_path))
            changed_rows.append(row)
        if changed_rows and not dry_run:
            payload: Dict[str, Any] = {"actionid": object_id}
            for section in sorted(touched_sections):
                payload[section] = action.get(section)
            api.call("action.update", payload)
        for row in changed_rows:
            add_result(row, "dry_run_changed" if dry_run else "applied")

    usergroups_by_id = {str(row.get("usrgrpid") or "").strip(): row for row in fetch_usergroups(api, sorted(usergroup_rows_by_id))}
    for object_id, rows in usergroup_rows_by_id.items():
        usergroup = usergroups_by_id.get(object_id)
        if usergroup is None:
            for row in rows:
                add_result(row, "object_not_found", "usergroup.get did not return this usrgrpid")
            continue
        rights = [_normalize_right(item) for item in (usergroup.get("hostgroup_rights") or [])]
        rights = [item for item in rights if item]
        rights_by_id = {item["id"]: item for item in rights}
        changed_rows: List[Dict[str, Any]] = []
        for row in rows:
            new_groupid = str(row.get("new_groupid") or "").strip()
            permission = ""
            details = str(row.get("details") or "")
            if details.startswith("permission="):
                permission = details.split("=", 1)[1]
            if new_groupid in rights_by_id:
                add_result(row, "already_present")
                continue
            rights.append({"id": new_groupid, "permission": permission})
            rights_by_id[new_groupid] = rights[-1]
            changed_rows.append(row)
        if changed_rows and not dry_run:
            api.call("usergroup.update", {"usrgrpid": object_id, "hostgroup_rights": rights})
        for row in changed_rows:
            add_result(row, "dry_run_changed" if dry_run else "applied")

    maintenances_by_id = {str(row.get("maintenanceid") or "").strip(): row for row in fetch_maintenances(api, sorted(maintenance_rows_by_id))}
    for object_id, rows in maintenance_rows_by_id.items():
        maintenance = maintenances_by_id.get(object_id)
        if maintenance is None:
            for row in rows:
                add_result(row, "object_not_found", "maintenance.get did not return this maintenanceid")
            continue
        groups = maintenance.get("groups") or []
        changed_rows: List[Dict[str, Any]] = []
        for row in rows:
            field_path = str(row.get("field_path") or "").strip()
            old_groupid = str(row.get("old_groupid") or "").strip()
            new_groupid = str(row.get("new_groupid") or "").strip()
            try:
                before = str(_get_path_value(maintenance, field_path) or "").strip()
            except Exception as exc:
                add_result(row, "field_not_found", str(exc))
                continue
            if before == new_groupid:
                add_result(row, "already_present")
                continue
            if before != old_groupid:
                add_result(row, "source_mismatch", f"current={before} expected_old={old_groupid}")
                continue
            _set_path_value(maintenance, field_path, new_groupid)
            changed_rows.append(row)
        if changed_rows and not dry_run:
            api.call("maintenance.update", {"maintenanceid": object_id, "groups": groups})
        for row in changed_rows:
            add_result(row, "dry_run_changed" if dry_run else "applied")

    summary = {
        "planned_object_rows": len(auto_rows),
        "changed_object_rows": sum(1 for row in object_results if row.get("status") in {"dry_run_changed", "applied"}),
        "already_present_object_rows": sum(1 for row in object_results if row.get("status") == "already_present"),
        "object_errors": sum(1 for row in object_results if row.get("status") in {"field_not_found", "object_not_found", "source_mismatch"}),
        "skipped_object_rows": len(skipped_rows),
    }
    return {
        "object_results": object_results,
        "skipped_rows": skipped_rows,
        "object_summary": summary,
    }



def apply_zabbix_changes(api: ZabbixAPI, impact_plan: Dict[str, Any], dry_run: bool) -> Dict[str, Any]:
    host_data = apply_host_enrichment(api, impact_plan, dry_run=dry_run)
    object_data = apply_object_changes(api, impact_plan, dry_run=dry_run)
    summary = impact_plan.get("summary") or {}
    out_summary = {
        "scope_as": summary.get("scope_as") or [],
        "scope_env": str(summary.get("scope_env") or "").strip(),
        "scope_gas": summary.get("scope_gas") or [],
        **host_data["host_summary"],
        **object_data["object_summary"],
        "mode": "DRY-RUN" if dry_run else "APPLY",
    }
    return {
        "meta": {
            "created_at": datetime.now().isoformat(timespec="seconds"),
            "version": "2.0",
            "impact_plan_path": "",
            "zabbix_url": str(getattr(api, "api_url", "")),
            "mode": "DRY-RUN" if dry_run else "APPLY",
        },
        "summary": out_summary,
        "host_results": host_data["host_results"],
        "object_results": object_data["object_results"],
        "skipped_rows": object_data["skipped_rows"],
    }



def write_apply_xlsx(data: Dict[str, Any], path: str) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    summary_ws = wb.create_sheet("SUMMARY")
    summary_ws.append(["key", "value"])
    for key, value in (data.get("summary") or {}).items():
        if isinstance(value, list):
            rendered = ", ".join(str(item) for item in value)
        else:
            rendered = value
        summary_ws.append([key, rendered])
    autosize_columns(summary_ws)

    hosts_ws = wb.create_sheet("HOST_APPLY")
    host_headers = [
        "hostid",
        "host",
        "name",
        "status",
        "mode",
        "current_groups",
        "planned_groups",
        "added_groups",
        "details",
    ]
    hosts_ws.append(host_headers)
    for row in data.get("host_results") or []:
        hosts_ws.append([row.get(header, "") for header in host_headers])
    autosize_columns(hosts_ws)

    object_ws = wb.create_sheet("OBJECT_APPLY")
    object_headers = [
        "object_type",
        "object_id",
        "object_name",
        "field_path",
        "change_kind",
        "old_group",
        "new_group",
        "status",
        "mode",
        "message",
        "details",
    ]
    object_ws.append(object_headers)
    for row in data.get("object_results") or []:
        object_ws.append([row.get(header, "") for header in object_headers])
    autosize_columns(object_ws)

    skipped_ws = wb.create_sheet("SKIPPED_OBJECTS")
    skipped_headers = [
        "object_type",
        "object_id",
        "object_name",
        "field_path",
        "change_kind",
        "old_group",
        "new_group",
        "reason",
        "details",
    ]
    skipped_ws.append(skipped_headers)
    for row in data.get("skipped_rows") or []:
        skipped_ws.append([row.get(header, "") for header in skipped_headers])
    autosize_columns(skipped_ws)

    wb.save(path)



def main() -> int:
    parser = argparse.ArgumentParser(description="Apply Zabbix plan from impact plan")
    parser.add_argument("--out-xlsx", dest="out_xlsx", help="Path to Zabbix apply XLSX")
    parser.add_argument("--out-json", dest="out_json", help="Path to Zabbix apply JSON")
    args = parser.parse_args()

    impact_plan_path = resolve_input_artifact(
        config.SOURCE_IMPACT_PLAN_JSON,
        config.IMPACT_PLAN_PREFIX,
        ".json",
        scope_as=config.SCOPE_AS,
        scope_env=config.SCOPE_ENV,
        scope_gas=config.SCOPE_GAS,
        label="impact plan JSON",
        strict_scope_match=True,
    )
    impact_plan = load_impact_plan(impact_plan_path)
    summary = impact_plan.get("summary") or {}
    scope_as = normalize_values(summary.get("scope_as") or [])
    scope_env = str(summary.get("scope_env") or "").strip()
    scope_gas = normalize_values(summary.get("scope_gas") or [])

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_xlsx = args.out_xlsx or build_artifact_path(config.ZABBIX_APPLY_PREFIX, scope_as, scope_env, scope_gas, ".xlsx", timestamp=timestamp)
    out_json = args.out_json or build_artifact_path(config.ZABBIX_APPLY_PREFIX, scope_as, scope_env, scope_gas, ".json", timestamp=timestamp)

    dry_run = not bool(config.ZABBIX_APPLY_CHANGES)
    connection = config.load_zabbix_connection()
    api = ZabbixAPI(connection.api_url, timeout_sec=int(config.HTTP_TIMEOUT_SEC))
    api.authenticate(connection.username, connection.password, connection.api_token)

    if not str(config.SOURCE_IMPACT_PLAN_JSON or "").strip():
        print(f"Using latest impact plan JSON: {impact_plan_path}")

    if not dry_run:
        auto_rows = _auto_object_rows(impact_plan)
        methods_to_check: Set[str] = set()
        if any(str(row.get("object_id") or "").strip() for row in impact_plan.get("host_enrich_plan") or []):
            methods_to_check.add("host.massadd")
        if any(str(row.get("object_type") or "") == "action" for row in auto_rows):
            methods_to_check.add("action.update")
        if any(str(row.get("object_type") or "") == "usergroup" for row in auto_rows):
            methods_to_check.add("usergroup.update")
        if any(str(row.get("object_type") or "") == "maintenance" for row in auto_rows):
            methods_to_check.add("maintenance.update")
        for method in sorted(methods_to_check):
            print(f"Checking Zabbix API permissions for {method}")
            params = {"hosts": [], "groups": []} if method == "host.massadd" else {}
            _assert_method_available(api, method, params)

        backup_path = resolve_input_artifact(
            config.SOURCE_BACKUP_FILE,
            config.BACKUP_PREFIX,
            ".json.gz",
            scope_as=scope_as,
            scope_env=scope_env,
            scope_gas=scope_gas,
            label="backup file",
            strict_scope_match=True,
        )
        if not str(config.SOURCE_BACKUP_FILE or "").strip():
            print(f"Using latest backup file: {backup_path}")
    prepare_zabbix_apply(api, impact_plan, backup_path, dry_run=dry_run, log=print)

    print(f"Applying Zabbix plan from: {impact_plan_path}")
    print(f"Mode: {'DRY-RUN' if dry_run else 'APPLY'}")
    result = apply_zabbix_changes(api, impact_plan, dry_run=dry_run)
    result.setdefault("meta", {})["impact_plan_path"] = impact_plan_path

    print(f"Writing Zabbix apply XLSX: {out_xlsx}")
    write_apply_xlsx(result, out_xlsx)
    print(f"Writing Zabbix apply JSON: {out_json}")
    with open(out_json, "w", encoding="utf-8") as handle:
        json.dump(result, handle, ensure_ascii=False, indent=2)
    print("Zabbix apply completed.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
