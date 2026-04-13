from __future__ import annotations

import json
import os
from collections import defaultdict
from datetime import datetime
from typing import Any, Dict, List

from openpyxl import Workbook  # type: ignore

import config
from api_clients import ZabbixAPI
from common import (
    autosize_columns,
    canonical_env_value,
    detect_discovery_host,
    get_tag_value,
    is_excluded_group,
    join_sorted,
    normalize_upper_tag_value,
    resolve_host_org,
    resolve_os_family,
    sample_host_names,
)
from zabbix_audit import fetch_hosts


QUEUE_HEADERS = [
    "as_scope",
    "as_variants",
    "org_values",
    "hosts_total",
    "enabled_hosts",
    "disabled_hosts",
    "discovery_hosts",
    "env_raw_values",
    "gas_values",
    "sample_hosts",
]

HOST_HEADERS = [
    "hostid",
    "host",
    "name",
    "status_label",
    "ORG",
    "AS",
    "GAS",
    "ENV_RAW",
    "ENV_SCOPE",
    "OS_FAMILY",
    "current_groups",
    "issue",
]

DISCOVERY_HEADERS = [
    "as_scope",
    "hostid",
    "host",
    "name",
    "status_label",
    "ORG",
    "AS",
    "GAS",
    "ENV_RAW",
    "ENV_SCOPE",
    "OS_FAMILY",
    "current_groups",
    "flags",
    "discovery_parent_hostid",
    "discovery_parent_itemid",
    "discovery_status",
    "discovery_disable_source",
    "discovery_ts_disable",
    "discovery_ts_delete",
    "discovery_reason",
]


def _status_label(status: Any) -> str:
    return "enabled" if str(status or "") == "0" else "disabled"


def _append_rows(ws, rows: List[Dict[str, Any]], headers: List[str]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    autosize_columns(ws)


def _build_paths(timestamp: str) -> tuple[str, str]:
    os.makedirs(config.OUTPUT_DIR, exist_ok=True)
    return (
        os.path.join(config.OUTPUT_DIR, f"{config.AS_QUEUE_PREFIX}_{timestamp}.xlsx"),
        os.path.join(config.OUTPUT_DIR, f"{config.AS_QUEUE_PREFIX}_{timestamp}.json"),
    )


def build_report(hosts: List[Dict[str, Any]]) -> Dict[str, Any]:
    queue_bucket: Dict[str, Dict[str, Any]] = defaultdict(
        lambda: {
            "raw_values": set(),
            "org_values": set(),
            "hostids": set(),
            "enabled_hostids": set(),
            "disabled_hostids": set(),
            "discovery_hostids": set(),
            "env_raw_values": set(),
            "gas_values": set(),
            "host_names": set(),
        }
    )
    missing_as_rows: List[Dict[str, Any]] = []
    discovery_rows: List[Dict[str, Any]] = []

    for host in hosts:
        tags = host.get("tags") or []
        as_value = get_tag_value(tags, config.TAG_AS)
        as_scope = str(as_value or "").strip().lower()
        env_raw = get_tag_value(tags, config.TAG_ENV) or ""
        env_scope = canonical_env_value(env_raw)
        gas_value = get_tag_value(tags, config.TAG_GAS) or ""
        guest_name = get_tag_value(tags, config.TAG_GUEST_NAME) or ""
        os_family = resolve_os_family(guest_name)
        filtered_groups = [
            str(group.get("name") or "")
            for group in (host.get("groups") or [])
            if str(group.get("name") or "") and not is_excluded_group(str(group.get("name") or ""))
        ]
        org_value, _ = resolve_host_org([str(host.get("host") or ""), str(host.get("name") or "")], filtered_groups)
        status_label = _status_label(host.get("status"))

        discovery = detect_discovery_host(host, filtered_groups)
        if discovery:
            discovery_rows.append(
                {
                    "as_scope": as_scope,
                    "hostid": str(host.get("hostid") or ""),
                    "host": str(host.get("host") or ""),
                    "name": str(host.get("name") or ""),
                    "status_label": status_label,
                    "ORG": org_value,
                    "AS": as_value or "",
                    "GAS": gas_value,
                    "ENV_RAW": env_raw,
                    "ENV_SCOPE": env_scope,
                    "OS_FAMILY": os_family,
                    "current_groups": join_sorted(filtered_groups),
                    "flags": discovery.get("flags", ""),
                    "discovery_parent_hostid": discovery.get("discovery_parent_hostid", ""),
                    "discovery_parent_itemid": discovery.get("discovery_parent_itemid", ""),
                    "discovery_status": discovery.get("discovery_status", ""),
                    "discovery_disable_source": discovery.get("discovery_disable_source", ""),
                    "discovery_ts_disable": discovery.get("discovery_ts_disable", ""),
                    "discovery_ts_delete": discovery.get("discovery_ts_delete", ""),
                    "discovery_reason": discovery.get("discovery_reason", ""),
                }
            )

        if not as_scope or normalize_upper_tag_value(as_value) == normalize_upper_tag_value(config.UNKNOWN_TAG_VALUE):
            missing_as_rows.append(
                {
                    "hostid": str(host.get("hostid") or ""),
                    "host": str(host.get("host") or ""),
                    "name": str(host.get("name") or ""),
                    "status_label": status_label,
                    "ORG": org_value,
                    "AS": as_value or "",
                    "GAS": gas_value,
                    "ENV_RAW": env_raw,
                    "ENV_SCOPE": env_scope,
                    "OS_FAMILY": os_family,
                    "current_groups": join_sorted(filtered_groups),
                    "issue": "missing or UNKNOWN AS",
                }
            )
            continue

        bucket = queue_bucket[as_scope]
        hostid = str(host.get("hostid") or "")
        bucket["raw_values"].add(str(as_value or "").strip())
        if org_value:
            bucket["org_values"].add(org_value)
        bucket["hostids"].add(hostid)
        bucket["host_names"].add(str(host.get("name") or host.get("host") or ""))
        if status_label == "enabled":
            bucket["enabled_hostids"].add(hostid)
        else:
            bucket["disabled_hostids"].add(hostid)
        if discovery:
            bucket["discovery_hostids"].add(hostid)
        if env_raw:
            bucket["env_raw_values"].add(env_raw)
        if gas_value:
            bucket["gas_values"].add(gas_value)

    queue_rows: List[Dict[str, Any]] = []
    for as_scope in sorted(queue_bucket.keys(), key=str.lower):
        bucket = queue_bucket[as_scope]
        queue_rows.append(
            {
                "as_scope": as_scope,
                "as_variants": join_sorted(bucket["raw_values"]),
                "org_values": join_sorted(bucket["org_values"]),
                "hosts_total": len(bucket["hostids"]),
                "enabled_hosts": len(bucket["enabled_hostids"]),
                "disabled_hosts": len(bucket["disabled_hostids"]),
                "discovery_hosts": len(bucket["discovery_hostids"]),
                "env_raw_values": join_sorted(bucket["env_raw_values"]),
                "gas_values": join_sorted(bucket["gas_values"]),
                "sample_hosts": sample_host_names(bucket["host_names"], 20),
            }
        )

    summary = {
        "hosts_total": len(hosts),
        "unique_as_total": len(queue_rows),
        "discovery_hosts_total": len({row["hostid"] for row in discovery_rows if row.get("hostid")}),
        "missing_or_unknown_as_hosts": len(missing_as_rows),
    }

    return {
        "summary": summary,
        "queue": queue_rows,
        "missing_as": sorted(missing_as_rows, key=lambda row: (str(row.get("name") or "").lower(), str(row.get("hostid") or ""))),
        "discovery_hosts": sorted(
            discovery_rows,
            key=lambda row: (str(row.get("as_scope") or "").lower(), str(row.get("name") or "").lower(), str(row.get("hostid") or "")),
        ),
    }


def write_workbook(data: Dict[str, Any], path: str) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    summary_ws = wb.create_sheet("SUMMARY")
    summary_ws.append(["key", "value"])
    for key, value in data["summary"].items():
        summary_ws.append([key, value])
    autosize_columns(summary_ws)

    queue_ws = wb.create_sheet("AS_QUEUE")
    _append_rows(queue_ws, data["queue"], QUEUE_HEADERS)

    missing_ws = wb.create_sheet("MISSING_AS")
    _append_rows(missing_ws, data["missing_as"], HOST_HEADERS)

    discovery_ws = wb.create_sheet("DISCOVERY_HOSTS")
    _append_rows(discovery_ws, data["discovery_hosts"], DISCOVERY_HEADERS)

    wb.save(path)


def save_json(data: Dict[str, Any], path: str) -> None:
    with open(path, "w", encoding="utf-8") as handle:
        json.dump(data, handle, ensure_ascii=False, indent=2)


def main() -> int:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_xlsx, out_json = _build_paths(timestamp)

    connection = config.load_zabbix_connection()
    api = ZabbixAPI(connection.api_url, timeout_sec=int(config.HTTP_TIMEOUT_SEC))
    api.authenticate(connection.username, connection.password, connection.api_token)

    hosts = fetch_hosts(api)
    data = build_report(hosts)
    write_workbook(data, out_xlsx)
    save_json(data, out_json)

    print(f"AS queue XLSX: {out_xlsx}")
    print(f"AS queue JSON: {out_json}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
