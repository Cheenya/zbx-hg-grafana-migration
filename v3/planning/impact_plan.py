from __future__ import annotations

import json
from datetime import datetime
from typing import Any, Dict, Iterable, List, Sequence, Set, Tuple

from openpyxl import Workbook  # type: ignore

import config
from clients.api_clients import ZabbixAPI
from core.common import autosize_columns, extract_action_groupids, join_sorted


def load_audit_report(path: str) -> Dict[str, Any]:
    with open(path, "r", encoding="utf-8") as handle:
        return json.load(handle)


def save_impact_plan_json(data: Dict[str, Any], path: str) -> None:
    with open(path, "w", encoding="utf-8") as handle:
        json.dump(data, handle, ensure_ascii=False, indent=2)


def write_impact_plan_xlsx(data: Dict[str, Any], path: str) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    summary_ws = wb.create_sheet("SUMMARY")
    summary_ws.append(["key", "value"])
    for key, value in (data.get("summary") or {}).items():
        if isinstance(value, list):
            rendered = ", ".join(str(item) for item in value)
        else:
            rendered = value
        summary_ws.append([key, rendered])
    autosize_columns(summary_ws)

    mapping_ws = wb.create_sheet("SELECTED_MAPPINGS")
    mapping_headers = ["AS", "old_group", "old_groupid", "new_group", "new_groupid"]
    mapping_ws.append(mapping_headers)
    for row in data.get("selected_mappings") or []:
        mapping_ws.append([row.get(header, "") for header in mapping_headers])
    autosize_columns(mapping_ws)

    zbx_ws = wb.create_sheet("ZABBIX_CHANGES")
    zbx_headers = [
        "object_type",
        "object_id",
        "object_name",
        "field_path",
        "change_kind",
        "old_group",
        "old_groupid",
        "new_group",
        "new_groupid",
        "manual_required",
        "details",
    ]
    zbx_ws.append(zbx_headers)
    for row in data.get("zabbix_changes") or []:
        zbx_ws.append([row.get(header, "") for header in zbx_headers])
    autosize_columns(zbx_ws)

    host_ws = wb.create_sheet("HOST_ENRICH_PLAN")
    host_ws.append(zbx_headers)
    for row in data.get("host_enrich_plan") or []:
        host_ws.append([row.get(header, "") for header in zbx_headers])
    autosize_columns(host_ws)

    object_ws = wb.create_sheet("OBJECT_MAPPING_PLAN")
    object_ws.append(zbx_headers)
    for row in data.get("object_mapping_plan") or []:
        object_ws.append([row.get(header, "") for header in zbx_headers])
    autosize_columns(object_ws)

    grafana_ws = wb.create_sheet("GRAFANA_CHANGES")
    grafana_headers = [
        "grafana_org_id",
        "dashboard_uid",
        "dashboard_title",
        "dashboard_url",
        "panel_url",
        "panel_id",
        "panel_title",
        "panel_type",
        "variable_name",
        "variable_type",
        "location_kind",
        "field_kind",
        "reference_kind",
        "json_path",
        "match_type",
        "change_kind",
        "old_group",
        "new_group",
        "matched_string",
        "source_text",
        "pattern_key",
        "manual_required",
        "details",
    ]
    grafana_ws.append(grafana_headers)
    for row in data.get("grafana_changes") or []:
        grafana_ws.append([row.get(header, "") for header in grafana_headers])
    autosize_columns(grafana_ws)

    grafana_review_ws = wb.create_sheet("GRAFANA_MANUAL_REVIEW")
    grafana_review_headers = grafana_headers + ["status", "message"]
    grafana_review_ws.append(grafana_review_headers)
    for row in data.get("grafana_manual_review") or []:
        grafana_review_ws.append([row.get(header, "") for header in grafana_review_headers])
    autosize_columns(grafana_review_ws)

    backup_ws = wb.create_sheet("BACKUP_SCOPE")
    backup_ws.append(["section", "value"])
    for key, value in (data.get("backup_scope") or {}).items():
        backup_ws.append([key, json.dumps(value, ensure_ascii=False)])
    autosize_columns(backup_ws)

    wb.save(path)


def fetch_actions(api: ZabbixAPI, actionids: Sequence[str]) -> List[Dict[str, Any]]:
    if not actionids:
        return []
    return api.call(
        "action.get",
        {
            "output": "extend",
            "actionids": list(actionids),
            "selectOperations": "extend",
            "selectRecoveryOperations": "extend",
            "selectUpdateOperations": "extend",
            "selectFilter": "extend",
        },
    )


def fetch_usergroups(api: ZabbixAPI, usrgrpids: Sequence[str]) -> List[Dict[str, Any]]:
    if not usrgrpids:
        return []
    return api.call(
        "usergroup.get",
        {
            "output": ["usrgrpid", "name"],
            "usrgrpids": list(usrgrpids),
            "selectHostGroupRights": "extend",
            "selectTagFilters": "extend",
            "selectUsers": ["userid", "username", "alias", "name", "surname"],
        },
    )


def fetch_maintenances(api: ZabbixAPI, maintenanceids: Sequence[str]) -> List[Dict[str, Any]]:
    if not maintenanceids:
        return []
    return api.call(
        "maintenance.get",
        {
            "output": ["maintenanceid", "name", "active_since", "active_till"],
            "maintenanceids": list(maintenanceids),
            "selectGroups": ["groupid", "name"],
        },
    )


def _iter_groupid_paths(node: Any, path: str, hits: List[Tuple[str, str]]) -> None:
    if isinstance(node, dict):
        for key, value in node.items():
            child_path = f"{path}.{key}" if path else str(key)
            if key == "groupid" and value is not None:
                hits.append((child_path, str(value)))
            else:
                _iter_groupid_paths(value, child_path, hits)
        return
    if isinstance(node, list):
        for index, item in enumerate(node):
            child_path = f"{path}[{index}]" if path else f"[{index}]"
            _iter_groupid_paths(item, child_path, hits)


def _extract_action_recipients(action: Dict[str, Any]) -> Tuple[Set[str], Set[str]]:
    usergroup_ids: Set[str] = set()
    user_ids: Set[str] = set()

    for key in ("operations", "recovery_operations", "update_operations"):
        for operation in action.get(key) or []:
            for row in operation.get("opmessage_grp") or []:
                if row.get("usrgrpid") is not None:
                    usergroup_ids.add(str(row.get("usrgrpid")))
            for row in operation.get("opmessage_usr") or []:
                if row.get("userid") is not None:
                    user_ids.add(str(row.get("userid")))

    return usergroup_ids, user_ids


def _mapping_requires_manual_object_review(mapping: Dict[str, str]) -> tuple[bool, str]:
    if str(mapping.get("manual_required") or "").strip().lower() == "yes":
        return True, "mapping row already marked manual"

    target_kind = str(mapping.get("target_kind") or "").strip().upper()
    old_env_scopes = {
        item.strip().upper()
        for item in str(mapping.get("old_env_scopes") or "").split(",")
        if item.strip()
    }
    if target_kind == "AS" and old_env_scopes:
        return True, "env-scoped old group cannot be auto-replaced with base AS"

    return False, ""


def _append_manual_detail(existing: str, extra: str) -> str:
    base = str(existing or "").strip()
    addition = str(extra or "").strip()
    if not addition:
        return base
    if not base:
        return addition
    if addition in base:
        return base
    return f"{base}; {addition}"


def build_impact_plan(
    api: ZabbixAPI,
    audit_report: Dict[str, Any],
    selected_mappings: Sequence[Dict[str, str]],
    audit_json_path: str,
    mapping_plan_path: str,
) -> Dict[str, Any]:
    inventory = audit_report.get("inventory") or {}
    scope_as = inventory.get("scope_as") or []
    scope_env = str(inventory.get("scope_env") or "").strip()
    scope_gas = [str(item).strip() for item in (inventory.get("scope_gas") or []) if str(item).strip()]
    if not scope_env:
        legacy_scope_envs = inventory.get("scope_envs") or []
        if legacy_scope_envs:
            scope_env = str(legacy_scope_envs[0] or "").strip()

    mappings_by_oldid = {str(item["old_groupid"]): dict(item) for item in selected_mappings}
    mappings_by_oldname = {str(item["old_group"]): dict(item) for item in selected_mappings}

    actions = fetch_actions(api, inventory.get("actionids") or [])
    usergroups = fetch_usergroups(api, inventory.get("usergroupids") or [])
    maintenances = fetch_maintenances(api, inventory.get("maintenanceids") or [])

    zabbix_changes: List[Dict[str, Any]] = []
    impacted_host_ids: Set[str] = set()
    impacted_action_ids: Set[str] = set()
    impacted_usergroup_ids: Set[str] = set()
    impacted_maintenance_ids: Set[str] = set()
    recipient_usergroup_ids: Set[str] = set()
    direct_user_ids: Set[str] = set()

    for action in actions:
        action_id = str(action.get("actionid") or "")
        action_name = str(action.get("name") or "")
        action_changed = False
        condition_ids, operation_ids = extract_action_groupids(action)
        action_all_groupids = condition_ids.union(operation_ids)

        for index, condition in enumerate((action.get("filter") or {}).get("conditions") or []):
            if str(condition.get("conditiontype") or "") != "0":
                continue
            group_id = str(condition.get("value") or "")
            mapping = mappings_by_oldid.get(group_id)
            if not mapping:
                continue
            manual_required, manual_details = _mapping_requires_manual_object_review(mapping)
            new_groupid = str(mapping["new_groupid"] or "").strip()
            if new_groupid and new_groupid in action_all_groupids:
                manual_required = True
                manual_details = _append_manual_detail(manual_details, "target group already present elsewhere in action")
            zabbix_changes.append(
                {
                    "object_type": "action",
                    "object_id": action_id,
                    "object_name": action_name,
                    "field_path": f"filter.conditions[{index}].value",
                    "change_kind": "manual_review_reference" if manual_required else "replace_groupid",
                    "old_group": mapping["old_group"],
                    "old_groupid": mapping["old_groupid"],
                    "new_group": mapping["new_group"],
                    "new_groupid": new_groupid,
                    "manual_required": "yes" if manual_required else "",
                    "details": manual_details or "condition hostgroup",
                }
            )
            action_changed = True

        groupid_hits: List[Tuple[str, str]] = []
        for key in ("operations", "recovery_operations", "update_operations"):
            _iter_groupid_paths(action.get(key), key, groupid_hits)
        for field_path, group_id in groupid_hits:
            mapping = mappings_by_oldid.get(group_id)
            if not mapping:
                continue
            manual_required, manual_details = _mapping_requires_manual_object_review(mapping)
            new_groupid = str(mapping["new_groupid"] or "").strip()
            if new_groupid and new_groupid in action_all_groupids:
                manual_required = True
                manual_details = _append_manual_detail(manual_details, "target group already present elsewhere in action")
            zabbix_changes.append(
                {
                    "object_type": "action",
                    "object_id": action_id,
                    "object_name": action_name,
                    "field_path": field_path,
                    "change_kind": "manual_review_reference" if manual_required else "replace_groupid",
                    "old_group": mapping["old_group"],
                    "old_groupid": mapping["old_groupid"],
                    "new_group": mapping["new_group"],
                    "new_groupid": new_groupid,
                    "manual_required": "yes" if manual_required else "",
                    "details": manual_details or "operation group reference",
                }
            )
            action_changed = True

        if action_changed:
            impacted_action_ids.add(action_id)
            action_usergroup_ids, action_user_ids = _extract_action_recipients(action)
            recipient_usergroup_ids.update(action_usergroup_ids)
            direct_user_ids.update(action_user_ids)

    for usergroup in usergroups:
        usergroup_id = str(usergroup.get("usrgrpid") or "")
        usergroup_name = str(usergroup.get("name") or "")
        changed = False
        existing_rights_by_groupid = {
            str((right.get("groupid") or right.get("id") or right.get("hostgroupid") or "")).strip(): str(right.get("permission") or "")
            for right in (usergroup.get("hostgroup_rights") or [])
            if str((right.get("groupid") or right.get("id") or right.get("hostgroupid") or "")).strip()
        }
        for index, right in enumerate(usergroup.get("hostgroup_rights") or []):
            group_id = str(right.get("groupid") or right.get("id") or right.get("hostgroupid") or "")
            mapping = mappings_by_oldid.get(group_id)
            if not mapping:
                continue
            manual_required, manual_details = _mapping_requires_manual_object_review(mapping)
            new_groupid = str(mapping["new_groupid"] or "").strip()
            old_permission = str(right.get("permission") or "")
            existing_new_permission = str(existing_rights_by_groupid.get(new_groupid) or "")
            if manual_required:
                zabbix_changes.append(
                    {
                        "object_type": "usergroup",
                        "object_id": usergroup_id,
                        "object_name": usergroup_name,
                        "field_path": f"hostgroup_rights[{index}]",
                        "change_kind": "manual_review_permission",
                        "old_group": mapping["old_group"],
                        "old_groupid": mapping["old_groupid"],
                        "new_group": mapping["new_group"],
                        "new_groupid": new_groupid,
                        "manual_required": "yes",
                        "details": manual_details or f"permission={old_permission}",
                    }
                )
                changed = True
                continue
            if existing_new_permission:
                if existing_new_permission != old_permission:
                    zabbix_changes.append(
                        {
                            "object_type": "usergroup",
                            "object_id": usergroup_id,
                            "object_name": usergroup_name,
                            "field_path": f"hostgroup_rights[{index}]",
                            "change_kind": "manual_review_permission",
                            "old_group": mapping["old_group"],
                            "old_groupid": mapping["old_groupid"],
                            "new_group": mapping["new_group"],
                            "new_groupid": new_groupid,
                            "manual_required": "yes",
                            "details": f"old_permission={old_permission}; new_permission={existing_new_permission}",
                        }
                    )
                    changed = True
                continue
            zabbix_changes.append(
                {
                    "object_type": "usergroup",
                    "object_id": usergroup_id,
                    "object_name": usergroup_name,
                    "field_path": "hostgroup_rights[+]",
                    "change_kind": "add_group_permission",
                    "old_group": mapping["old_group"],
                    "old_groupid": mapping["old_groupid"],
                    "new_group": mapping["new_group"],
                    "new_groupid": new_groupid,
                    "manual_required": "",
                    "details": f"permission={old_permission}",
                }
            )
            changed = True
        if changed:
            impacted_usergroup_ids.add(usergroup_id)

    for maintenance in maintenances:
        maintenance_id = str(maintenance.get("maintenanceid") or "")
        maintenance_name = str(maintenance.get("name") or "")
        changed = False
        maintenance_groupids = {
            str(group.get("groupid") or "")
            for group in (maintenance.get("groups") or [])
            if str(group.get("groupid") or "").strip()
        }
        for index, group in enumerate(maintenance.get("groups") or []):
            group_id = str(group.get("groupid") or "")
            mapping = mappings_by_oldid.get(group_id)
            if not mapping:
                continue
            manual_required, manual_details = _mapping_requires_manual_object_review(mapping)
            new_groupid = str(mapping["new_groupid"] or "").strip()
            if new_groupid and new_groupid in maintenance_groupids:
                manual_required = True
                manual_details = _append_manual_detail(manual_details, "target group already present in maintenance")
            zabbix_changes.append(
                {
                    "object_type": "maintenance",
                    "object_id": maintenance_id,
                    "object_name": maintenance_name,
                    "field_path": f"groups[{index}].groupid",
                    "change_kind": "manual_review_reference" if manual_required else "replace_groupid",
                    "old_group": mapping["old_group"],
                    "old_groupid": mapping["old_groupid"],
                    "new_group": mapping["new_group"],
                    "new_groupid": new_groupid,
                    "manual_required": "yes" if manual_required else "",
                    "details": manual_details or "maintenance group reference",
                }
            )
            changed = True
        if changed:
            impacted_maintenance_ids.add(maintenance_id)

    for row in audit_report.get("host_expected_groups") or []:
        hostid = str(row.get("hostid") or "").strip()
        groupid = str(row.get("groupid") or "").strip()
        group_name = str(row.get("group_name") or "").strip()
        if not hostid or not groupid or not group_name:
            continue
        if str(row.get("exists_in_zabbix") or "") != "yes":
            continue
        if str(row.get("on_host") or "") == "yes":
            continue
        zabbix_changes.append(
            {
                "object_type": "host",
                "object_id": hostid,
                "object_name": str(row.get("name") or row.get("host") or ""),
                "field_path": "groups",
                "change_kind": "add_group",
                "old_group": "",
                "old_groupid": "",
                "new_group": group_name,
                "new_groupid": groupid,
                "manual_required": "",
                "details": f"expected {str(row.get('group_kind') or '').strip()}",
            }
        )
        impacted_host_ids.add(hostid)

    backup_usergroup_ids: Set[str] = set(impacted_usergroup_ids)
    backup_usergroup_ids.update(recipient_usergroup_ids)
    backup_usergroups = fetch_usergroups(api, sorted(backup_usergroup_ids))

    backup_user_ids: Set[str] = set(direct_user_ids)
    for usergroup in backup_usergroups:
        for user in usergroup.get("users") or []:
            if user.get("userid") is not None:
                backup_user_ids.add(str(user.get("userid")))

    grafana_changes: List[Dict[str, Any]] = []
    grafana_manual_review: List[Dict[str, Any]] = []
    for row in audit_report.get("grafana") or []:
        match_type = str(row.get("match_type") or "")
        matched_string = str(row.get("matched_string") or "")
        location_kind = str(row.get("location_kind") or "")

        if match_type == "OLD":
            mapping = mappings_by_oldname.get(matched_string)
            if not mapping:
                continue
            out_row = {
                "grafana_org_id": str(row.get("grafana_org_id") or ""),
                "dashboard_uid": str(row.get("dashboard_uid") or ""),
                "dashboard_title": str(row.get("dashboard_title") or ""),
                "dashboard_url": str(row.get("dashboard_url") or ""),
                "panel_url": str(row.get("panel_url") or ""),
                "panel_id": str(row.get("panel_id") or ""),
                "panel_title": str(row.get("panel_title") or ""),
                "panel_type": str(row.get("panel_type") or ""),
                "variable_name": str(row.get("variable_name") or ""),
                "variable_type": str(row.get("variable_type") or ""),
                "location_kind": location_kind,
                "field_kind": str(row.get("field_kind") or ""),
                "reference_kind": str(row.get("reference_kind") or ""),
                "json_path": str(row.get("json_path") or ""),
                "source_text": str(row.get("source_text") or ""),
                "match_type": match_type,
                "change_kind": "replace_exact_string",
                "old_group": mapping["old_group"],
                "new_group": mapping["new_group"],
                "matched_string": matched_string,
                "pattern_key": str(row.get("pattern_key") or ""),
                "manual_required": "",
                "details": "",
            }
            if location_kind == "variable":
                grafana_changes.append(out_row)
            else:
                grafana_manual_review.append(
                    {
                        **out_row,
                        "status": "unsupported_location",
                        "message": "Non-variable Grafana matches stay in manual review and are not included in executable plan.",
                    }
                )
            continue

        if match_type != "OLD_PATTERN":
            continue

        related = [
            mapping
            for old_group, mapping in mappings_by_oldname.items()
            if old_group.lower() in matched_string.lower()
        ]
        if len(related) == 1:
            old_group = related[0]["old_group"]
            new_group = related[0]["new_group"]
            details = "pattern match requires manual review"
        else:
            old_group = ""
            new_group = ""
            details = "pattern match could not be mapped uniquely"

        out_row = {
            "grafana_org_id": str(row.get("grafana_org_id") or ""),
            "dashboard_uid": str(row.get("dashboard_uid") or ""),
            "dashboard_title": str(row.get("dashboard_title") or ""),
            "dashboard_url": str(row.get("dashboard_url") or ""),
            "panel_url": str(row.get("panel_url") or ""),
            "panel_id": str(row.get("panel_id") or ""),
            "panel_title": str(row.get("panel_title") or ""),
            "panel_type": str(row.get("panel_type") or ""),
            "variable_name": str(row.get("variable_name") or ""),
            "variable_type": str(row.get("variable_type") or ""),
            "location_kind": location_kind,
            "field_kind": str(row.get("field_kind") or ""),
            "reference_kind": str(row.get("reference_kind") or ""),
            "json_path": str(row.get("json_path") or ""),
            "source_text": str(row.get("source_text") or ""),
            "match_type": match_type,
            "change_kind": "review_pattern",
            "old_group": old_group,
            "new_group": new_group,
            "matched_string": matched_string,
            "pattern_key": str(row.get("pattern_key") or ""),
            "manual_required": "yes",
            "details": details,
        }
        if location_kind == "variable":
            grafana_changes.append(out_row)
        else:
            grafana_manual_review.append(
                {
                    **out_row,
                    "status": "unsupported_location",
                    "message": "Non-variable Grafana matches stay in manual review and are not included in executable plan.",
                }
            )

    hostgroups_scope: List[Dict[str, str]] = []
    seen_groupids: Set[str] = set()
    for item in selected_mappings:
        for kind, key_id, key_name in (("OLD", "old_groupid", "old_group"), ("NEW", "new_groupid", "new_group")):
            group_id = str(item.get(key_id) or "").strip()
            group_name = str(item.get(key_name) or "").strip()
            if not group_id or group_id in seen_groupids:
                continue
            seen_groupids.add(group_id)
            hostgroups_scope.append({"groupid": group_id, "name": group_name, "kind": kind})

    for row in audit_report.get("host_expected_groups") or []:
        group_id = str(row.get("groupid") or "").strip()
        group_name = str(row.get("group_name") or "").strip()
        if not group_id or not group_name:
            continue
        if str(row.get("exists_in_zabbix") or "") != "yes":
            continue
        if group_id in seen_groupids:
            continue
        seen_groupids.add(group_id)
        hostgroups_scope.append({"groupid": group_id, "name": group_name, "kind": str(row.get("group_kind") or "STANDARD")})

    backup_scope = {
        "hostids": sorted(impacted_host_ids),
        "hostgroups": sorted(hostgroups_scope, key=lambda item: (item["kind"], item["name"].lower())),
        "actionids": sorted(impacted_action_ids),
        "usergroupids": sorted(backup_usergroup_ids),
        "userids": sorted(backup_user_ids),
        "maintenanceids": sorted(impacted_maintenance_ids),
    }

    summary = {
        "scope_as": scope_as,
        "scope_env": scope_env,
        "scope_gas": scope_gas,
        "selected_mappings": len(selected_mappings),
        "zabbix_changes": len(zabbix_changes),
        "host_changes": len(impacted_host_ids),
        "host_enrich_plan_rows": sum(1 for row in zabbix_changes if str(row.get("object_type") or "") == "host"),
        "object_mapping_plan_rows": sum(1 for row in zabbix_changes if str(row.get("object_type") or "") != "host"),
        "grafana_changes": len(grafana_changes),
        "grafana_manual_review": len(grafana_manual_review),
        "backup_hostids": len(backup_scope["hostids"]),
        "backup_hostgroups": len(backup_scope["hostgroups"]),
        "backup_actions": len(backup_scope["actionids"]),
        "backup_usergroups": len(backup_scope["usergroupids"]),
        "backup_users": len(backup_scope["userids"]),
        "backup_maintenances": len(backup_scope["maintenanceids"]),
    }

    return {
        "meta": {
            "created_at": datetime.now().isoformat(timespec="seconds"),
            "version": "2.0",
            "audit_json_path": audit_json_path,
            "mapping_plan_path": mapping_plan_path,
            "zabbix_url": str(getattr(api, "api_url", "")),
        },
        "summary": summary,
        "selected_mappings": list(selected_mappings),
        "backup_scope": backup_scope,
        "zabbix_changes": zabbix_changes,
        "host_enrich_plan": [row for row in zabbix_changes if str(row.get("object_type") or "") == "host"],
        "object_mapping_plan": [row for row in zabbix_changes if str(row.get("object_type") or "") != "host"],
        "grafana_changes": grafana_changes,
        "grafana_manual_review": grafana_manual_review,
    }
