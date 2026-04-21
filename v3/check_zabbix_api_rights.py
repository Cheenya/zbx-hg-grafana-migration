#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Небоевая проверка API-прав текущей Zabbix УЗ."""

from __future__ import annotations

import argparse
import json
from datetime import datetime
from typing import Any, Dict, List

from openpyxl import Workbook  # type: ignore

import config
from clients.api_clients import ZabbixAPI
from core.common import autosize_columns, build_artifact_path


METHODS_TO_PROBE = (
    "host.massadd",
    "host.update",
    "action.update",
    "maintenance.update",
    "usergroup.update",
    "user.update",
)


def _probe_method(api: ZabbixAPI, method: str) -> Dict[str, str]:
    try:
        api.call(method, {})
    except RuntimeError as exc:
        message = str(exc)
        if f'No permissions to call "{method}"' in message:
            return {
                "method": method,
                "status": "denied_by_role",
                "details": message,
            }
        return {
            "method": method,
            "status": "allowed_by_role_bad_params",
            "details": message,
        }
    return {
        "method": method,
        "status": "allowed_by_role",
        "details": "Method accepted empty params unexpectedly",
    }


def _get_auth_info(api: ZabbixAPI) -> Dict[str, Any]:
    if getattr(api, "api_token", ""):
        auth_info = api.call(
            "user.checkAuthentication",
            {"token": api.api_token},
            include_auth=False,
            include_bearer=False,
        )
    else:
        auth_info = api.call(
            "user.checkAuthentication",
            {"sessionid": api.auth},
            include_auth=False,
            include_bearer=False,
        )

    userid = str(auth_info.get("userid") or "").strip()
    user_row: Dict[str, Any] = {}
    role_row: Dict[str, Any] = {}
    if userid:
        rows = api.call(
            "user.get",
            {
                "output": ["userid", "username", "name", "surname", "roleid"],
                "userids": [userid],
            },
        )
        if rows:
            user_row = rows[0]
            if str(user_row.get("roleid") or "").strip():
                roles = api.call(
                    "role.get",
                    {
                        "output": "extend",
                        "selectRules": "extend",
                        "roleids": [str(user_row.get("roleid") or "").strip()],
                    },
                )
                if roles:
                    role_row = roles[0]

    return {
        "auth_info": auth_info,
        "user": user_row,
        "role": role_row,
    }


def _write_xlsx(data: Dict[str, Any], path: str) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    summary_ws = wb.create_sheet("SUMMARY")
    summary_ws.append(["key", "value"])
    for key, value in (data.get("summary") or {}).items():
        if isinstance(value, list):
            rendered = ", ".join(str(item) for item in value)
        else:
            rendered = value
        summary_ws.append([key, rendered])
    autosize_columns(summary_ws)

    user_ws = wb.create_sheet("USER")
    user_ws.append(["key", "value"])
    for key, value in (data.get("user") or {}).items():
        user_ws.append([key, value])
    autosize_columns(user_ws)

    role_ws = wb.create_sheet("ROLE")
    role_ws.append(["key", "value"])
    for key, value in (data.get("role") or {}).items():
        role_ws.append([key, json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value])
    autosize_columns(role_ws)

    methods_ws = wb.create_sheet("METHODS")
    headers = ["method", "status", "details"]
    methods_ws.append(headers)
    for row in data.get("methods") or []:
        methods_ws.append([row.get(header, "") for header in headers])
    autosize_columns(methods_ws)

    wb.save(path)


def main() -> int:
    parser = argparse.ArgumentParser(description="Non-destructive Zabbix API rights probe")
    parser.add_argument("--out-xlsx", dest="out_xlsx", help="Path to XLSX report")
    parser.add_argument("--out-json", dest="out_json", help="Path to JSON report")
    args = parser.parse_args()

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_xlsx = args.out_xlsx or build_artifact_path("zabbix_api_rights", config.SCOPE_AS, config.SCOPE_ENV, config.SCOPE_GAS, ".xlsx", timestamp=timestamp)
    out_json = args.out_json or build_artifact_path("zabbix_api_rights", config.SCOPE_AS, config.SCOPE_ENV, config.SCOPE_GAS, ".json", timestamp=timestamp)

    connection = config.load_zabbix_connection()
    api = ZabbixAPI(connection.api_url, timeout_sec=int(config.HTTP_TIMEOUT_SEC))
    api.authenticate(connection.username, connection.password, connection.api_token)

    auth_data = _get_auth_info(api)
    method_rows: List[Dict[str, str]] = [_probe_method(api, method) for method in METHODS_TO_PROBE]

    data = {
        "meta": {
            "created_at": datetime.now().isoformat(timespec="seconds"),
            "zabbix_url": connection.api_url,
            "auth_mode": "token" if str(connection.api_token or "").strip() else "password",
        },
        "summary": {
            "zabbix_url": connection.api_url,
            "auth_mode": "token" if str(connection.api_token or "").strip() else "password",
            "username": str((auth_data.get("user") or {}).get("username") or (auth_data.get("auth_info") or {}).get("username") or "").strip(),
            "userid": str((auth_data.get("auth_info") or {}).get("userid") or "").strip(),
            "roleid": str((auth_data.get("user") or {}).get("roleid") or (auth_data.get("role") or {}).get("roleid") or "").strip(),
            "role_name": str((auth_data.get("role") or {}).get("name") or "").strip(),
            "role_type": str((auth_data.get("role") or {}).get("type") or (auth_data.get("auth_info") or {}).get("type") or "").strip(),
            "methods_checked": len(method_rows),
            "methods_denied": sum(1 for row in method_rows if row["status"] == "denied_by_role"),
        },
        "auth_info": auth_data.get("auth_info") or {},
        "user": auth_data.get("user") or {},
        "role": auth_data.get("role") or {},
        "methods": method_rows,
    }

    with open(out_json, "w", encoding="utf-8") as handle:
        json.dump(data, handle, ensure_ascii=False, indent=2)
    _write_xlsx(data, out_xlsx)

    print(f"Auth mode: {data['summary']['auth_mode']}")
    print(f"User: {data['summary']['username'] or '-'} (userid={data['summary']['userid'] or '-'})")
    print(f"Role: {data['summary']['role_name'] or '-'} (roleid={data['summary']['roleid'] or '-'}, type={data['summary']['role_type'] or '-'})")
    for row in method_rows:
        print(f"{row['method']}: {row['status']}")
    print(f"Wrote XLSX: {out_xlsx}")
    print(f"Wrote JSON: {out_json}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
