from __future__ import annotations

from typing import Any, Dict, List, Sequence

from openpyxl import Workbook, load_workbook  # type: ignore

from common import autosize_columns


MAPPING_PLAN_HEADERS: List[str] = [
    "selected",
    "AS",
    "ORG",
    "old_group",
    "old_groupid",
    "old_group_kind",
    "legacy_env_token",
    "new_group",
    "new_groupid",
    "target_kind",
    "target_exists",
    "candidate_rank",
    "candidate_count",
    "old_hosts_count",
    "target_scope_hosts",
    "new_hosts_count",
    "host_action",
    "hosts_need_add_new",
    "hosts_already_have_new",
    "old_orgs",
    "old_envs",
    "old_env_scopes",
    "target_env_raw",
    "auto_reason",
    "manual_required",
    "status",
    "comment",
]


def write_mapping_plan_xlsx(rows: Sequence[Dict[str, Any]], path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "MAPPING_PLAN"
    ws.append(MAPPING_PLAN_HEADERS)
    for row in rows:
        ws.append([row.get(header, "") for header in MAPPING_PLAN_HEADERS])
    autosize_columns(ws)
    wb.save(path)


def load_mapping_plan_rows(path: str) -> List[Dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=False)
    ws = wb["MAPPING_PLAN"] if "MAPPING_PLAN" in wb.sheetnames else wb[wb.sheetnames[0]]
    raw_rows = list(ws.iter_rows(values_only=True))
    if not raw_rows:
        return []

    headers = [str(item or "").strip() for item in raw_rows[0]]
    rows: List[Dict[str, Any]] = []
    for raw in raw_rows[1:]:
        if not raw:
            continue
        row = {headers[index]: raw[index] for index in range(min(len(headers), len(raw))) if headers[index]}
        if any(str(value or "").strip() for value in row.values()):
            rows.append(row)
    return rows


def is_selected(value: Any) -> bool:
    text = str(value or "").strip().lower()
    return text in {"1", "y", "yes", "true", "x"}


def get_selected_mappings(rows: Sequence[Dict[str, Any]]) -> List[Dict[str, str]]:
    selected: List[Dict[str, str]] = []
    old_seen: Dict[str, str] = {}

    for row in rows:
        if not is_selected(row.get("selected")):
            continue

        old_group = str(row.get("old_group") or "").strip()
        old_groupid = str(row.get("old_groupid") or "").strip()
        new_group = str(row.get("new_group") or "").strip()
        new_groupid = str(row.get("new_groupid") or "").strip()
        as_value = str(row.get("AS") or "").strip()

        if not old_group or not old_groupid or not new_group or not new_groupid:
            raise RuntimeError("Selected mapping row must contain old_group, old_groupid, new_group and new_groupid.")

        if old_group in old_seen:
            raise RuntimeError(f"Duplicate selected old_group in mapping plan: {old_group}")

        old_seen[old_group] = new_group
        selected.append(
            {
                "AS": as_value,
                "old_group": old_group,
                "old_groupid": old_groupid,
                "new_group": new_group,
                "new_groupid": new_groupid,
                "old_group_kind": str(row.get("old_group_kind") or "").strip(),
                "target_kind": str(row.get("target_kind") or "").strip(),
                "manual_required": str(row.get("manual_required") or "").strip(),
                "status": str(row.get("status") or "").strip(),
                "candidate_count": str(row.get("candidate_count") or "").strip(),
                "old_envs": str(row.get("old_envs") or "").strip(),
                "old_env_scopes": str(row.get("old_env_scopes") or "").strip(),
                "target_env_raw": str(row.get("target_env_raw") or "").strip(),
            }
        )

    if not selected:
        raise RuntimeError("No selected mappings found in mapping plan.")

    return selected
