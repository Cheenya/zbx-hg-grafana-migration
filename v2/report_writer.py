from __future__ import annotations

import json
from datetime import datetime
from typing import Any, Dict, Iterable, List, Sequence

from openpyxl import Workbook  # type: ignore

from .common import autosize_columns, safe_sheet_title


def _append_rows(ws, rows: Sequence[Dict[str, Any]], headers: Sequence[str]) -> None:
    ws.append(list(headers))
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    autosize_columns(ws)


def save_inventory_json(report: Dict[str, Any], path: str) -> None:
    payload = {
        "meta": {
            "created_at": datetime.now().isoformat(timespec="seconds"),
            "version": "2.0",
        },
        "summary": report["summary"],
        "inventory": report["inventory"],
        "hosts": report["hosts"],
        "hosts_skipped_env": report["hosts_skipped_env"],
        "groups_old": report["groups_old"],
        "groups_new": report["groups_new"],
        "actions": report["actions"],
        "usergroups": report["usergroups"],
        "maintenances": report["maintenances"],
        "grafana": report["grafana"],
    }
    with open(path, "w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2)


def write_workbook(report: Dict[str, Any], out_path: str) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    summary_ws = wb.create_sheet("SUMMARY")
    summary_ws.append(["key", "value"])
    for key, value in report["summary"].items():
        if isinstance(value, list):
            rendered = ", ".join(str(item) for item in value)
        else:
            rendered = value
        summary_ws.append([key, rendered])
    autosize_columns(summary_ws)

    hosts_ws = wb.create_sheet("HOSTS")
    _append_rows(
        hosts_ws,
        report["hosts"],
        ["hostid", "host", "name", "status", "AS", "ASN", "ENV_RAW", "ENV_SCOPE", "old_groups", "new_groups", "other_groups"],
    )

    skipped_ws = wb.create_sheet("HOSTS_SKIPPED_ENV")
    _append_rows(
        skipped_ws,
        report["hosts_skipped_env"],
        ["hostid", "host", "name", "status", "AS", "ASN", "ENV_RAW", "ENV_SCOPE", "skip_reason"],
    )

    old_ws = wb.create_sheet("GROUPS_OLD")
    _append_rows(old_ws, report["groups_old"], ["group_name", "groupid", "hosts_count", "as_values", "env_values", "sample_hosts"])

    new_ws = wb.create_sheet("GROUPS_NEW")
    _append_rows(new_ws, report["groups_new"], ["group_name", "groupid", "hosts_count", "as_values", "env_values", "sample_hosts"])

    actions_ws = wb.create_sheet("ACTIONS")
    _append_rows(
        actions_ws,
        report["actions"],
        [
            "actionid",
            "name",
            "status",
            "where_found",
            "matched_groupids",
            "matched_group_names",
            "recipient_usergroups",
            "recipient_users",
            "recipients_media",
        ],
    )

    usergroups_ws = wb.create_sheet("USERGROUPS")
    _append_rows(
        usergroups_ws,
        report["usergroups"],
        [
            "usrgrpid",
            "name",
            "rights_on_scope_groups",
            "matching_tag_filters",
            "users",
            "users_media",
            "is_action_recipient",
        ],
    )

    maintenances_ws = wb.create_sheet("MAINTENANCES")
    _append_rows(
        maintenances_ws,
        report["maintenances"],
        ["maintenanceid", "name", "matched_groupids", "matched_group_names", "active_since", "active_till"],
    )

    grafana_ws = wb.create_sheet("GRAFANA")
    _append_rows(
        grafana_ws,
        report["grafana"],
        ["AS", "dashboard_uid", "dashboard_title", "match_type", "matched_string", "json_path", "count"],
    )

    inventory_ws = wb.create_sheet("INVENTORY")
    inventory_ws.append(["section", "value"])
    for key, value in report["inventory"].items():
        inventory_ws.append([key, json.dumps(value, ensure_ascii=False)])
    autosize_columns(inventory_ws)

    wb.save(out_path)
