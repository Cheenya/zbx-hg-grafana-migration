#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import json
import os
import re
from typing import Any, Dict, List, Set, Tuple

import urllib3  # type: ignore
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

from api_clients import GrafanaAPI, ZabbixAPI
from openpyxl import load_workbook  # type: ignore


# =========================
# НАСТРОЙКИ (пока прямо тут)
# =========================

AS_VALUE = "ELK"  # <-- ваша AS, которая мигрирует первой

# Zabbix
ZBX_URL = "https://zabbix-api.example.ru/api_jsonrpc.php"
ZBX_USER = "login"
ZBX_PASSWORD = "password"

# Grafana
GRAFANA_URL = "https://grafana.example.ru"
GRAFANA_USER = "login"
GRAFANA_PASSWORD = "password"
GRAFANA_TOKEN = ""  # опционально, если используете token

# Файл от аудита Zabbix с листом MAPPING
ZBX_AUDIT_XLSX = "hostgroup_mapping_audit.xlsx"
MAPPING_SHEET = "MAPPING"

# План миграции (JSON). Если пусто — путь берётся как <ZBX_AUDIT_XLSX>_migration_plan.json
MIGRATION_PLAN_JSON = ""
REQUIRE_MIGRATION_PLAN = True

# Режимы
DRY_RUN_ZABBIX = True
DRY_RUN_GRAFANA = True


# =========================
# Load mapping from XLSX
# =========================

def load_mapping_for_as(xlsx_path: str, sheet: str, as_value: str) -> Dict[str, str]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        raise RuntimeError(f"No sheet '{sheet}' in {xlsx_path}")

    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}

    header = [str(x).strip() if x is not None else "" for x in rows[0]]
    idx_as = header.index("AS")
    idx_old = header.index("old_group")
    idx_new = header.index("new_group")

    mp: Dict[str, str] = {}
    for r in rows[1:]:
        if not r:
            continue
        a = str(r[idx_as] or "").strip()
        if a.lower() != as_value.strip().lower():
            continue
        oldg = str(r[idx_old] or "").strip()
        newg = str(r[idx_new] or "").strip()
        if oldg and newg:
            mp[oldg] = newg
    return mp


def _default_plan_path(xlsx_path: str) -> str:
    base = xlsx_path
    if base.lower().endswith(".xlsx"):
        base = base[:-5]
    return f"{base}_migration_plan.json"


def load_mapping_from_plan(plan_path: str, as_value: str) -> Dict[str, str]:
    with open(plan_path, "r", encoding="utf-8") as f:
        raw = json.load(f)
    items = raw.get("items") or []
    mp: Dict[str, str] = {}
    for r in items:
        a = str(r.get("AS") or "").strip()
        if a.lower() != as_value.strip().lower():
            continue
        if r.get("enabled") is False:
            continue
        oldg = str(r.get("old_group") or "").strip()
        newg = str(r.get("new_group") or "").strip()
        if oldg and newg:
            mp[oldg] = newg
    return mp


# =========================
# Zabbix migration
# =========================

def zbx_get_hosts_by_as(api: ZabbixAPI, as_value: str) -> List[Dict[str, Any]]:
    # В Zabbix host.get можно фильтровать по tags, но в разных версиях нюансы.
    # Самый надёжный вариант: взять hosts + tags и отфильтровать локально.
    hosts = api.call("host.get", {
        "output": ["hostid", "host", "name"],
        "selectTags": ["tag", "value"],
        "selectGroups": ["groupid", "name"],
    })

    out: List[Dict[str, Any]] = []
    for h in hosts:
        tags = h.get("tags") or []
        for t in tags:
            if t.get("tag") == "AS" and str(t.get("value") or "").strip().lower() == as_value.strip().lower():
                out.append(h)
                break
    return out


def zbx_get_groupids_by_names(api: ZabbixAPI, names: Set[str]) -> Dict[str, str]:
    if not names:
        return {}
    # hostgroup.get supports filter by name array
    res = api.call("hostgroup.get", {"output": ["groupid", "name"], "filter": {"name": list(names)}})
    m: Dict[str, str] = {}
    for g in res:
        if g.get("name") and g.get("groupid"):
            m[str(g["name"])] = str(g["groupid"])
    return m


def zbx_update_host_groups(
    api: ZabbixAPI,
    host: Dict[str, Any],
    old_to_new_name: Dict[str, str],
    name_to_id: Dict[str, str],
    dry_run: bool,
) -> bool:
    """
    Удаляем OLD-группы (ключи old_to_new_name).
    NEW-группы считаются уже присутствующими, поэтому не добавляются.
    Остальные группы на хосте не трогаем.
    """
    hostid = str(host["hostid"])
    groups = host.get("groups") or []
    current_names = [str(g.get("name")) for g in groups if g.get("name")]

    old_present = [g for g in current_names if g in old_to_new_name]
    if not old_present:
        return False

    # Сохраняем все группы, кроме удаляемых old
    keep_names = [g for g in current_names if g not in old_present]
    final_names = sorted(set(keep_names))

    # name -> groupid
    final_groupids = []
    for n in final_names:
        gid = name_to_id.get(n)
        if gid:
            final_groupids.append({"groupid": gid})

    if dry_run:
        print(f"[DRY] host={host.get('host')}: remove={old_present}")
        return True

    api.call("host.update", {"hostid": hostid, "groups": final_groupids})
    print(f"[OK] host={host.get('host')}: remove={old_present}")
    return True


def zbx_replace_groupids_in_action(action_obj: Any, oldid_to_newid: Dict[str, str]) -> Tuple[Any, bool]:
    """
    Рекурсивно заменяет поля 'groupid' в структуре operations/recovery/update.
    Возвращает (new_obj, changed).
    """
    changed = False

    if isinstance(action_obj, dict):
        newd = {}
        for k, v in action_obj.items():
            if k == "groupid" and v is not None:
                sv = str(v)
                if sv in oldid_to_newid:
                    newd[k] = oldid_to_newid[sv]
                    changed = True
                else:
                    newd[k] = v
            else:
                nv, ch = zbx_replace_groupids_in_action(v, oldid_to_newid)
                newd[k] = nv
                changed = changed or ch
        return newd, changed

    if isinstance(action_obj, list):
        newl = []
        for x in action_obj:
            nx, ch = zbx_replace_groupids_in_action(x, oldid_to_newid)
            newl.append(nx)
            changed = changed or ch
        return newl, changed

    return action_obj, False


def zbx_migrate_actions(api: ZabbixAPI, oldid_to_newid: Dict[str, str], dry_run: bool) -> int:
    actions = api.call("action.get", {
        "output": "extend",
        "selectOperations": "extend",
        "selectRecoveryOperations": "extend",
        "selectUpdateOperations": "extend",
        "selectFilter": "extend",
        "filter": {"eventsource": 0},
    })

    updated = 0
    for a in actions:
        aid = str(a.get("actionid"))
        flt = a.get("filter") or {}
        conditions = flt.get("conditions") or []

        cond_changed = False
        new_conditions = []
        for c in conditions:
            cc = dict(c)
            # conditiontype=0 обычно host group
            val = cc.get("value")
            if val is not None and str(val) in oldid_to_newid:
                cc["value"] = oldid_to_newid[str(val)]
                cond_changed = True
            new_conditions.append(cc)

        ops, ch1 = zbx_replace_groupids_in_action(a.get("operations"), oldid_to_newid)
        rops, ch2 = zbx_replace_groupids_in_action(a.get("recovery_operations"), oldid_to_newid)
        uops, ch3 = zbx_replace_groupids_in_action(a.get("update_operations"), oldid_to_newid)

        if not (cond_changed or ch1 or ch2 or ch3):
            continue

        updated += 1
        if dry_run:
            print(f"[DRY] action.update actionid={aid} name={a.get('name')}")
            continue

        api.call("action.update", {
            "actionid": aid,
            "filter": {**flt, "conditions": new_conditions},
            "operations": ops,
            "recovery_operations": rops,
            "update_operations": uops,
        })
        print(f"[OK] action.update actionid={aid} name={a.get('name')}")

    return updated


def zbx_migrate_usergroup_rights(api: ZabbixAPI, oldid_to_newid: Dict[str, str], dry_run: bool) -> int:
    ugs = api.call("usergroup.get", {
        "output": ["usrgrpid", "name"],
        "selectHostGroupRights": "extend",
        "selectTagFilters": "extend",
        "selectUsers": "extend",
    })

    updated = 0
    for ug in ugs:
        rights = ug.get("hostgroup_rights") or []
        changed = False
        new_rights = []
        for r in rights:
            rr = dict(r)
            gid = rr.get("groupid") or rr.get("id") or rr.get("hostgroupid")
            if gid is not None and str(gid) in oldid_to_newid:
                rr["groupid"] = oldid_to_newid[str(gid)]
                # на всякий: уберём альтернативные ключи
                rr.pop("id", None)
                rr.pop("hostgroupid", None)
                changed = True
            new_rights.append(rr)

        if not changed:
            continue

        updated += 1
        if dry_run:
            print(f"[DRY] usergroup.update usrgrpid={ug.get('usrgrpid')} name={ug.get('name')}")
            continue

        api.call("usergroup.update", {"usrgrpid": ug.get("usrgrpid"), "hostgroup_rights": new_rights})
        print(f"[OK] usergroup.update usrgrpid={ug.get('usrgrpid')} name={ug.get('name')}")

    return updated


def zbx_migrate_maintenance(api: ZabbixAPI, oldid_to_newid: Dict[str, str], dry_run: bool) -> int:
    maint = api.call("maintenance.get", {"output": "extend", "selectGroups": "extend"})
    updated = 0
    for m in maint:
        groups = m.get("groups") or []
        changed = False
        new_groups = []
        for g in groups:
            gg = dict(g)
            gid = gg.get("groupid")
            if gid is not None and str(gid) in oldid_to_newid:
                gg["groupid"] = oldid_to_newid[str(gid)]
                changed = True
            new_groups.append(gg)

        if not changed:
            continue

        updated += 1
        if dry_run:
            print(f"[DRY] maintenance.update maintenanceid={m.get('maintenanceid')} name={m.get('name')}")
            continue

        api.call("maintenance.update", {"maintenanceid": m.get("maintenanceid"), "groups": new_groups})
        print(f"[OK] maintenance.update maintenanceid={m.get('maintenanceid')} name={m.get('name')}")
    return updated


# =========================
# Grafana migration
# =========================

OLD_TOKEN_ALLOWED = re.compile(r"^[A-Za-z0-9._-]+$")

def walk_strings(node: Any, path: str = "") -> List[Tuple[str, str]]:
    out: List[Tuple[str, str]] = []
    if isinstance(node, str):
        out.append((path, node))
        return out
    if isinstance(node, list):
        for i, v in enumerate(node):
            out.extend(walk_strings(v, f"{path}[{i}]"))
        return out
    if isinstance(node, dict):
        for k, v in node.items():
            np = f"{path}.{k}" if path else str(k)
            out.extend(walk_strings(v, np))
        return out
    return out


def replace_in_dashboard_strings(dash_obj: Dict[str, Any], old_to_new: Dict[str, str]) -> Tuple[Dict[str, Any], int]:
    """
    Делает замены только в строковых значениях JSON.
    Возвращает (new_dashboard_obj, number_of_replacements).
    """
    # Компилируем regex по OLD именам (строго по границам токена)
    # граница: не [A-Za-z0-9._-] слева/справа
    patterns: List[Tuple[re.Pattern, str]] = []
    for old, new in old_to_new.items():
        if "/" in old:
            continue
        # немного защиты от мусора
        tail = old.split("-", 1)[-1]
        if not tail or not OLD_TOKEN_ALLOWED.match(tail.replace("-", "")):
            pass
        pat = re.compile(rf"(?<![A-Za-z0-9._-]){re.escape(old)}(?![A-Za-z0-9._-])")
        patterns.append((pat, new))

    def rec(x: Any) -> Tuple[Any, int]:
        if isinstance(x, str):
            s = x
            cnt = 0
            for pat, repl in patterns:
                s2, n = pat.subn(repl, s)
                if n:
                    s = s2
                    cnt += n
            return s, cnt

        if isinstance(x, list):
            total = 0
            out = []
            for v in x:
                nv, n = rec(v)
                out.append(nv)
                total += n
            return out, total

        if isinstance(x, dict):
            total = 0
            outd = {}
            for k, v in x.items():
                nv, n = rec(v)
                outd[k] = nv
                total += n
            return outd, total

        return x, 0

    new_dash, n = rec(dash_obj)
    return new_dash, n


def grafana_migrate(api: GrafanaAPI, old_to_new: Dict[str, str], dry_run: bool) -> int:
    dashboards = api.list_dashboards()
    changed = 0

    for d in dashboards:
        uid = d.get("uid")
        if not uid:
            continue
        uid = str(uid)

        dj = api.get_dashboard_by_uid(uid)
        meta = dj.get("meta") or {}
        dash = dj.get("dashboard") or {}
        folder_id = int(meta.get("folderId") or 0)

        new_dash, n = replace_in_dashboard_strings(dash, old_to_new)
        if n == 0:
            continue

        title = str(dash.get("title") or "")
        changed += 1
        if dry_run:
            print(f"[DRY] grafana dashboard uid={uid} title={title} replacements={n}")
            continue

        api.update_dashboard(new_dash, folder_id, message=f"Auto migrate AS={AS_VALUE}: replace old host-groups")
        print(f"[OK] grafana dashboard uid={uid} title={title} replacements={n}")

    return changed


# =========================
# Main orchestration
# =========================

def main() -> int:
    plan_path = MIGRATION_PLAN_JSON or _default_plan_path(ZBX_AUDIT_XLSX)
    if REQUIRE_MIGRATION_PLAN:
        if not os.path.exists(plan_path):
            raise RuntimeError(f"Migration plan not found: {plan_path}")
        print(f"Loading migration plan for AS={AS_VALUE} from {plan_path} ...")
        old_to_new = load_mapping_from_plan(plan_path, AS_VALUE)
    else:
        if os.path.exists(plan_path):
            print(f"Loading migration plan for AS={AS_VALUE} from {plan_path} ...")
            old_to_new = load_mapping_from_plan(plan_path, AS_VALUE)
        else:
            print(f"Loading mapping for AS={AS_VALUE} from {ZBX_AUDIT_XLSX}:{MAPPING_SHEET} ...")
            old_to_new = load_mapping_for_as(ZBX_AUDIT_XLSX, MAPPING_SHEET, AS_VALUE)
    if not old_to_new:
        print("Mapping is empty for this AS. Nothing to migrate.")
        return 2

    # --- Zabbix ---
    print("\n=== Zabbix migration ===")
    zapi = ZabbixAPI(ZBX_URL, timeout_sec=120)
    zapi.login(ZBX_USER, ZBX_PASSWORD)

    hosts = zbx_get_hosts_by_as(zapi, AS_VALUE)
    print(f"Hosts in AS={AS_VALUE}: {len(hosts)}")

    # group name -> id for all involved old/new
    needed_names = set(old_to_new.keys()) | set(old_to_new.values())
    name_to_id = zbx_get_groupids_by_names(zapi, needed_names)

    # oldid->newid (только для тех, что реально существуют)
    oldid_to_newid: Dict[str, str] = {}
    for old, new in old_to_new.items():
        oldid = name_to_id.get(old)
        newid = name_to_id.get(new)
        if oldid and newid:
            oldid_to_newid[oldid] = newid

    # 1) update hosts (удаляем только OLD-группы)
    host_updates = 0
    for h in hosts:
        if zbx_update_host_groups(zapi, h, old_to_new, name_to_id, dry_run=DRY_RUN_ZABBIX):
            host_updates += 1
    print(f"Hosts updated: {host_updates}")

    # 2) actions / usergroups / maintenance
    a_upd = zbx_migrate_actions(zapi, oldid_to_newid, dry_run=DRY_RUN_ZABBIX)
    ug_upd = zbx_migrate_usergroup_rights(zapi, oldid_to_newid, dry_run=DRY_RUN_ZABBIX)
    m_upd = zbx_migrate_maintenance(zapi, oldid_to_newid, dry_run=DRY_RUN_ZABBIX)

    print(f"Actions updated: {a_upd}")
    print(f"User groups updated: {ug_upd}")
    print(f"Maintenances updated: {m_upd}")

    # --- Grafana ---
    print("\n=== Grafana migration ===")
    gapi = GrafanaAPI(GRAFANA_URL, GRAFANA_USER, GRAFANA_PASSWORD, GRAFANA_TOKEN, timeout_sec=120)
    g_changed = grafana_migrate(gapi, old_to_new, dry_run=DRY_RUN_GRAFANA)
    print(f"Grafana dashboards changed: {g_changed}")

    print("\nDONE.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
