#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""zbx_hg_mapping_audit.py — аудит соответствия OLD/NEW host-groups, actions и прав.

Что делает:
1) Через Zabbix API забирает хосты с группами и тегами.
2) UNKNOWN-хосты: AS==UNKNOWN, ASN==UNKNOWN, группа UNKNOWN, или отсутствует AS (в эту выборку не попадает). Исключения групп — regex из config.py.
3) По каждой AS (тег AS) строит частотное соответствие new<->old host-groups:
   - NEW: (BNK|DOM)/AS/<AS>/... (без учёта регистра в <AS>)
   - OLD: любые legacy-группы (BNK|DOM)-... (без '/')
   - метрики: intersection, precision, recall, jaccard
   - сохраняет top-N кандидатов, применяя пороги из config.py
   - outliers: old-группы в AS, не вошедшие в эталон
   - помечает конфликтные соответствия (несколько NEW на один OLD и т.д.)
4) Находит trigger actions (eventsource=0), где фигурируют группы этой AS.
5) Находит user groups с правами/tag-фильтрами на host-groups этой AS.

Выход: Excel-файл (см. config.py): лист UNKNOWN_HOSTS, лист на каждую AS (5 секций).
Запуск: pip install requests openpyxl; python zbx_hg_mapping_audit.py
Важно: config.py должен лежать рядом и содержать CONFIG + load_connection_from_env_or_prompt.
"""

from __future__ import annotations

import hashlib
import json
from datetime import datetime
import re
from collections import Counter, defaultdict
from typing import Any, Dict, Iterable, List, Optional, Set, Tuple

import urllib3  # type: ignore

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

from openpyxl import Workbook  # type: ignore
from openpyxl.styles import Alignment, Font  # type: ignore
from openpyxl.utils import get_column_letter  # type: ignore

# --- config (рядом) ---
from api_clients import ZabbixAPI
from artifact_paths import build_migration_plan_path, build_seed_path

from config import CONFIG, load_connection_from_env_or_prompt, load_grafana_from_module

# Исключённые группы: regex из config.py
EXCLUDED_GROUP_PATTERNS = [re.compile(p) for p in getattr(CONFIG.runtime, "excluded_group_patterns", ())]

def is_excluded_group(name: str) -> bool:
    if not name:
        return False
    return any(rx.search(name) for rx in EXCLUDED_GROUP_PATTERNS)


# =========================
# Утилиты
# =========================

def get_tag_value(tags: List[Dict[str, Any]], tag_name: str) -> Optional[str]:
    for t in tags or []:
        if t.get("tag") == tag_name:
            v = t.get("value")
            if v is None:
                return None
            v = str(v).strip()
            return v if v else None
    return None


# --- Правила NEW/OLD для эталона ---


def is_as_new_group(name: str, as_val: Optional[str]) -> bool:
    """NEW-группа для эталона: (BNK|DOM)/AS/<AS>/... (регистр <AS> не важен)."""
    if not as_val:
        return False
    n = str(name)
    parts = [p for p in n.split("/") if p != ""]
    # Expect: PREFIX/AS/<AS>/...
    if len(parts) < 3:
        return False
    prefix, kind, as_seg = parts[0], parts[1], parts[2]
    if prefix not in ("BNK", "DOM"):
        return False
    if kind != "AS":
        return False
    return str(as_seg).strip().lower() == str(as_val).strip().lower()


def is_old_legacy_group(name: str) -> bool:
    """OLD-группа для эталона: (BNK|DOM)-... (без '/')."""
    n = str(name)
    if "/" in n:
        return False
    return re.match(r"^(BNK|DOM)-", n) is not None


def is_unknown_host(h: Dict[str, Any]) -> bool:
    tags = h.get("tags") or []
    as_val = get_tag_value(tags, CONFIG.tags.AS)
    asn_val = get_tag_value(tags, CONFIG.tags.ASN)

    if as_val == CONFIG.unknown.unknown_tag_value or asn_val == CONFIG.unknown.unknown_tag_value:
        return True

    groups = [
        g.get("name")
        for g in (h.get("groups") or [])
        if g.get("name") and not is_excluded_group(g.get("name"))
    ]
    if CONFIG.unknown.unknown_group_name in groups:
        return True

    # Нет AS -> в UNKNOWN
    if not as_val:
        return True

    return False


def safe_sheet_title(raw: str, max_len: int = 31) -> str:
    s = str(raw)
    s = re.sub(r"[\[\]\*:/\\\?]", "_", s).strip()
    if not s:
        return "AS"
    if len(s) <= max_len:
        return s
    h = hashlib.md5(s.encode("utf-8")).hexdigest()[:6]
    head = s[: max(1, max_len - 8)]
    return f"{head}_{h}"[:max_len]


def autosize_columns(ws, min_width: int = 10, max_width: int = 70) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(min_width, min(max_width, max_len + 2))


def write_section_title(ws, row: int, title: str) -> int:
    c = ws.cell(row=row, column=1, value=title)
    c.font = Font(bold=True, size=14)
    c.alignment = Alignment(vertical="center")
    return row + 2


def permission_name(p: Any) -> str:
    try:
        p = int(p)
    except Exception:
        return str(p)
    return {0: "deny", 2: "read", 3: "read-write"}.get(p, str(p))


def save_zabbix_seed(
    report_as: Dict[str, Any],
    unknown_rows: List[Dict[str, Any]],
    mapping_rows: List[Dict[str, Any]],
    path: str,
) -> None:
    payload = {
        "meta": {"created_at": datetime.now().isoformat(timespec="seconds"), "version": "1.1"},
        "as": {k: {"groups_old": v.get("groups_old") or [], "groups_new": v.get("groups_new") or []} for k, v in report_as.items()},
        "mapping_rows": mapping_rows or [],
        "unknown_hosts": unknown_rows or [],
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def save_migration_plan(mapping_rows: List[Dict[str, Any]], path: str, scope_as: List[str]) -> None:
    items: List[Dict[str, Any]] = []
    for r in mapping_rows or []:
        item = dict(r)
        item["enabled"] = True
        item["note"] = ""
        items.append(item)
    payload = {
        "meta": {
            "created_at": datetime.now().isoformat(timespec="seconds"),
            "version": "1.0",
            "scope_as": scope_as,
        },
        "items": items,
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def _env_top(counter: Counter[str]) -> Tuple[str, bool]:
    if not counter:
        return "", False
    top_val = str(counter.most_common(1)[0][0])
    multi = len(counter) > 1
    return top_val, multi


# --- Права/тег-фильтры (разные версии Zabbix) ---
def _get_right_groupid(r: Dict[str, Any]) -> str:
    """Zabbix usergroup hostgroup_rights may use different keys across versions."""
    gid = r.get("groupid")
    if gid is None:
        gid = r.get("id")
    if gid is None:
        gid = r.get("hostgroupid")
    return str(gid) if gid is not None else ""


def _get_tagfilter_tag(tf: Dict[str, Any]) -> Optional[str]:
    return tf.get("tag") or tf.get("tag_name") or tf.get("tagname")


def _get_tagfilter_value(tf: Dict[str, Any]) -> Optional[str]:
    v = tf.get("value")
    if v is None:
        v = tf.get("tagvalue")
    if v is None:
        v = tf.get("val")
    return str(v) if v is not None else None


# =========================
# Fetchers
# =========================
def fetch_hosts(api: ZabbixAPI) -> List[Dict[str, Any]]:
    params: Dict[str, Any] = {
        "output": ["hostid", "host", "name"],
        "selectGroups": ["groupid", "name"],
        "selectTags": ["tag", "value"],
    }
    if CONFIG.runtime.monitored_hosts_only:
        params["monitored_hosts"] = True
    return api.call("host.get", params)


def fetch_trigger_actions(api: ZabbixAPI) -> List[Dict[str, Any]]:
    # Trigger actions: eventsource=0
    return api.call(
        "action.get",
        {
            "output": "extend",
            "selectOperations": "extend",
            "selectRecoveryOperations": "extend",
            "selectUpdateOperations": "extend",
            "selectFilter": "extend",
            "filter": {"eventsource": 0},
        },
    )



def fetch_usergroups_with_rights(api: ZabbixAPI) -> List[Dict[str, Any]]:
    return api.call(
        "usergroup.get",
        {
            "output": ["usrgrpid", "name"],
            "selectHostGroupRights": "extend",
            "selectTagFilters": "extend",
            "selectUsers": "extend",
        },
    )


def fetch_users(api: ZabbixAPI) -> List[Dict[str, Any]]:
    return api.call(
        "user.get",
        {
            "output": ["userid", "username", "alias", "name", "surname"],
            "selectMedias": "extend",
            "selectUsrgrps": ["usrgrpid", "name"],
        },
    )


# =========================
# Core computations
# =========================
def build_as_index(hosts: List[Dict[str, Any]]) -> Tuple[Dict[str, List[Dict[str, Any]]], List[Dict[str, Any]]]:
    by_as: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    unknown_hosts: List[Dict[str, Any]] = []

    for h in hosts:
        if is_unknown_host(h):
            unknown_hosts.append(h)
            continue
        tags = h.get("tags") or []
        as_val = get_tag_value(tags, CONFIG.tags.AS)

        if not as_val:
            unknown_hosts.append(h)
            continue

        by_as[as_val].append(h)

    return by_as, unknown_hosts


def compute_mapping_for_as(as_val: str, hosts_in_as: List[Dict[str, Any]]) -> Tuple[
    List[Dict[str, Any]],
    List[Dict[str, Any]],
    Dict[str, Set[str]],
    Dict[str, Set[str]],
]:
    hosts_by_new: Dict[str, Set[str]] = defaultdict(set)
    hosts_by_old: Dict[str, Set[str]] = defaultdict(set)
    env_by_new: Dict[str, Counter[str]] = defaultdict(Counter)
    env_by_old: Dict[str, Counter[str]] = defaultdict(Counter)
    co: Dict[str, Counter] = defaultdict(Counter)

    for h in hosts_in_as:
        hostid = str(h["hostid"])
        groups = [g.get("name") for g in (h.get("groups") or []) if g.get("name") and not is_excluded_group(g.get("name"))]
        env_val = get_tag_value(h.get("tags") or [], CONFIG.tags.ENV)

        # Эталон по группам: NEW (..../AS/<AS>/..), OLD (BNK|DOM-..)
        new_groups = [gn for gn in groups if is_as_new_group(gn, as_val)]
        old_groups = [gn for gn in groups if is_old_legacy_group(gn)]

        for ng in new_groups:
            hosts_by_new[ng].add(hostid)
            if env_val:
                env_by_new[ng][str(env_val)] += 1
        for og in old_groups:
            hosts_by_old[og].add(hostid)
            if env_val:
                env_by_old[og][str(env_val)] += 1

        for ng in new_groups:
            for og in old_groups:
                co[ng][og] += 1

    # build candidates
    top_n = int(CONFIG.mapping.top_n_candidates)
    min_inter = int(CONFIG.mapping.min_intersection)
    min_precision = float(CONFIG.mapping.min_precision)

    etalon_rows: List[Dict[str, Any]] = []
    used_old_in_etalon: Set[str] = set()

    per_new_selected: Dict[str, List[Tuple[str, int, float, float, float, int, int]]] = {}

    for ng, old_counter in co.items():
        n_hosts = len(hosts_by_new.get(ng, set()))
        if n_hosts == 0:
            continue

        candidates: List[Tuple[str, int, float, float, float, int, int]] = []
        for og, inter in old_counter.items():
            o_hosts = len(hosts_by_old.get(og, set()))
            if o_hosts == 0:
                continue
            precision = inter / n_hosts
            recall = inter / o_hosts
            denom = n_hosts + o_hosts - inter
            jaccard = (inter / denom) if denom else 0.0
            candidates.append((og, inter, precision, recall, jaccard, n_hosts, o_hosts))

        candidates.sort(key=lambda x: (x[4], x[1]), reverse=True)  # jaccard, intersection

        selected: List[Tuple[str, int, float, float, float, int, int]] = []
        for item in candidates:
            og, inter, precision, recall, jaccard, n_hosts, o_hosts = item
            # ENV-политика: при запрете несовпадений не допускаем пары PROD<->NONPROD и т.п.
            if CONFIG.mapping.forbid_env_mismatch:
                env_new_top, _ = _env_top(env_by_new.get(ng, Counter()))
                env_old_top, _ = _env_top(env_by_old.get(og, Counter()))
                if env_new_top and env_old_top:
                    if str(env_new_top).strip().upper() != str(env_old_top).strip().upper():
                        continue
            if inter < min_inter:
                continue
            if precision < min_precision:
                continue
            selected.append(item)
            if len(selected) >= top_n:
                break

        per_new_selected[ng] = selected

    # conflicts: old group used as top-1 by multiple new groups
    old_top1_counter: Counter = Counter()
    for ng, selected in per_new_selected.items():
        if not selected:
            continue
        top1_old = selected[0][0]
        old_top1_counter[top1_old] += 1

    # flatten to rows
    for ng, selected in per_new_selected.items():
        if not selected:
            continue
        has_multi_candidates = len(selected) > 1
        top1_old = selected[0][0]
        for rank, (og, inter, precision, recall, jaccard, n_hosts, o_hosts) in enumerate(selected, start=1):
            used_old_in_etalon.add(og)
            env_new_top, env_new_multi = _env_top(env_by_new.get(ng, Counter()))
            env_old_top, env_old_multi = _env_top(env_by_old.get(og, Counter()))
            env_mismatch = ""
            if env_new_top and env_old_top and str(env_new_top).strip().upper() != str(env_old_top).strip().upper():
                env_mismatch = "yes"
            etalon_rows.append(
                {
                    "new_group": ng,
                    "old_group": og,
                    "rank": rank,
                    "is_top1": "yes" if rank == 1 else "",
                    "top1_old": top1_old,
                    "old_top1_conflict": "yes" if (rank == 1 and old_top1_counter[top1_old] > 1) else "",
                    "new_has_multi_candidates": "yes" if has_multi_candidates else "",
                    "intersection": inter,
                    "hosts_in_new": n_hosts,
                    "hosts_in_old": o_hosts,
                    "precision": round(precision, 4),
                    "recall": round(recall, 4),
                    "jaccard": round(jaccard, 4),
                    "env_new_top": env_new_top,
                    "env_old_top": env_old_top,
                    "env_new_multi": "yes" if env_new_multi else "",
                    "env_old_multi": "yes" if env_old_multi else "",
                    "env_mismatch": env_mismatch,
                }
            )

    etalon_rows.sort(key=lambda r: (r["new_group"], r["rank"]))

    # outliers
    outlier_rows: List[Dict[str, Any]] = []
    for og, hostset in hosts_by_old.items():
        if og not in used_old_in_etalon:
            outlier_rows.append({"old_group": og, "hosts": len(hostset), "note": "not in etalon"})
    outlier_rows.sort(key=lambda r: r["hosts"], reverse=True)

    return etalon_rows, outlier_rows, hosts_by_new, hosts_by_old


# =========================
# Actions (поиск упоминаний групп)
# =========================
def _recursive_collect_groupids(obj: Any, hits: Set[str]) -> None:
    if isinstance(obj, dict):
        for k, v in obj.items():
            if k == "groupid" and v is not None:
                hits.add(str(v))
            else:
                _recursive_collect_groupids(v, hits)
    elif isinstance(obj, list):
        for x in obj:
            _recursive_collect_groupids(x, hits)



def action_mentions_groupids(action: Dict[str, Any], relevant_groupids: Set[str]) -> Tuple[bool, str, Set[str]]:
    """Возвращает (mentions?, where, matched_groupids) где where: 'conditions', 'operations', 'both'."""
    where = set()
    matched: Set[str] = set()

    # условия
    flt = action.get("filter") or {}
    for c in flt.get("conditions") or []:
        val = c.get("value")
        if val is None:
            continue
        v = str(val)
        if v in relevant_groupids:
            where.add("conditions")
            matched.add(v)

    # операции: ищем groupid рекурсивно
    op_hits: Set[str] = set()
    for ops_key in ("operations", "recovery_operations", "update_operations"):
        _recursive_collect_groupids(action.get(ops_key), op_hits)
    op_matched = op_hits.intersection(relevant_groupids)
    if op_matched:
        where.add("operations")
        matched.update(op_matched)

    if not where:
        return False, "", set()
    if where == {"conditions"}:
        return True, "conditions", matched
    if where == {"operations"}:
        return True, "operations", matched
    return True, "both", matched



def extract_action_recipients(action: Dict[str, Any]) -> Tuple[Set[str], Set[str]]:
    usrgrpids: Set[str] = set()
    userids: Set[str] = set()

    def scan_ops(ops: Any) -> None:
        for op in ops or []:
            for grp in (op.get("opmessage_grp") or []):
                ugid = grp.get("usrgrpid")
                if ugid is not None:
                    usrgrpids.add(str(ugid))
            for usr in (op.get("opmessage_usr") or []):
                uid = usr.get("userid")
                if uid is not None:
                    userids.add(str(uid))

    scan_ops(action.get("operations"))
    scan_ops(action.get("recovery_operations"))
    scan_ops(action.get("update_operations"))

    return usrgrpids, userids


def extract_active_media_sendto(medias: List[Dict[str, Any]]) -> List[str]:
    """Возвращает список sendto для активных media (active==0)."""
    out: List[str] = []
    for m in medias or []:
        active = m.get("active")
        if active is None:
            active = m.get("status")
        if str(active) not in ("0", "False", "false", "active"):
            continue
        sendto = m.get("sendto")
        if sendto is None:
            continue
        s = str(sendto).strip()
        if s:
            out.append(s)
    return out


def preprocess_actions(actions: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Предподготовка actions: кэшируем groupids (conditions/operations) + получателей."""
    cached: List[Dict[str, Any]] = []
    for a in actions:
        # groupids in conditions
        cond_ids: Set[str] = set()
        flt = a.get("filter") or {}
        for c in flt.get("conditions") or []:
            if c.get("value") is None:
                continue
            if str(c.get("conditiontype")) == "0":
                # host group
                cond_ids.add(str(c.get("value")))
            else:
                # иногда groupid встречается и без conditiontype==0
                if str(c.get("value")).isdigit():
                    cond_ids.add(str(c.get("value")))

        # groupids in operations/recovery/update
        op_ids: Set[str] = set()
        for ops_key in ("operations", "recovery_operations", "update_operations"):
            _recursive_collect_groupids(a.get(ops_key), op_ids)

        usrgrpids, userids = extract_action_recipients(a)

        cached.append(
            {
                "action": a,
                "cond_groupids": cond_ids,
                "op_groupids": op_ids,
                "recipient_usrgrpids": usrgrpids,
                "recipient_userids": userids,
            }
        )
    return cached


def build_mapping_rows(report_as: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Плоская таблица соответствий OLD->NEW по результатам эталона.

    Берём только top1-кандидаты (rank=1) из эталона каждой AS.
    Если один OLD встречается как top1 у нескольких NEW — помечаем ambiguous.
    """
    rows: List[Dict[str, Any]] = []

    for as_val, data in report_as.items():
        et = data.get("etalon") or []
        # top1 rows: rank==1
        top1 = [r for r in et if int(r.get("rank") or 0) == 1]
        if not top1:
            continue

        # group by old_group
        by_old: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
        for r in top1:
            og = str(r.get("old_group") or "").strip()
            if not og:
                continue
            by_old[og].append(r)

        for og, lst in by_old.items():
            # choose best by jaccard then intersection
            lst_sorted = sorted(
                lst,
                key=lambda x: (float(x.get("jaccard") or 0.0), int(x.get("intersection") or 0)),
                reverse=True,
            )
            best = lst_sorted[0]
            new_best = str(best.get("new_group") or "")

            # ambiguous if multiple NEW for same OLD
            ambiguous = "yes" if len(lst_sorted) > 1 else ""
            others = ", ".join(sorted({str(x.get("new_group") or "") for x in lst_sorted[1:] if x.get("new_group")}))

            rows.append(
                {
                    "AS": as_val,
                    "old_group": og,
                    "new_group": new_best,
                    "jaccard": best.get("jaccard"),
                    "precision": best.get("precision"),
                    "intersection": best.get("intersection"),
                    "hosts_in_new": best.get("hosts_in_new"),
                    "hosts_in_old": best.get("hosts_in_old"),
                    "old_top1_conflict": best.get("old_top1_conflict"),
                    "ambiguous_old_to_many_new": ambiguous,
                    "other_new_top1": others,
                    "env_new_top": best.get("env_new_top"),
                    "env_old_top": best.get("env_old_top"),
                    "env_new_multi": best.get("env_new_multi"),
                    "env_old_multi": best.get("env_old_multi"),
                    "env_mismatch": best.get("env_mismatch"),
                }
            )

    # stable order
    rows.sort(key=lambda r: (str(r.get("AS") or ""), str(r.get("old_group") or "")))
    return rows

# =========================
# Workbook builder
# =========================
def _build_workbook_single(
    as_items: List[Tuple[str, Any]],
    unknown_hosts_rows: List[Dict[str, Any]],
    out_path: str,
    mapping_rows: Optional[List[Dict[str, Any]]] = None,
) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    # UNKNOWN sheet
    ws_u = wb.create_sheet(CONFIG.excel.sheet_unknown)
    ws_u.append(["host_name", "hostid", CONFIG.tags.AS, CONFIG.tags.ASN, "groups", "tags_json"])
    for r in unknown_hosts_rows:
        ws_u.append(
            [
                r.get("host_name"),
                r.get("hostid"),
                r.get("AS"),
                r.get("ASN"),
                r.get("groups"),
                r.get("tags_json"),
            ]
        )
    autosize_columns(ws_u)

    # MAPPING sheet (для Grafana/рефакторинга): OLD -> NEW (top1 + ambiguous)
    if mapping_rows is not None:
        sheet_name = getattr(CONFIG.excel, "sheet_mapping", "MAPPING")
        ws_m = wb.create_sheet(sheet_name)
        ws_m.append(
            [
                "AS",
                "old_group",
                "new_group",
                "jaccard",
                "precision",
                "intersection",
                "hosts_in_new",
                "hosts_in_old",
                "old_top1_conflict",
                "ambiguous_old_to_many_new",
                "other_new_top1",
                "env_new_top",
                "env_old_top",
                "env_new_multi",
                "env_old_multi",
                "env_mismatch",
            ]
        )
        for r0 in mapping_rows:
            ws_m.append(
                [
                    r0.get("AS"),
                    r0.get("old_group"),
                    r0.get("new_group"),
                    r0.get("jaccard"),
                    r0.get("precision"),
                    r0.get("intersection"),
                    r0.get("hosts_in_new"),
                    r0.get("hosts_in_old"),
                    r0.get("old_top1_conflict"),
                    r0.get("ambiguous_old_to_many_new"),
                    r0.get("other_new_top1"),
                    r0.get("env_new_top"),
                    r0.get("env_old_top"),
                    r0.get("env_new_multi"),
                    r0.get("env_old_multi"),
                    r0.get("env_mismatch"),
                ]
            )
        autosize_columns(ws_m)

    for as_val, data in as_items:
        title = safe_sheet_title(as_val, max_len=int(CONFIG.excel.sheet_name_max))
        ws = wb.create_sheet(title)

        r = 1
        r = write_section_title(ws, r, f"AS = {as_val}")

        # 0) ASN
        r = write_section_title(ws, r, "0) ASN, найденные на хостах этой AS")
        ws.append(["Всего хостов", data.get("hosts_total")])
        ws.append([])
        ws.append(["ASN", "hosts", "share"])
        for row in data.get("asn") or []:
            ws.append([row.get("asn"), row.get("hosts"), row.get("share")])
        r = ws.max_row + 2

        # 1) Etalon
        r = write_section_title(ws, r, "1) Эталон: new → old (частотное соответствие)")
        ws.append(
            [
                "new_group",
                "old_group",
                "rank",
                "is_top1",
                "top1_old",
                "old_top1_conflict",
                "new_has_multi_candidates",
                "intersection",
                "hosts_in_new",
                "hosts_in_old",
                "precision",
                "recall",
                "jaccard",
                "env_new_top",
                "env_old_top",
                "env_new_multi",
                "env_old_multi",
                "env_mismatch",
            ]
        )
        for row in data["etalon"]:
            ws.append(
                [
                    row["new_group"],
                    row["old_group"],
                    row["rank"],
                    row["is_top1"],
                    row["top1_old"],
                    row["old_top1_conflict"],
                    row["new_has_multi_candidates"],
                    row["intersection"],
                    row["hosts_in_new"],
                    row["hosts_in_old"],
                    row["precision"],
                    row["recall"],
                    row["jaccard"],
                    row.get("env_new_top"),
                    row.get("env_old_top"),
                    row.get("env_new_multi"),
                    row.get("env_old_multi"),
                    row.get("env_mismatch"),
                ]
            )
        r = ws.max_row + 2

        # 2) Outliers
        r = write_section_title(ws, r, "2) Не вошло в эталон (старые группы / выбросы)")
        ws.append(["old_group", "hosts", "note"])
        for row in data["outliers"]:
            ws.append([row["old_group"], row["hosts"], row["note"]])
        r = ws.max_row + 2

        # 3) Trigger actions
        r = write_section_title(ws, r, "3) Trigger actions (eventsource=0), затрагивающие группы этой AS")
        ws.append([
            "actionid",
            "name",
            "status",
            "where_found",
            "matched_groups",
            "recipients_usergroups",
            "recipients_users",
            "recipients_media",
        ])
        for a in data["actions"]:
            ws.append(
                [
                    a.get("actionid"),
                    a.get("name"),
                    a.get("status"),
                    a.get("where_found"),
                    a.get("matched_groups"),
                    a.get("recipients_usergroups"),
                    a.get("recipients_users"),
                    a.get("recipients_media"),
                ]
            )
        r = ws.max_row + 2

        # 4) Permissions
        r = write_section_title(ws, r, "4) User groups / Users / Permissions (host groups + tag-based)")
        ws.append(["usrgrpid", "usergroup", "rights_on_relevant_groups", "tag_filters", "users", "users_media"])
        for ug in data["permissions"]:
            ws.append([ug["usrgrpid"], ug["name"], ug["rights"], ug["tag_filters"], ug["users"], ug.get("users_media")])
        r = ws.max_row + 2

        # 5) Grafana
        r = write_section_title(ws, r, "5) Grafana dashboards (поиск OLD/NEW host-groups)")
        ws.append(["dashboard_uid", "dashboard_title", "matched_string", "match_type", "count"])
        for row in data.get("grafana") or []:
            ws.append(
                [
                    row.get("dashboard_uid"),
                    row.get("dashboard_title"),
                    row.get("matched_string"),
                    row.get("match_type"),
                    row.get("count"),
                ]
            )

        autosize_columns(ws)

    wb.save(out_path)


def build_workbooks(
    report_as: Dict[str, Any],
    unknown_hosts_rows: List[Dict[str, Any]],
    out_path: str,
) -> List[str]:
    """Пишет один или несколько xlsx.

    Если AS много — режем отчёт на части по N листов.
    """

    # Excel hard/soft limit: обычно 255 листов, плюс 1 лист UNKNOWN.
    max_sheets = int(getattr(CONFIG.excel, "max_sheets_per_workbook", 200))
    if max_sheets < 10:
        max_sheets = 10

    items = list(report_as.items())

    mapping_rows_full = build_mapping_rows(report_as)

    if not items:
        _build_workbook_single([], unknown_hosts_rows, out_path, mapping_rows=[])
        return [out_path]

    # Split into chunks
    written: List[str] = []

    base = out_path
    if base.lower().endswith(".xlsx"):
        base_no_ext = base[:-5]
        ext = ".xlsx"
    else:
        base_no_ext = base
        ext = ".xlsx"

    # Put UNKNOWN sheet only into the first workbook to avoid duplication
    for part_idx in range(0, len(items), max_sheets):
        chunk = items[part_idx : part_idx + max_sheets]
        part_no = part_idx // max_sheets + 1

        if len(items) <= max_sheets:
            part_path = out_path
        else:
            part_path = f"{base_no_ext}_part{part_no:03d}{ext}"

        unknown_for_this = unknown_hosts_rows if part_no == 1 else []
        mapping_for_this = mapping_rows_full if part_no == 1 else None
        _build_workbook_single(chunk, unknown_for_this, part_path, mapping_rows=mapping_for_this)
        written.append(part_path)

    return written


# =========================
# Main / Runner
# =========================
def run_audit(as_filter: Optional[Iterable[str]] = None, output_xlsx: Optional[str] = None) -> List[str]:
    conn = load_connection_from_env_or_prompt(interactive=False)

    api = ZabbixAPI(conn.api_url, timeout_sec=int(CONFIG.runtime.http_timeout_sec))
    api.login(conn.username, conn.password)

    print("Fetching hosts...")
    hosts = fetch_hosts(api)
    print(f"Hosts fetched: {len(hosts)}")

    by_as, unknown_hosts = build_as_index(hosts)

    # collect unknown rows for sheet
    unknown_rows: List[Dict[str, Any]] = []
    for h in unknown_hosts:
        host_name = h.get("name") or h.get("host") or h.get("hostid")
        tags = {t.get("tag"): t.get("value") for t in (h.get("tags") or [])}
        groups = ", ".join(
            sorted(
                g.get("name")
                for g in (h.get("groups") or [])
                if g.get("name") and not is_excluded_group(g.get("name"))
            )
        )
        unknown_rows.append(
            {
                "host_name": host_name,
                "hostid": h.get("hostid"),
                "AS": tags.get(CONFIG.tags.AS),
                "ASN": tags.get(CONFIG.tags.ASN),
                "groups": groups,
                "tags_json": json.dumps(tags, ensure_ascii=False),
            }
        )

    # prefetch actions + usergroups once
    print("Fetching trigger actions (eventsource=0)...")
    actions = fetch_trigger_actions(api)
    print(f"Actions fetched: {len(actions)}")

    print("Fetching user groups (rights + tag filters + users)...")
    usergroups = fetch_usergroups_with_rights(api)
    print(f"User groups fetched: {len(usergroups)}")

    print("Fetching users...")
    users = fetch_users(api)
    print(f"Users fetched: {len(users)}")

    user_by_id: Dict[str, str] = {}
    user_media_by_id: Dict[str, List[str]] = {}
    for u in users:
        uid = u.get("userid")
        if uid is None:
            continue
        uname = u.get("username") or u.get("alias") or ""
        full = f"{(u.get('name') or '').strip()} {(u.get('surname') or '').strip()}".strip()
        if uname and full:
            user_by_id[str(uid)] = f"{uname} ({full})"
        elif uname:
            user_by_id[str(uid)] = str(uname)
        else:
            user_by_id[str(uid)] = f"userid={uid}"
        user_media_by_id[str(uid)] = extract_active_media_sendto(u.get("medias") or [])

    usrgrp_by_id: Dict[str, str] = {str(ug.get("usrgrpid")): str(ug.get("name")) for ug in usergroups if ug.get("usrgrpid") is not None and ug.get("name")}
    usrgrp_userids: Dict[str, Set[str]] = defaultdict(set)
    for ug in usergroups:
        gid = ug.get("usrgrpid")
        if gid is None:
            continue
        for u in (ug.get("users") or []):
            if u.get("userid") is None:
                continue
            usrgrp_userids[str(gid)].add(str(u.get("userid")))

    actions_cached = preprocess_actions(actions)

    # build name->id mapping from hosts (enough for our needs)
    groupname_to_id: Dict[str, str] = {}
    groupid_to_name: Dict[str, str] = {}
    for h in hosts:
        for g in (h.get("groups") or []):
            gid = str(g.get("groupid"))
            gname = g.get("name")
            if not gid or not gname:
                continue
            groupname_to_id[gname] = gid
            groupid_to_name[gid] = gname

    # report per AS
    report_as: Dict[str, Any] = {}

    as_items = list(by_as.items())
    if as_filter:
        lower_filter = {str(x).strip().lower() for x in as_filter if str(x).strip()}
        as_items = [(k, v) for (k, v) in as_items if str(k).strip().lower() in lower_filter]
        # В скоупе по AS UNKNOWN-хосты неинформативны
        unknown_rows = []
    if CONFIG.runtime.limit_as:
        as_items = as_items[: int(CONFIG.runtime.limit_as)]

    print(f"AS count (after UNKNOWN filtering): {len(as_items)}")

    for idx, (as_val, as_hosts) in enumerate(as_items, start=1):
        if idx % 25 == 0:
            print(f"  processing AS {idx}/{len(as_items)}...")

        # ASN в этой AS (для прав/фильтров + отчёта)
        asn_counter: Counter[str] = Counter()
        for h in as_hosts:
            tags = h.get("tags") or []
            asn_val = get_tag_value(tags, CONFIG.tags.ASN)
            if asn_val:
                asn_counter[str(asn_val)] += 1

        asn_values_in_bucket: Set[str] = set(asn_counter.keys())
        as_hosts_total = len(as_hosts)
        asn_rows = [
            {
                "asn": asn,
                "hosts": cnt,
                "share": round((cnt / as_hosts_total), 4) if as_hosts_total else 0.0,
            }
            for asn, cnt in asn_counter.most_common()
        ]

        etalon, outliers, hosts_by_new, hosts_by_old = compute_mapping_for_as(as_val, as_hosts)

        # relevant groupids for this AS = all groups seen on its hosts (both old & new)
        relevant_groupids: Set[str] = set()
        for gname in list(hosts_by_new.keys()) + list(hosts_by_old.keys()):
            gid = groupname_to_id.get(gname)
            if gid:
                relevant_groupids.add(str(gid))

        # actions that mention these groups
        as_actions: List[Dict[str, Any]] = []
        if relevant_groupids:
            for ac in actions_cached:
                a = ac["action"]
                cond_matched = ac["cond_groupids"].intersection(relevant_groupids)
                op_matched = ac["op_groupids"].intersection(relevant_groupids)

                if not cond_matched and not op_matched:
                    continue

                if cond_matched and op_matched:
                    where_found = "both"
                elif cond_matched:
                    where_found = "conditions"
                else:
                    where_found = "operations"

                matched = set()
                matched.update(cond_matched)
                matched.update(op_matched)

                matched_groups = ", ".join(
                    sorted(
                        f"{gid}:{groupid_to_name.get(gid, '')}" if groupid_to_name.get(gid) else str(gid)
                        for gid in matched
                    )
                )

                # recipients: ids -> names
                rec_usrgrp = ", ".join(
                    sorted(
                        f"{ugid}:{usrgrp_by_id.get(ugid, '')}" if usrgrp_by_id.get(ugid) else str(ugid)
                        for ugid in (ac.get("recipient_usrgrpids") or set())
                    )
                )
                rec_users = ", ".join(
                    sorted(
                        f"{uid}:{user_by_id.get(uid, '')}" if user_by_id.get(uid) else str(uid)
                        for uid in (ac.get("recipient_userids") or set())
                    )
                )
                recipient_userids: Set[str] = set(ac.get("recipient_userids") or set())
                for ugid in (ac.get("recipient_usrgrpids") or set()):
                    recipient_userids.update(usrgrp_userids.get(str(ugid), set()))

                recipients_media_set: Set[str] = set()
                for uid in recipient_userids:
                    for sendto in user_media_by_id.get(str(uid), []) or []:
                        recipients_media_set.add(str(sendto))
                recipients_media = ", ".join(sorted(recipients_media_set))

                as_actions.append(
                    {
                        "actionid": a.get("actionid"),
                        "name": a.get("name"),
                        "status": a.get("status"),
                        "where_found": where_found,
                        "matched_groups": matched_groups,
                        "recipients_usergroups": rec_usrgrp,
                        "recipients_users": rec_users,
                        "recipients_media": recipients_media,
                    }
                )

        # permissions (only those touching relevant groups OR having tag filters mentioning this AS/ASN)
        perms_rows: List[Dict[str, Any]] = []
        for ug in usergroups:
            rights = ug.get("hostgroup_rights") or []
            touched = [r for r in rights if _get_right_groupid(r) in relevant_groupids]

            tag_filters = ug.get("tag_filters") or []

            # Keep usergroup only if it either touches relevant host-groups by explicit rights,
            # or its tag-filters mention this AS/ASN value.
            tf_mentions_this_as = False
            for tf in tag_filters:
                t = _get_tagfilter_tag(tf)
                v = _get_tagfilter_value(tf)
                if not t or v is None:
                    continue
                if t == CONFIG.tags.AS and str(v).strip().lower() == str(as_val).strip().lower():
                    tf_mentions_this_as = True
                    break
                if t == CONFIG.tags.ASN and str(v) in asn_values_in_bucket:
                    tf_mentions_this_as = True
                    break

            if not touched and not tf_mentions_this_as:
                continue

            rights_str = ""
            if touched:
                parts = []
                for rgt in touched:
                    gid = _get_right_groupid(rgt)
                    gname = groupid_to_name.get(gid, "")
                    perm = permission_name(rgt.get("permission"))
                    parts.append(f"{gname or gid}:{perm}")
                rights_str = "; ".join(parts)

            users = ug.get("users") or []
            user_chunks: List[str] = []
            users_media_set: Set[str] = set()
            for u in users:
                uname = u.get("username") or u.get("alias") or ""
                full = f"{(u.get('name') or '').strip()} {(u.get('surname') or '').strip()}".strip()
                if uname and full:
                    user_chunks.append(f"{uname} ({full})")
                elif uname:
                    user_chunks.append(uname)
                elif u.get("userid") is not None:
                    user_chunks.append(f"userid={u.get('userid')}")
                if u.get("userid") is not None:
                    for sendto in user_media_by_id.get(str(u.get("userid")), []) or []:
                        users_media_set.add(str(sendto))
            users_str = ", ".join(user_chunks)
            users_media_str = ", ".join(sorted(users_media_set))

            perms_rows.append(
                {
                    "usrgrpid": ug.get("usrgrpid"),
                    "name": ug.get("name"),
                    "rights": rights_str,
                    "tag_filters": json.dumps(tag_filters, ensure_ascii=False) if tag_filters else "",
                    "users": users_str,
                    "users_media": users_media_str,
                }
            )

        report_as[as_val] = {
            "asn": asn_rows,
            "hosts_total": as_hosts_total,
            "etalon": etalon,
            "outliers": outliers,
            "actions": as_actions,
            "permissions": perms_rows,
            "groups_old": list(hosts_by_old.keys()),
            "groups_new": list(hosts_by_new.keys()),
        }

    out_path = output_xlsx or CONFIG.excel.output_xlsx

    seed_mapping_rows = build_mapping_rows(report_as)

    # Сохранить seed для Grafana-only аудита
    if CONFIG.runtime.save_zabbix_seed_on_audit:
        seed_path = CONFIG.runtime.zabbix_seed_path or build_seed_path(out_path)
        print(f"Saving Zabbix seed: {seed_path}")
        save_zabbix_seed(report_as, unknown_rows, seed_mapping_rows, seed_path)

    # Сохранить план миграции (полуручной)
    plan_path = CONFIG.excel.migration_plan_path or build_migration_plan_path(out_path)
    print(f"Saving migration plan: {plan_path}")
    save_migration_plan(seed_mapping_rows, plan_path, [as_val for as_val, _h in as_items])

    # Бэкап при аудите (если включено)
    if CONFIG.runtime.create_backup_on_audit:
        if not as_items:
            raise RuntimeError("Backup scope is empty. Provide AS list for audit.")
        try:
            from make_backup import build_backup_filename, create_backup  # локальный импорт для избежания циклов
        except Exception as e:
            raise RuntimeError(f"Backup module import failed: {e}") from e

        scope_values = [as_val for (as_val, _hosts) in as_items]
        out_base = output_xlsx or CONFIG.excel.output_xlsx
        backup_path = build_backup_filename(scope_values, base_path=out_base)
        print(f"Creating backup: {backup_path}")
        create_backup(
            api,
            scope_values,
            backup_path,
            hosts=hosts,
            actions=actions,
            usergroups=usergroups,
            users=users,
        )

    # Grafana аудит
    if CONFIG.runtime.enable_grafana_audit and report_as:
        try:
            print("Fetching Grafana dashboards...")
            from grafana_audit import collect_grafana_matches  # локальный импорт

            grafana_conn = load_grafana_from_module()
            grafana_rows_by_as = collect_grafana_matches(
                grafana_conn, report_as, scope_as=[as_val for as_val, _h in as_items]
            )
            for as_val, rows in grafana_rows_by_as.items():
                if as_val in report_as:
                    report_as[as_val]["grafana"] = rows
        except Exception as e:
            print(f"Grafana audit failed: {e}")
    print(f"Writing Excel: {out_path}")
    written_files = build_workbooks(report_as, unknown_rows, out_path)
    if len(written_files) == 1:
        print(f"Excel saved: {written_files[0]}")
    else:
        print("Excel saved (multiple parts):")
        for p in written_files:
            print(f"  - {p}")
    print("OK")
    return written_files


def main() -> int:
    run_audit()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
