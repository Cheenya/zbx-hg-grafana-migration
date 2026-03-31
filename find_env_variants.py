from __future__ import annotations

import json
import os
from collections import defaultdict
from datetime import datetime
from typing import Any, Dict, List, Sequence, Set

from openpyxl import Workbook  # type: ignore

import config
from api_clients import ZabbixAPI
from common import autosize_columns, canonical_env_value, join_sorted, sample_host_names


OUTPUT_PREFIX = "env_variants"


def fetch_hosts(api: ZabbixAPI) -> List[Dict[str, Any]]:
    return api.call(
        "host.get",
        {
            "output": ["hostid", "host", "name", "status"],
            "selectTags": ["tag", "value"],
        },
    )


def _status_label(status: Any) -> str:
    return "enabled" if str(status or "") == "0" else "disabled"


def _raw_env_values(tags: Sequence[Dict[str, Any]]) -> List[str]:
    values: List[str] = []
    for tag in tags or []:
        if str(tag.get("tag") or "") != config.TAG_ENV:
            continue
        value = tag.get("value")
        values.append("" if value is None else str(value))
    return values


def _json_string(value: str) -> str:
    return json.dumps(value, ensure_ascii=False)


def build_report(hosts: Sequence[Dict[str, Any]]) -> Dict[str, Any]:
    value_bucket: Dict[str, Dict[str, Any]] = defaultdict(
        lambda: {
            "hostids": set(),
            "host_names": set(),
            "enabled_hostids": set(),
            "disabled_hostids": set(),
            "tag_instances": 0,
        }
    )
    host_rows: List[Dict[str, Any]] = []
    missing_rows: List[Dict[str, Any]] = []
    blank_rows: List[Dict[str, Any]] = []
    multi_tag_rows: List[Dict[str, Any]] = []

    for host in hosts:
        hostid = str(host.get("hostid") or "")
        host_name = str(host.get("name") or host.get("host") or "")
        status = str(host.get("status") or "")
        status_label = _status_label(status)
        values = _raw_env_values(host.get("tags") or [])

        if not values:
            missing_rows.append(
                {
                    "hostid": hostid,
                    "host": str(host.get("host") or ""),
                    "name": host_name,
                    "status": status,
                    "status_label": status_label,
                    "issue": "missing ENV tag",
                }
            )
            continue

        if len(values) > 1:
            multi_tag_rows.append(
                {
                    "hostid": hostid,
                    "host": str(host.get("host") or ""),
                    "name": host_name,
                    "status": status,
                    "status_label": status_label,
                    "env_values_json": join_sorted(_json_string(item) for item in values),
                    "env_values_trimmed": join_sorted(item.strip() for item in values),
                }
            )

        for index, raw_value in enumerate(values, start=1):
            trimmed = raw_value.strip()
            canonical = canonical_env_value(raw_value) if trimmed else ""
            bucket = value_bucket[raw_value]
            bucket["hostids"].add(hostid)
            bucket["host_names"].add(host_name)
            bucket["tag_instances"] += 1
            if status_label == "enabled":
                bucket["enabled_hostids"].add(hostid)
            else:
                bucket["disabled_hostids"].add(hostid)

            host_row = {
                "hostid": hostid,
                "host": str(host.get("host") or ""),
                "name": host_name,
                "status": status,
                "status_label": status_label,
                "tag_index": index,
                "env_value_exact": raw_value,
                "env_value_json": _json_string(raw_value),
                "length": len(raw_value),
                "trimmed_value": trimmed,
                "upper_trimmed_value": trimmed.upper(),
                "canonical_env_scope": canonical,
            }
            host_rows.append(host_row)

            if not trimmed:
                blank_rows.append(
                    {
                        **host_row,
                        "issue": "blank ENV value",
                    }
                )

    value_rows: List[Dict[str, Any]] = []
    for raw_value, bucket in sorted(value_bucket.items(), key=lambda item: (item[0].strip().upper(), item[0])):
        hosts_count = len(bucket["hostids"])
        disabled_count = len(bucket["disabled_hostids"])
        trimmed = raw_value.strip()
        canonical = canonical_env_value(raw_value) if trimmed else ""
        value_rows.append(
            {
                "env_value_exact": raw_value,
                "env_value_json": _json_string(raw_value),
                "length": len(raw_value),
                "trimmed_value": trimmed,
                "upper_trimmed_value": trimmed.upper(),
                "canonical_env_scope": canonical,
                "hosts_count": hosts_count,
                "enabled_hosts": hosts_count - disabled_count,
                "disabled_hosts": disabled_count,
                "tag_instances": int(bucket["tag_instances"]),
                "sample_hosts": sample_host_names(bucket["host_names"], 20),
                "sample_hostids": join_sorted(sorted(bucket["hostids"])[:20]),
            }
        )

    summary = {
        "hosts_total": len(hosts),
        "hosts_with_env_tag": len({row["hostid"] for row in host_rows}),
        "hosts_without_env_tag": len(missing_rows),
        "hosts_with_blank_env": len({row["hostid"] for row in blank_rows}),
        "hosts_with_multiple_env_tags": len(multi_tag_rows),
        "unique_exact_env_values": len(value_rows),
    }

    return {
        "summary": summary,
        "values": value_rows,
        "host_values": sorted(host_rows, key=lambda row: (row["upper_trimmed_value"], row["env_value_exact"], row["name"].lower(), row["hostid"])),
        "missing": missing_rows,
        "blank": blank_rows,
        "multi": multi_tag_rows,
    }


def write_workbook(data: Dict[str, Any], path: str) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    summary_ws = wb.create_sheet("SUMMARY")
    summary_ws.append(["key", "value"])
    for key, value in data["summary"].items():
        summary_ws.append([key, value])
    autosize_columns(summary_ws)

    values_ws = wb.create_sheet("ENV_VALUES")
    values_headers = [
        "env_value_exact",
        "env_value_json",
        "length",
        "trimmed_value",
        "upper_trimmed_value",
        "canonical_env_scope",
        "hosts_count",
        "enabled_hosts",
        "disabled_hosts",
        "tag_instances",
        "sample_hosts",
        "sample_hostids",
    ]
    values_ws.append(values_headers)
    for row in data["values"]:
        values_ws.append([row.get(header, "") for header in values_headers])
    autosize_columns(values_ws)

    host_ws = wb.create_sheet("HOST_VALUES")
    host_headers = [
        "hostid",
        "host",
        "name",
        "status",
        "status_label",
        "tag_index",
        "env_value_exact",
        "env_value_json",
        "length",
        "trimmed_value",
        "upper_trimmed_value",
        "canonical_env_scope",
    ]
    host_ws.append(host_headers)
    for row in data["host_values"]:
        host_ws.append([row.get(header, "") for header in host_headers])
    autosize_columns(host_ws)

    missing_ws = wb.create_sheet("MISSING_ENV")
    missing_headers = ["hostid", "host", "name", "status", "status_label", "issue"]
    missing_ws.append(missing_headers)
    for row in data["missing"]:
        missing_ws.append([row.get(header, "") for header in missing_headers])
    autosize_columns(missing_ws)

    blank_ws = wb.create_sheet("BLANK_ENV")
    blank_headers = ["hostid", "host", "name", "status", "status_label", "env_value_exact", "env_value_json", "length", "issue"]
    blank_ws.append(blank_headers)
    for row in data["blank"]:
        blank_ws.append([row.get(header, "") for header in blank_headers])
    autosize_columns(blank_ws)

    multi_ws = wb.create_sheet("MULTI_ENV_TAGS")
    multi_headers = ["hostid", "host", "name", "status", "status_label", "env_values_json", "env_values_trimmed"]
    multi_ws.append(multi_headers)
    for row in data["multi"]:
        multi_ws.append([row.get(header, "") for header in multi_headers])
    autosize_columns(multi_ws)

    wb.save(path)


def build_output_path() -> str:
    os.makedirs(config.OUTPUT_DIR, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return os.path.join(config.OUTPUT_DIR, f"{OUTPUT_PREFIX}_{timestamp}.xlsx")


def main() -> int:
    connection = config.load_zabbix_connection()
    api = ZabbixAPI(connection.api_url, timeout_sec=int(config.HTTP_TIMEOUT_SEC))
    api.login(connection.username, connection.password)

    print("Fetching hosts...")
    hosts = fetch_hosts(api)
    print(f"Hosts fetched: {len(hosts)}")

    report = build_report(hosts)
    out_path = build_output_path()
    write_workbook(report, out_path)

    print(f"ENV variants workbook saved: {out_path}")
    print(f"Unique exact ENV values: {report['summary']['unique_exact_env_values']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
