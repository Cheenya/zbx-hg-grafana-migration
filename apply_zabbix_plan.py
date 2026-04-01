#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Применение Zabbix impact plan: пока только донасыщение host groups на хостах."""

from __future__ import annotations

import argparse
import json
from datetime import datetime
from typing import Any, Dict, List, Sequence, Set, Tuple

from openpyxl import Workbook  # type: ignore

import config
from api_clients import ZabbixAPI
from backup_io import load_backup
from common import (
    autosize_columns,
    build_artifact_path,
    join_sorted,
    normalize_scope_env,
    normalize_values,
    resolve_input_artifact,
)


def load_impact_plan(path: str) -> Dict[str, Any]:
    with open(path, "r", encoding="utf-8") as handle:
        return json.load(handle)


def fetch_hosts(api: ZabbixAPI, hostids: Sequence[str]) -> List[Dict[str, Any]]:
    if not hostids:
        return []
    return api.call(
        "host.get",
        {
            "output": ["hostid", "host", "name", "status"],
            "hostids": list(hostids),
            "selectGroups": ["groupid", "name"],
        },
    )


def _validate_backup_scope(backup_path: str, impact_plan: Dict[str, Any], api: ZabbixAPI) -> None:
    data = load_backup(backup_path)
    summary = impact_plan.get("summary") or {}
    impact_scope_as = normalize_values(summary.get("scope_as") or [])
    impact_scope_env = normalize_scope_env(summary.get("scope_env") or "")
    impact_scope_gas = normalize_values(summary.get("scope_gas") or [])

    if str(data.meta.zabbix_url or "").strip() and str(data.meta.zabbix_url).strip() != str(getattr(api, "api_url", "")).strip():
        raise RuntimeError(
            f"Backup Zabbix URL mismatch: backup={data.meta.zabbix_url} current={getattr(api, 'api_url', '')}"
        )

    backup_scope_as = normalize_values(data.meta.scope_as)
    if sorted(item.lower() for item in backup_scope_as) != sorted(item.lower() for item in impact_scope_as):
        raise RuntimeError(f"Backup scope_as mismatch: backup={backup_scope_as} impact={impact_scope_as}")

    backup_scope_env = normalize_scope_env(data.meta.scope_env)
    if str(backup_scope_env) != str(impact_scope_env):
        raise RuntimeError(f"Backup scope_env mismatch: backup={backup_scope_env} impact={impact_scope_env}")

    backup_scope_gas = normalize_values(data.meta.scope_gas)
    if sorted(item.upper() for item in backup_scope_gas) != sorted(item.upper() for item in impact_scope_gas):
        raise RuntimeError(f"Backup scope_gas mismatch: backup={backup_scope_gas} impact={impact_scope_gas}")

    backup_hostids = {str(item.hostid).strip() for item in data.hosts if str(item.hostid).strip()}
    impact_hostids = {
        str(row.get("object_id") or "").strip()
        for row in impact_plan.get("host_enrich_plan") or []
        if str(row.get("object_id") or "").strip()
    }
    missing_hostids = sorted(impact_hostids.difference(backup_hostids))
    if missing_hostids:
        raise RuntimeError(f"Backup coverage error for hosts: missing ids: {', '.join(missing_hostids)}")


def _group_host_rows(impact_plan: Dict[str, Any]) -> Dict[str, Dict[str, Any]]:
    grouped: Dict[str, Dict[str, Any]] = {}
    for row in impact_plan.get("host_enrich_plan") or []:
        if str(row.get("change_kind") or "") != "add_group":
            continue
        hostid = str(row.get("object_id") or "").strip()
        groupid = str(row.get("new_groupid") or "").strip()
        group_name = str(row.get("new_group") or "").strip()
        if not hostid or not groupid or not group_name:
            continue

        bucket = grouped.setdefault(
            hostid,
            {
                "hostid": hostid,
                "host_name": str(row.get("object_name") or "").strip(),
                "groups": {},
                "details": {},
            },
        )
        bucket["groups"][groupid] = group_name
        detail_key = str(group_name).strip()
        if detail_key:
            bucket["details"][detail_key] = str(row.get("details") or "").strip()
    return grouped


def _assert_host_massadd_available(api: ZabbixAPI) -> None:
    """Проверяем доступность метода host.massadd для текущей роли."""
    try:
        api.call("host.massadd", {"hosts": [], "groups": []})
    except RuntimeError as exc:
        message = str(exc)
        if 'No permissions to call "host.massadd"' in message:
            raise RuntimeError(
                'Zabbix API denies "host.massadd" for this user/role. '
                'Check role API permissions for host.massadd.'
            ) from exc
        if "Zabbix API error (host.massadd):" in message:
            return
        raise


def _build_skipped_rows(impact_plan: Dict[str, Any]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for row in impact_plan.get("object_mapping_plan") or []:
        object_type = str(row.get("object_type") or "").strip()
        change_kind = str(row.get("change_kind") or "").strip()
        manual_required = str(row.get("manual_required") or "").strip().lower() == "yes"
        if manual_required:
            reason = "manual_required"
        else:
            reason = f"{object_type}:{change_kind or 'unsupported'}"
        rows.append(
            {
                "object_type": object_type,
                "object_id": str(row.get("object_id") or "").strip(),
                "object_name": str(row.get("object_name") or "").strip(),
                "field_path": str(row.get("field_path") or "").strip(),
                "change_kind": change_kind,
                "old_group": str(row.get("old_group") or "").strip(),
                "new_group": str(row.get("new_group") or "").strip(),
                "reason": reason,
                "details": str(row.get("details") or "").strip(),
            }
        )
    return rows


def _batch_massadd_rows(pending_rows: Sequence[Dict[str, Any]]) -> Dict[Tuple[str, ...], List[Dict[str, Any]]]:
    batches: Dict[Tuple[str, ...], List[Dict[str, Any]]] = {}
    for row in pending_rows:
        key = tuple(sorted(str(item).strip() for item in (row.get("add_groupids") or []) if str(item).strip()))
        if not key:
            continue
        batches.setdefault(key, []).append(row)
    return batches


def apply_host_enrichment(api: ZabbixAPI, impact_plan: Dict[str, Any], dry_run: bool) -> Dict[str, Any]:
    grouped = _group_host_rows(impact_plan)
    hostids = sorted(grouped)
    live_hosts = fetch_hosts(api, hostids)
    live_by_id = {str(row.get("hostid") or "").strip(): row for row in live_hosts if str(row.get("hostid") or "").strip()}

    results: List[Dict[str, Any]] = []
    pending_rows: List[Dict[str, Any]] = []
    applied_hosts = 0
    changed_hosts = 0
    already_hosts = 0
    missing_hosts = 0
    applied_batches = 0

    for hostid in hostids:
        plan_row = grouped[hostid]
        host = live_by_id.get(hostid)
        if not host:
            missing_hosts += 1
            results.append(
                {
                    "hostid": hostid,
                    "host": "",
                    "name": plan_row.get("host_name") or "",
                    "status": "host_not_found",
                    "mode": "DRY-RUN" if dry_run else "APPLY",
                    "current_groups": "",
                    "planned_groups": join_sorted(plan_row["groups"].values()),
                    "added_groups": "",
                    "details": "host.get did not return this hostid",
                }
            )
            continue

        current_groups = host.get("groups") or []
        current_groupids = {
            str(item.get("groupid") or "").strip()
            for item in current_groups
            if str(item.get("groupid") or "").strip()
        }
        current_group_names = [
            str(item.get("name") or "").strip()
            for item in current_groups
            if str(item.get("name") or "").strip()
        ]
        planned_groupids = sorted(plan_row["groups"])
        add_groupids = [groupid for groupid in planned_groupids if groupid not in current_groupids]
        add_group_names = [plan_row["groups"][groupid] for groupid in add_groupids]

        if not add_groupids:
            already_hosts += 1
            results.append(
                {
                    "hostid": hostid,
                    "host": str(host.get("host") or "").strip(),
                    "name": str(host.get("name") or "").strip(),
                    "status": "already_present",
                    "mode": "DRY-RUN" if dry_run else "APPLY",
                    "current_groups": join_sorted(current_group_names),
                    "planned_groups": join_sorted(plan_row["groups"].values()),
                    "added_groups": "",
                    "details": join_sorted(plan_row["details"].values()),
                }
            )
            continue

        pending_rows.append(
            {
                "hostid": hostid,
                "host": str(host.get("host") or "").strip(),
                "name": str(host.get("name") or "").strip(),
                "status": "dry_run_add" if dry_run else "pending_apply",
                "mode": "DRY-RUN" if dry_run else "APPLY",
                "current_groups": join_sorted(current_group_names),
                "planned_groups": join_sorted(plan_row["groups"].values()),
                "added_groups": join_sorted(add_group_names),
                "details": join_sorted(plan_row["details"].values()),
                "add_groupids": sorted(add_groupids),
            }
        )
        changed_hosts += 1

    if not dry_run and pending_rows:
        for add_groupids, rows in _batch_massadd_rows(pending_rows).items():
            api.call(
                "host.massadd",
                {
                    "hosts": [{"hostid": row["hostid"]} for row in rows],
                    "groups": [{"groupid": groupid} for groupid in add_groupids],
                },
            )
            applied_batches += 1
            applied_hosts += len(rows)
            for row in rows:
                row["status"] = "applied"

    for row in pending_rows:
        row.pop("add_groupids", None)
        results.append(row)

    skipped_rows = _build_skipped_rows(impact_plan)
    summary = impact_plan.get("summary") or {}
    return {
        "meta": {
            "created_at": datetime.now().isoformat(timespec="seconds"),
            "version": "2.0",
            "impact_plan_path": "",
            "zabbix_url": str(getattr(api, "api_url", "")),
            "mode": "DRY-RUN" if dry_run else "APPLY",
        },
        "summary": {
            "scope_as": summary.get("scope_as") or [],
            "scope_env": str(summary.get("scope_env") or "").strip(),
            "scope_gas": summary.get("scope_gas") or [],
            "planned_host_rows": len(impact_plan.get("host_enrich_plan") or []),
            "planned_hosts": len(hostids),
            "changed_hosts": changed_hosts,
            "applied_hosts": applied_hosts,
            "applied_batches": applied_batches,
            "already_present_hosts": already_hosts,
            "missing_hosts": missing_hosts,
            "skipped_object_rows": len(skipped_rows),
            "mode": "DRY-RUN" if dry_run else "APPLY",
        },
        "host_results": results,
        "skipped_rows": skipped_rows,
    }


def write_apply_xlsx(data: Dict[str, Any], path: str) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    summary_ws = wb.create_sheet("SUMMARY")
    summary_ws.append(["key", "value"])
    for key, value in (data.get("summary") or {}).items():
        if isinstance(value, list):
            rendered = ", ".join(str(item) for item in value)
        else:
            rendered = value
        summary_ws.append([key, rendered])
    autosize_columns(summary_ws)

    hosts_ws = wb.create_sheet("HOST_APPLY")
    host_headers = [
        "hostid",
        "host",
        "name",
        "status",
        "mode",
        "current_groups",
        "planned_groups",
        "added_groups",
        "details",
    ]
    hosts_ws.append(host_headers)
    for row in data.get("host_results") or []:
        hosts_ws.append([row.get(header, "") for header in host_headers])
    autosize_columns(hosts_ws)

    skipped_ws = wb.create_sheet("SKIPPED_OBJECTS")
    skipped_headers = [
        "object_type",
        "object_id",
        "object_name",
        "field_path",
        "change_kind",
        "old_group",
        "new_group",
        "reason",
        "details",
    ]
    skipped_ws.append(skipped_headers)
    for row in data.get("skipped_rows") or []:
        skipped_ws.append([row.get(header, "") for header in skipped_headers])
    autosize_columns(skipped_ws)

    wb.save(path)


def main() -> int:
    parser = argparse.ArgumentParser(description="Apply Zabbix host enrichment from impact plan")
    parser.add_argument("--out-xlsx", dest="out_xlsx", help="Path to Zabbix apply XLSX")
    parser.add_argument("--out-json", dest="out_json", help="Path to Zabbix apply JSON")
    args = parser.parse_args()

    impact_plan_path = resolve_input_artifact(
        config.SOURCE_IMPACT_PLAN_JSON,
        config.IMPACT_PLAN_PREFIX,
        ".json",
        scope_as=config.SCOPE_AS,
        scope_env=config.SCOPE_ENV,
        scope_gas=config.SCOPE_GAS,
        label="impact plan JSON",
    )
    impact_plan = load_impact_plan(impact_plan_path)
    summary = impact_plan.get("summary") or {}
    scope_as = normalize_values(summary.get("scope_as") or [])
    scope_env = str(summary.get("scope_env") or "").strip()
    scope_gas = normalize_values(summary.get("scope_gas") or [])

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_xlsx = args.out_xlsx or build_artifact_path(config.ZABBIX_APPLY_PREFIX, scope_as, scope_env, scope_gas, ".xlsx", timestamp=timestamp)
    out_json = args.out_json or build_artifact_path(config.ZABBIX_APPLY_PREFIX, scope_as, scope_env, scope_gas, ".json", timestamp=timestamp)

    dry_run = not bool(config.ZABBIX_APPLY_CHANGES)
    connection = config.load_zabbix_connection()
    api = ZabbixAPI(connection.api_url, timeout_sec=int(config.HTTP_TIMEOUT_SEC))
    api.authenticate(connection.username, connection.password, connection.api_token)

    if not str(config.SOURCE_IMPACT_PLAN_JSON or "").strip():
        print(f"Using latest impact plan JSON: {impact_plan_path}")

    if not dry_run:
        print("Checking Zabbix API permissions for host.massadd")
        _assert_host_massadd_available(api)
        backup_path = resolve_input_artifact(
            config.SOURCE_BACKUP_FILE,
            config.BACKUP_PREFIX,
            ".json.gz",
            scope_as=scope_as,
            scope_env=scope_env,
            scope_gas=scope_gas,
            label="backup file",
        )
        if not str(config.SOURCE_BACKUP_FILE or "").strip():
            print(f"Using latest backup file: {backup_path}")
        print(f"Validating backup against impact plan: {backup_path}")
        _validate_backup_scope(backup_path, impact_plan, api)

    print(f"Applying Zabbix plan from: {impact_plan_path}")
    print(f"Mode: {'DRY-RUN' if dry_run else 'APPLY'}")
    result = apply_host_enrichment(api, impact_plan, dry_run=dry_run)
    result.setdefault("meta", {})["impact_plan_path"] = impact_plan_path

    print(f"Writing Zabbix apply XLSX: {out_xlsx}")
    write_apply_xlsx(result, out_xlsx)
    print(f"Writing Zabbix apply JSON: {out_json}")
    with open(out_json, "w", encoding="utf-8") as handle:
        json.dump(result, handle, ensure_ascii=False, indent=2)
    print("Zabbix apply completed.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
