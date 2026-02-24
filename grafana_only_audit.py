#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""grafana_only_audit.py — запуск аудита только по Grafana на основе seed."""

from __future__ import annotations

import argparse
import json
import os
import re
from typing import Dict, Iterable, List, Optional

from openpyxl import Workbook  # type: ignore
from openpyxl.styles import Alignment, Font  # type: ignore
from openpyxl.utils import get_column_letter  # type: ignore

from config import CONFIG, load_grafana_from_module
from grafana_audit import collect_grafana_matches


def safe_sheet_title(raw: str, max_len: int = 31) -> str:
    s = str(raw)
    s = re.sub(r"[\[\]\*:/\\\?]", "_", s).strip()
    if not s:
        return "AS"
    if len(s) <= max_len:
        return s
    head = s[: max(1, max_len - 8)]
    return f"{head}_seed"[:max_len]


def autosize_columns(ws, min_width: int = 10, max_width: int = 70) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(min_width, min(max_width, max_len + 2))


def write_section_title(ws, row: int, title: str) -> int:
    c = ws.cell(row=row, column=1, value=title)
    c.font = Font(bold=True, size=14)
    c.alignment = Alignment(vertical="center")
    return row + 2


def load_seed(path: str) -> tuple[Dict[str, Dict[str, List[str]]], List[Dict[str, str]], List[Dict[str, str]]]:
    with open(path, "r", encoding="utf-8") as f:
        raw = json.load(f)
    as_data = raw.get("as") or {}
    out: Dict[str, Dict[str, List[str]]] = {}
    for as_val, v in as_data.items():
        out[str(as_val)] = {
            "groups_old": list(v.get("groups_old") or []),
            "groups_new": list(v.get("groups_new") or []),
        }
    mapping_rows = list(raw.get("mapping_rows") or [])
    unknown_rows = list(raw.get("unknown_hosts") or [])
    return out, mapping_rows, unknown_rows


def build_seed_path(output_xlsx: str) -> str:
    base = output_xlsx
    if base.lower().endswith(".xlsx"):
        base = base[:-5]
    return f"{base}_zbx_seed.json"


def write_grafana_workbook(
    rows_by_as: Dict[str, List[Dict[str, str]]],
    out_path: str,
    mapping_rows: List[Dict[str, str]],
    unknown_rows: List[Dict[str, str]],
) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    # UNKNOWN sheet
    ws_u = wb.create_sheet(CONFIG.excel.sheet_unknown)
    ws_u.append(["host_name", "hostid", CONFIG.tags.AS, CONFIG.tags.ASN, "groups", "tags_json"])
    for r in unknown_rows or []:
        ws_u.append(
            [
                r.get("host_name"),
                r.get("hostid"),
                r.get("AS"),
                r.get("ASN"),
                r.get("groups"),
                r.get("tags_json"),
            ]
        )
    autosize_columns(ws_u)

    # MAPPING sheet
    ws_m = wb.create_sheet(CONFIG.excel.sheet_mapping)
    ws_m.append(
        [
            "AS",
            "old_group",
            "new_group",
            "jaccard",
            "precision",
            "intersection",
            "hosts_in_new",
            "hosts_in_old",
            "old_top1_conflict",
            "ambiguous_old_to_many_new",
            "other_new_top1",
            "env_new_top",
            "env_old_top",
            "env_new_multi",
            "env_old_multi",
            "env_mismatch",
        ]
    )
    for r in mapping_rows or []:
        ws_m.append(
            [
                r.get("AS"),
                r.get("old_group"),
                r.get("new_group"),
                r.get("jaccard"),
                r.get("precision"),
                r.get("intersection"),
                r.get("hosts_in_new"),
                r.get("hosts_in_old"),
                r.get("old_top1_conflict"),
                r.get("ambiguous_old_to_many_new"),
                r.get("other_new_top1"),
                r.get("env_new_top"),
                r.get("env_old_top"),
                r.get("env_new_multi"),
                r.get("env_old_multi"),
                r.get("env_mismatch"),
            ]
        )
    autosize_columns(ws_m)

    # Summary sheet
    ws_sum = wb.create_sheet(getattr(CONFIG.excel, "sheet_grafana", "GRAFANA"))
    ws_sum.append(["AS", "dashboard_uid", "dashboard_title", "matched_string", "match_type", "count"])
    for as_val, rows in rows_by_as.items():
        for r in rows:
            ws_sum.append(
                [
                    as_val,
                    r.get("dashboard_uid"),
                    r.get("dashboard_title"),
                    r.get("matched_string"),
                    r.get("match_type"),
                    r.get("count"),
                ]
            )
    autosize_columns(ws_sum)

    # Per-AS sheets (в стиле основного отчёта)
    for as_val, rows in rows_by_as.items():
        ws = wb.create_sheet(safe_sheet_title(as_val, max_len=int(CONFIG.excel.sheet_name_max)))
        r = 1
        r = write_section_title(ws, r, f"AS = {as_val}")
        r = write_section_title(ws, r, "5) Grafana dashboards (поиск OLD/NEW host-groups)")
        ws.append(["dashboard_uid", "dashboard_title", "matched_string", "match_type", "count"])
        for row in rows:
            ws.append(
                [
                    row.get("dashboard_uid"),
                    row.get("dashboard_title"),
                    row.get("matched_string"),
                    row.get("match_type"),
                    row.get("count"),
                ]
            )
        autosize_columns(ws)

    wb.save(out_path)


def main() -> int:
    parser = argparse.ArgumentParser(description="Grafana-only audit (using Zabbix seed)")
    parser.add_argument("--seed", help="Seed JSON (from zbx_hg_mapping_audit)")
    parser.add_argument("--out", dest="output", help="XLSX output path")
    args = parser.parse_args()

    scope_as = list(CONFIG.runtime.audit_scope_as) if CONFIG.runtime.audit_scope_as else []

    seed_path = args.seed or CONFIG.runtime.zabbix_seed_path or build_seed_path(CONFIG.excel.output_xlsx)
    if not os.path.exists(seed_path):
        raise RuntimeError(f"Seed file not found: {seed_path}")

    seed_as, mapping_rows, unknown_rows = load_seed(seed_path)
    report_as = {k: {"groups_old": v["groups_old"], "groups_new": v["groups_new"]} for k, v in seed_as.items()}
    if scope_as:
        scope_lower = {s.strip().lower() for s in scope_as}
        report_as = {k: v for k, v in report_as.items() if k.strip().lower() in scope_lower}
        mapping_rows = [r for r in mapping_rows if str(r.get("AS") or "").strip().lower() in scope_lower]
        unknown_rows = []

    grafana_conn = load_grafana_from_module()
    rows_by_as = collect_grafana_matches(grafana_conn, report_as, scope_as=scope_as or None)

    out_path = args.output
    if not out_path:
        base_dir = os.path.dirname(CONFIG.excel.output_xlsx) or "."
        out_path = os.path.join(base_dir, "grafana_audit_only.xlsx")

    write_grafana_workbook(rows_by_as, out_path, mapping_rows, unknown_rows)
    print(f"Grafana audit saved: {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
